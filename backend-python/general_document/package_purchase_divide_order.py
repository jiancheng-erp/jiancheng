import shutil
from openpyxl import load_workbook
import os
from openpyxl.styles import Border, Side, Alignment
from openpyxl.drawing.image import Image
from logger import logger
# Function to load the Excel template and prepare for modification
def load_template(template_path, new_file_path):
    # Copy the template to a new file
    shutil.copy(template_path, new_file_path)
    # Load the new workbook
    wb = load_workbook(new_file_path)
    ws = wb.active
    return wb, ws
def copy_text_to_new_sheet(source_file, dest_wb, sheet_name="Package Info"):
    """Copy text from source file into a new worksheet in destination workbook, with error handling."""
    try:
        # Load source workbook
        source_wb = load_workbook(source_file, data_only=False)
        source_ws = source_wb.active

        # Create a new worksheet in the destination workbook
        if sheet_name in dest_wb.sheetnames:
            dest_wb.remove(dest_wb[sheet_name])  # Remove existing sheet if any
        dest_ws = dest_wb.create_sheet(sheet_name)

        max_col = source_ws.max_column
        max_row = source_ws.max_row

        # Copy cell values
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = source_ws.cell(row=row, column=col)
                dest_ws.cell(row=row, column=col, value=cell.value)

        logger.debug(f"Successfully copied text to new sheet: {sheet_name}")
        return source_ws, dest_ws  # Return both worksheets

    except Exception as e:
        logger.debug(f"⚠️ Error copying text to new sheet: {e}")
        return None, None  # Return None to indicate failure but continue execution

def copy_images_with_absolute_positioning(source_ws, dest_ws):
    """Copy images from source worksheet to destination worksheet while keeping absolute positions, with error handling."""
    try:
        if source_ws is None or dest_ws is None:
            logger.debug("⚠️ Skipping image copying due to missing source/destination worksheet.")
            return

        for img in source_ws._images:
            new_img = Image(img.ref)  # Copy the image

            # Check the type of anchor (oneCellAnchor or twoCellAnchor)
            if hasattr(img, "anchor"):
                new_img.anchor = img.anchor  # Copy exact position anchor
                dest_ws.add_image(new_img)  # Place image at the same position

        logger.debug("✅ Successfully copied images with absolute positioning.")

    except Exception as e:
        logger.debug(f"⚠️ Error copying images: {e}")
def add_borders(ws, start_cell, end_cell):
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws[start_cell:end_cell]:
        for cell in row:
            cell.border = border  # Apply border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # Center text

# Function to insert series data starting from row 4
def merge_cells(ws, row):
    ws.merge_cells(f"A{row+1}:D{row+1}")
    ws.merge_cells(f"B{row+2}:H{row+2}")
    ws.merge_cells(f"B{row+3}:H{row+3}")
    ws.merge_cells(f"B{row+4}:D{row+4}")
    ws.merge_cells(f"E{row+4}:H{row+4}")
    ws.merge_cells(f"A3:A{row}")
    add_borders(ws, f"A3", f"H{row+1}")
def insert_series_data(ws, series_data, start_row=4):
    required_rows = 5  # Minimum rows to keep
    row = start_row - 1  # To calculate the last row after loop

    for i, item in enumerate(series_data):
        row = start_row + i
        logger.debug(f"Inserting series data into row {row}")
        logger.debug(item)

        ws[f"B{row}"] = i + 1
        ws[f"C{row}"] = item.get("物品名称", "")
        ws[f"D{row}"] = item.get("单位", "")
        ws[f"E{row}"] = item.get("数量", "")
        ws[f"F{row}"] = item.get("单价", "")
        ws[f"G{row}"] = item.get("用途说明", "")
        ws[f"H{row}"] = item.get("备注", "")

        # Ensure alignment for each cell in the row
        for col in "BCDEFGH":
            ws[f"{col}{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i in range(len(series_data), required_rows):
        row = start_row + i
        logger.debug(f"Adding empty row at {row}")
        ws[f"B{row}"] = i + 1  # Continue numbering for empty rows

        # Ensure alignment for each empty row
        for col in "BCDEFGH":
            ws[f"{col}{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    return row

# Function to save the workbook after modification
def save_workbook(wb, new_file_path):
    wb.save(new_file_path)

# Main function to generate the Excel file
def generate_package_excel_file(template_path, new_file_path, package_info_file, order_data):
    logger.debug(f"Generating Excel file for order {order_data.get('订单信息', '')}")
    wb, ws = load_template(template_path, new_file_path)

    # Insert order details
    ws["D2"] = order_data.get("订单信息", "")
    ws["B2"] = order_data.get("供应商", "")
    ws["H2"] = order_data.get("日期", "")
    ws["H1"] = order_data.get("客户名", "")

    # Insert series data from row 4 onwards
    row = insert_series_data(ws, order_data.get("seriesData", []))

    # Summary Section
    ws[f"A{row+1}"] = "合计"
    total_sum = sum(ws[f"E{r}"].value for r in range(5, row, 2) if ws[f"E{r}"].value)
    ws[f"E{row+1}"] = total_sum
    ws[f"F{row+1}"] = order_data.get("备注", "")
    ws[f"A{row+2}"] = "环境要求:"
    ws[f"A{row+3}"] = "发货地址:"
    ws[f"A{row+4}"] = "交货期限:"
    ws[f"B{row+2}"] = order_data.get("环保要求", "")
    ws[f"B{row+3}"] = order_data.get("发货地址", "")
    ws[f"B{row+4}"] = order_data.get("交货期限", "")
    ws[f"E{row+4}"] = "如有特殊情况提前5天反馈，无故延期有贵公司承担后续责任。"
    ws[f"A{row+5}"] = "制表:"
    ws[f"E{row+5}"] = "审核:"

    # Ensure summary section is also centered
    for col in "ABCDEFGH":
        for r in range(row+1, row+6):
            ws[f"{col}{r}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Merge and apply borders
    merge_cells(ws, row)
    add_borders(ws, "A3", f"H{row+1}")
    source_ws, dest_ws = copy_text_to_new_sheet(package_info_file, wb, "Package Info")

    # Copy images into the new worksheet with absolute positioning
    copy_images_with_absolute_positioning(source_ws, dest_ws)

    # Save the workbook
    save_workbook(wb, new_file_path)

    logger.debug(f"Workbook saved as {new_file_path}")


# template_path = "H:/git-projects/jiancheng/backend-python/general_document/标准采购订单.xlsx"
# new_file_path = "H:/git-projects/jiancheng/backend-python/general_document/pur_test.xlsx"
# package_file = "H:/git-projects/jiancheng/backend-python/general_document/包装资料.xlsx"
# order_data = {
#     "订单信息": "订单编号12345",
#     "供应商": "供应商A",
#     "日期": "2024-11-19",
#     "seriesData": [
#         {"物品名称": "商品1", "单位": "个", "数量": 10, "单价": 5.5, "用途说明": "用途1", "备注": "备注1"},
#         {"物品名称": "商品2", "单位": "箱", "数量": 20, "单价": 15.0, "用途说明": "用途2", "备注": "备注2"},
#     ],
#     "合计": 350,
#     "备注": "总备注",
#     "环保要求": "符合环保标准",
#     "发货地址": "测试地址",
#     "交货期限": "2024-12-01",
# }
# generate_package_excel_file(template_path, new_file_path, package_file, order_data)
# logger.debug("Test Case 1: Basic functionality passed.")

