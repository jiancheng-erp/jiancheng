import os
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, column_index_from_string

from constants import SHOESIZERANGE
from file_locations import IMAGE_UPLOAD_PATH
from logger import logger

TABLE_START_COLUMN = 2
NOTES_BLOCK_WIDTH = 5
HEADER_FILL = PatternFill("solid", fgColor="E8EFF7")
THIN_BORDER = Border(
    left=Side(style="thin", color="9CA3AF"),
    right=Side(style="thin", color="9CA3AF"),
    top=Side(style="thin", color="9CA3AF"),
    bottom=Side(style="thin", color="9CA3AF"),
)
CAUTION_LINES = [
    "所有尺寸以实裁为准，严控色差与装配。",
    "如遇生产异常请及时反馈生产管理部。",
]
CRAFT_TALL_LABELS = {"针车要求", "成型要求", "后整理"}


def generate_production_instruction_order_excel_file(
    template_path,
    new_file_path,
    order_data: Dict,
    metadata: Dict,
    instruction_payload: Optional[Dict] = None,
):
    """Build the instruction workbook without relying on a template file."""

    if not order_data:
        raise ValueError("order_data is required")

    workbook = Workbook()
    ws = workbook.active
    ws.title = "生产指令单"
    _configure_sheet(ws)

    size_names = _normalize_size_names(metadata.get("sizeNames"))
    payload = instruction_payload or {}
    current_row = 1

    for index, order_entry in enumerate(order_data.values()):
        if index > 0:
            current_row = ws.max_row + 3
        current_row = _render_instruction_block(
            ws, start_row=current_row, order_info=order_entry, size_names=size_names, payload=payload
        )

    workbook.save(new_file_path)
    logger.debug("生产指令单生成完成: %s", new_file_path)


def _configure_sheet(ws):
    column_widths = {
        "A": 14,
        "B": 16,
        "C": 14,
        "D": 10,
        "E": 16,
        "F": 12,
        "G": 12,
    }
    for offset in range(len(SHOESIZERANGE) + 6):
        letter = get_column_letter(TABLE_START_COLUMN + offset)
        column_widths.setdefault(letter, 8.5)
    for letter, width in column_widths.items():
        ws.column_dimensions[letter].width = width
    ws.sheet_view.showGridLines = False
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def _normalize_size_names(size_names: Optional[List[str]]) -> List[str]:
    fallback = [str(size) for size in SHOESIZERANGE]
    normalized: List[str] = []
    incoming = size_names or []
    for idx in range(len(fallback)):
        label = incoming[idx] if idx < len(incoming) else None
        text = str(label).strip() if label else ""
        normalized.append(text or fallback[idx])
    return normalized


def _render_instruction_block(ws, start_row: int, order_info: Dict, size_names: List[str], payload: Dict) -> int:
    title_row = start_row
    _render_title(ws, title_row, order_info, len(size_names))
    spacer_row = title_row + 1
    ws.row_dimensions[spacer_row].height = max(ws.row_dimensions[spacer_row].height or 0, 6)
    header_row = spacer_row + 1
    table_meta = _setup_instruction_table_headers(ws, header_row, size_names)
    data_start_row = header_row + 1
    last_data_row = _insert_instruction_rows(ws, data_start_row, order_info, table_meta, payload)
    if last_data_row < data_start_row:
        last_data_row = data_start_row
    notes_start_row = last_data_row + 1
    notes_end_row = _render_instruction_notes(ws, payload, order_info, notes_start_row, table_meta)
    caution_row = _merge_lower_left_block(ws, last_data_row, notes_end_row, table_meta)
    final_row = max(last_data_row, notes_end_row, caution_row or 0)
    _render_date_column(ws, header_row, final_row, payload, order_info)
    _apply_table_borders(ws, header_row, final_row, table_meta["last_col"], 1)
    return final_row


def _render_title(ws, row: int, order_info: Dict, size_count: int):
    end_column = TABLE_START_COLUMN + size_count + 6
    ws.merge_cells(start_row=row, start_column=TABLE_START_COLUMN, end_row=row, end_column=end_column)
    cell = ws.cell(row=row, column=TABLE_START_COLUMN)
    cell.value = order_info.get("title") or "生产指令单"
    cell.font = Font(size=18, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 32


def _derive_customer_name(order_info: Dict) -> Optional[str]:
    explicit = order_info.get("customerName")
    if explicit:
        return explicit
    title = order_info.get("title")
    if title:
        core = title
        if "健诚集团" in core:
            core = core.split("健诚集团", 1)[1]
        if "生产指令单" in core:
            core = core.replace("生产指令单", "")
        core = core.strip()
        if core:
            return core
    return order_info.get("customerProductName")


def _setup_instruction_table_headers(ws, row: int, size_names: List[str]) -> Dict:
    headers: List[Tuple[str, str]] = [
        ("customer_name", "客户名称"),
        ("shoe_model", "鞋型编号"),
        ("flow", "流程号"),
        ("customer", "客户款号"),
        ("color", "颜色"),
    ]
    column_map = {}
    column_index = TABLE_START_COLUMN
    for key, label in headers:
        column_map[f"{key}_col"] = column_index
        ws.cell(row=row, column=column_index, value=label)
        _style_header_cell(ws.cell(row=row, column=column_index))
        column_index += 1

    column_map["size_start_col"] = column_index
    for name in size_names:
        ws.cell(row=row, column=column_index, value=name)
        _style_header_cell(ws.cell(row=row, column=column_index))
        column_index += 1
    column_map["size_end_col"] = column_index - 1

    footer_headers: List[Tuple[str, str]] = [
        ("total_col", "合计"),
        ("remark_col", "备注"),
        ("last_type_col", "楦头"),
        ("outsole_col", "大底"),
        ("needle_status_col", "针车状态"),
        ("line_col", "线别"),
        ("order_no_col", "订单号"),
    ]
    for key, label in footer_headers:
        column_map[key] = column_index
        ws.cell(row=row, column=column_index, value=label)
        _style_header_cell(ws.cell(row=row, column=column_index))
        column_index += 1

    column_map["start_col"] = TABLE_START_COLUMN
    column_map["last_col"] = column_index - 1
    column_map["size_names"] = size_names
    return column_map


def _style_header_cell(cell):
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = HEADER_FILL
    cell.border = THIN_BORDER


def _style_body_cell(cell, wrap: bool = False):
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    cell.border = THIN_BORDER


def _insert_instruction_rows(ws, start_row: int, order_info: Dict, table_meta: Dict, payload: Dict) -> int:
    row = start_row
    shoes = order_info.get("shoes") or []
    size_pairs = list(zip(SHOESIZERANGE, table_meta["size_names"]))
    customer_name_value = _derive_customer_name(order_info)
    shoe_model_value = order_info.get("shoeRId")
    last_value = _extract_last_material(payload)
    outsole_value = _extract_outsole_material(payload)

    if not shoes:
        cell = ws.cell(row=row, column=table_meta["flow_col"])
        cell.value = "暂无数据"
        _style_body_cell(cell)
        return row

    for shoe_entry in shoes:
        packaging_entries = shoe_entry.get("packagingInfo") or [{}]
        color_row_start = row
        for packaging in packaging_entries:
            ws.row_dimensions[row].height = 22
            _write_packaging_row(ws, row, order_info, packaging, table_meta, size_pairs)
            flow_cell = ws.cell(row=row, column=table_meta["flow_col"])
            if flow_cell.value is None:
                flow_cell.value = ""
            _style_body_cell(flow_cell)
            for aux_col in ("needle_status_col", "line_col", "order_no_col"):
                column_index = table_meta.get(aux_col)
                if not column_index:
                    continue
                aux_cell = ws.cell(row=row, column=column_index)
                if aux_cell.value is None:
                    aux_cell.value = ""
                _style_body_cell(aux_cell)
            row += 1
        _merge_block_columns(
            ws,
            color_row_start,
            row - 1,
            [
                (table_meta["customer_name_col"], customer_name_value),
                (table_meta["shoe_model_col"], shoe_model_value),
                (table_meta["customer_col"], order_info.get("customerProductName")),
                (table_meta["color_col"], shoe_entry.get("colorName") or shoe_entry.get("color")),
                (table_meta["last_type_col"], last_value, ""),
                (table_meta["outsole_col"], outsole_value, ""),
            ],
        )
    return row - 1


def _write_packaging_row(ws, row: int, order_info: Dict, packaging: Dict, table_meta: Dict, size_pairs: List[Tuple[int, str]]):
    size_total = 0
    for idx, (size_value, _) in enumerate(size_pairs):
        column = table_meta["size_start_col"] + idx
        amount = _extract_size_amount(packaging, size_value)
        cell = ws.cell(row=row, column=column, value=amount if amount else None)
        _style_body_cell(cell)
        if isinstance(amount, (int, float)):
            size_total += amount

    total_amount = _coerce_number(packaging.get("totalQuantityRatio"))
    if total_amount is None:
        total_amount = _coerce_number(packaging.get("totalProductionAmount"))
    if total_amount is None:
        total_amount = size_total

    remark_text = packaging.get("remark") or order_info.get("remark")

    ws.cell(row=row, column=table_meta["total_col"], value=total_amount if total_amount is not None else "-")
    remark_cell = ws.cell(row=row, column=table_meta["remark_col"], value=remark_text or "")

    _style_body_cell(ws.cell(row=row, column=table_meta["total_col"]))
    _style_body_cell(remark_cell, wrap=True)


def _extract_size_amount(packaging: Dict, shoe_size: int):
    size_map = packaging.get("sizeProductionAmounts")
    if isinstance(size_map, dict):
        return _coerce_number(size_map.get(str(shoe_size)) or size_map.get(shoe_size))
    ratio_key = f"size{shoe_size}Ratio"
    amount_key = f"size{shoe_size}Amount"
    return _coerce_number(packaging.get(ratio_key) or packaging.get(amount_key))


def _merge_block_columns(ws, start_row: int, end_row: int, columns: Iterable[Tuple[int, Optional[str], Optional[str]]]):
    if end_row <= start_row:
        for entry in columns:
            column, value, placeholder = _normalize_column_entry(entry)
            cell = ws.cell(row=start_row, column=column, value=_resolve_placeholder(value, placeholder))
            _style_body_cell(cell, wrap=True)
        return

    for entry in columns:
        column, value, placeholder = _normalize_column_entry(entry)
        ws.merge_cells(start_row=start_row, start_column=column, end_row=end_row, end_column=column)
        cell = ws.cell(row=start_row, column=column)
        cell.value = _resolve_placeholder(value, placeholder)
        _style_body_cell(cell, wrap=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _merge_lower_left_block(ws, last_data_row: int, notes_end_row: int, table_meta: Dict) -> Optional[int]:
    start_row = last_data_row + 1
    end_row = max(notes_end_row, start_row)
    notes_start_col = max(table_meta["last_col"] - NOTES_BLOCK_WIDTH + 1, TABLE_START_COLUMN)
    max_left_col = column_index_from_string("U")
    start_col = 2
    end_col = min(max_left_col, notes_start_col - 1)

    if end_col >= start_col and end_row >= start_row:
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
        cell = ws.cell(row=start_row, column=start_col)
        cell.value = ""
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        base_row = end_row
    else:
        base_row = start_row - 1
        end_col = max_left_col

    caution_row = base_row + 1
    if caution_row <= base_row:
        return base_row
    ws.merge_cells(start_row=caution_row, start_column=start_col, end_row=caution_row, end_column=end_col)
    caution_cell = ws.cell(row=caution_row, column=start_col)
    caution_cell.value = "  ".join(CAUTION_LINES)
    caution_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    caution_cell.border = THIN_BORDER
    return caution_row


def _normalize_column_entry(entry: Tuple[int, Optional[str], Optional[str]]):
    if len(entry) == 3:
        column, value, placeholder = entry
    else:
        column, value = entry  # type: ignore[misc]
        placeholder = "-"
    return column, value, placeholder


def _resolve_placeholder(value: Optional[str], placeholder: Optional[str]) -> str:
    if value not in (None, ""):
        return str(value)
    if placeholder is None:
        return ""
    return placeholder


def _render_instruction_notes(ws, payload: Dict, order_info: Dict, start_row: int, table_meta: Dict) -> int:
    start_col = max(table_meta["last_col"] - NOTES_BLOCK_WIDTH + 1, TABLE_START_COLUMN)
    end_col = table_meta["last_col"]
    row = start_row

    sections: List[Tuple[str, List[Tuple[str, str]]]] = []

    spec_lines: List[Tuple[str, str]] = []
    printing_value = payload.get("printingProcess")
    if printing_value:
        spec_lines.append(("印刷工艺", printing_value))
    if spec_lines:
        sections.append(("包装/印刷", spec_lines))

    material_lines = _summarize_materials(payload.get("materials"))
    if material_lines:
        sections.append(("材料说明", material_lines))

    craft_lines = _summarize_craft(payload.get("craftRequirements"))
    if craft_lines:
        sections.append(("工艺要求", craft_lines))

    remark = order_info.get("remark")
    if remark:
        sections.append(("订单备注", [("备注", remark)]))

    for title, lines in sections:
        row = _write_notes_section(ws, row, start_col, end_col, title, lines)

    row = _maybe_attach_image(ws, row, start_col, end_col, order_info)

    return max(row - 1, start_row)


def _render_date_column(ws, header_row: int, last_row: int, payload: Dict, order_info: Dict):
    if last_row < header_row:
        return
    date_value = payload.get("instructionDate") or order_info.get("orderStartDate") or order_info.get("orderEndDate")
    if not date_value:
        return
    ws.merge_cells(start_row=header_row, start_column=1, end_row=last_row, end_column=1)
    cell = ws.cell(row=header_row, column=1)
    cell.value = f"排单日期\n{date_value}"
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.font = Font(bold=True)
    for r in range(header_row, last_row + 1):
        ws.cell(row=r, column=1).border = THIN_BORDER


def _maybe_attach_image(ws, row: int, start_col: int, end_col: int, order_info: Dict) -> int:
    shoes = order_info.get("shoes") or []
    image_url = shoes[0].get("imgUrl") if shoes else None
    image_path = _resolve_image_path(image_url)
    if not image_path:
        return row
    try:
        img = Image(image_path)
        img.width = 120
        img.height = 120
        anchor_col = max(start_col, end_col - 1)
        anchor = ws.cell(row=row, column=anchor_col).coordinate
        ws.add_image(img, anchor)
        for extra_row in range(row, row + 6):
            ws.row_dimensions[extra_row].height = max(ws.row_dimensions[extra_row].height or 0, 22)
        return row + 6
    except Exception as exc:  # pragma: no cover - defensive guard
        logger.warning("指令单图片插入失败: %s", exc)
        return row


def _write_notes_section(ws, row: int, start_col: int, end_col: int, title: str, lines: List[Tuple[str, str]]) -> int:
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    title_cell = ws.cell(row=row, column=start_col, value=title)
    title_cell.font = Font(bold=True)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    row += 1
    for label, value in lines:
        if not (label or value):
            continue
        label_cell = ws.cell(row=row, column=start_col, value=label or "")
        _style_body_cell(label_cell)
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        if start_col + 1 <= end_col:
            ws.merge_cells(start_row=row, start_column=start_col + 1, end_row=row, end_column=end_col)
            body_cell = ws.cell(row=row, column=start_col + 1, value=value or "")
        else:
            body_cell = ws.cell(row=row, column=start_col, value=value or "")
        _style_body_cell(body_cell, wrap=True)
        body_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        if label in CRAFT_TALL_LABELS:
            ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 0, 120)
        row += 1
    return row


def _resolve_image_path(image_url: Optional[str]) -> Optional[str]:
    if not image_url:
        return None
    if os.path.isabs(image_url) and os.path.exists(image_url):
        return image_url
    candidate = os.path.join(IMAGE_UPLOAD_PATH, image_url)
    return candidate if os.path.exists(candidate) else None


def _summarize_materials(materials: Optional[Dict]) -> List[Tuple[str, str]]:
    lines: List[Tuple[str, str]] = []
    if materials and materials.get("midsole"):
        text = materials["midsole"].get("displayText") or _join_material_parts(materials["midsole"])
        if text:
            lines.append(("中底", text))
    if materials and materials.get("hotsole"):
        text = materials["hotsole"].get("displayText") or _join_material_parts(materials["hotsole"])
        if text:
            lines.append(("烫底", text))
    lines.append(("鞋盒规格", ""))
    lines.append(("外箱规格", ""))
    return lines


def _extract_last_material(payload: Dict) -> Optional[str]:
    materials = (payload or {}).get("materials") or {}
    value = materials.get("lastType")
    return value if value else None


def _extract_outsole_material(payload: Dict) -> Optional[str]:
    materials = (payload or {}).get("materials") or {}
    outsole = materials.get("outsole") or {}
    if not outsole:
        return None
    return outsole.get("displayText") or _join_material_parts(outsole)


def _join_material_parts(entry: Dict) -> str:
    parts = [entry.get("supplierName"), entry.get("materialModel"), entry.get("materialSpecification")]
    return " ".join(filter(None, parts))


def _summarize_craft(craft: Optional[Dict]) -> List[Tuple[str, str]]:
    if not craft:
        return []
    mapping = {
        "productionRequirement": "生产要求",
        "cuttingRequirement": "裁断要求",
        "sewingRequirement": "针车要求",
        "moldingRequirement": "成型要求",
        "postProcessing": "后整理",
    }
    lines: List[Tuple[str, str]] = []
    for key, label in mapping.items():
        value = craft.get(key)
        if value:
            lines.append((label, value))
    return lines


def _coerce_number(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return value
    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return float(value)
        except (TypeError, ValueError):
            return None


def _apply_table_borders(ws, header_row: int, last_row: int, last_col: int, start_col: int):
    for row in range(header_row, last_row + 1):
        for column in range(start_col, last_col + 1):
            cell = ws.cell(row=row, column=column)
            if not cell.border or cell.border != THIN_BORDER:
                cell.border = THIN_BORDER
