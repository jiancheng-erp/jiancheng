# -*- coding: utf-8 -*-
# 总经理订单汇总导出（按工厂型号逐行）
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

THIN = Side(style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", fgColor="F2F2F2")
HEAD_FONT = Font(bold=True)
TITLE_FONT = Font(size=16, bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

HEADER = [
    "序号",
    "订单时间",
    "编号",
    "客户",
    "部门",
    "数量（双）",
    "金额",
    "币种",
    "订单交货期",
    "客人型号",
    "工厂型号",
]
CENTER_COLS = {1, 2, 6, 7, 8, 9}


def _fmt_date(d) -> str:
    return d.strftime("%Y.%m.%d") if d else ""


def _now_tag() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _text_width(val) -> int:
    if val is None:
        return 0
    w = 0
    for ch in str(val):
        w += 2 if ord(ch) > 255 else 1
    return max(w, 1)


def _filters_summary(f: dict) -> str:
    f = f or {}
    parts = []
    if f.get("start_date_from") or f.get("start_date_to"):
        parts.append(
            f"订单时间: {f.get('start_date_from') or '-'} ~ {f.get('start_date_to') or '-'}"
        )
    if f.get("end_date_from") or f.get("end_date_to"):
        parts.append(
            f"交货期: {f.get('end_date_from') or '-'} ~ {f.get('end_date_to') or '-'}"
        )
    if f.get("customer_name"):
        parts.append(f"客户: {f['customer_name']}")
    if f.get("department_name"):
        parts.append(f"部门: {f['department_name']}")
    if f.get("order_rid"):
        parts.append(f"编号: {f['order_rid']}")
    if f.get("customer_product_name"):
        parts.append(f"客人型号: {f['customer_product_name']}")
    if f.get("shoe_rid"):
        parts.append(f"工厂型号: {f['shoe_rid']}")
    return " | ".join(parts) if parts else "（无筛选条件）"


def build_order_summary_excel(rows: list[dict], filters: dict):
    """rows: 每行一个工厂型号，字段见 head_manager_api 组装。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "订单汇总"
    header_len = len(HEADER)

    ws.cell(row=1, column=1, value="订单汇总")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=header_len)
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = CENTER

    ws.cell(row=2, column=1, value=f"导出时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=header_len)
    ws.cell(row=2, column=1).alignment = RIGHT

    ws.cell(row=3, column=1, value=f"筛选条件：{_filters_summary(filters)}")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=header_len)
    ws.cell(row=3, column=1).alignment = LEFT

    header_row = 5
    widths = [0] * header_len
    for col_idx, val in enumerate(HEADER, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=val)
        cell.font = HEAD_FONT
        cell.alignment = CENTER
        cell.fill = HEAD_FILL
        cell.border = BORDER
        widths[col_idx - 1] = max(widths[col_idx - 1], _text_width(val))

    data_start = header_row + 1
    for idx, row in enumerate(rows, start=1):
        amount = row.get("amount")
        values = [
            idx,
            _fmt_date(row.get("start_date")),
            row.get("order_rid") or "",
            row.get("customer_name") or "",
            row.get("department_name") or "",
            row.get("total_pairs") or 0,
            float(amount) if amount is not None else "",
            row.get("currency") or "",
            _fmt_date(row.get("end_date")),
            row.get("customer_product_name") or "",
            row.get("shoe_rid") or "",
        ]
        r = data_start + idx - 1
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.border = BORDER
            if col_idx in CENTER_COLS:
                cell.alignment = CENTER
            widths[col_idx - 1] = max(widths[col_idx - 1], _text_width(val))

    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = min(
            max(8, w * 0.9 + 2), 48
        )
    ws.freeze_panes = ws[f"A{data_start}"]

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"订单汇总_{_now_tag()}.xlsx"
    return bio, filename
