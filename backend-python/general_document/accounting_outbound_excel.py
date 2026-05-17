"""
财务部 - 出库明细 Excel 生成
==================================
不依赖模板文件，使用 openpyxl 从零构建。列顺序与 OutboundDetail 页面
`name_en_cn_mapping_outbound` 保持一致。
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# (attr_camel_case, 中文表头, 列宽)
OUTBOUND_COLUMNS = [
    ("outboundRid",            "出库单据号", 18),
    ("outboundDatetime",       "出库时间",   20),
    ("outboundType",           "出库类型",   14),
    ("outboundDepartment",     "接收部门",   14),
    ("picker",                 "接收人",     12),
    ("supplierName",           "供应商",     18),
    ("materialName",           "材料名称",   18),
    ("materialModel",          "材料型号",   18),
    ("materialSpecification",  "材料规格",   20),
    ("color",                  "材料颜色",   12),
    ("unitPrice",              "库存单价",   12),
    ("actualInboundUnit",      "单位",       8),
    ("outboundAmount",         "出库数量",   12),
    ("itemTotalPrice",         "总价",       12),
    ("approvalStatus",         "审批状态",   12),
]


def generate_accounting_outbound_excel(
    save_path,
    warehouse_name,
    supplier_name,
    material_model,
    time_range,
    outbound_records,
):
    """构建出库总单 Excel 并保存到 save_path。"""
    wb = Workbook()
    sheet = wb.active
    sheet.title = "出库明细"

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    title_font = Font(name="微软雅黑", size=14, bold=True)
    meta_font = Font(name="微软雅黑", size=10)
    header_font = Font(name="微软雅黑", size=11, bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    n_cols = len(OUTBOUND_COLUMNS)
    last_col_letter = get_column_letter(n_cols)

    # 行 1: 标题
    sheet.cell(row=1, column=1, value="财务部 - 总仓出库明细").font = title_font
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    sheet.cell(row=1, column=1).alignment = center
    sheet.row_dimensions[1].height = 26

    # 行 2: 筛选条件
    meta_pairs = [
        ("仓库", warehouse_name or "全部"),
        ("供应商", supplier_name or "全部"),
        ("材料型号", material_model or "全部"),
        ("时间范围", time_range or "全部"),
    ]
    meta_text = "    ".join(f"{k}: {v}" for k, v in meta_pairs)
    cell = sheet.cell(row=2, column=1, value=meta_text)
    cell.font = meta_font
    cell.alignment = left
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    sheet.row_dimensions[2].height = 20

    # 行 3: 列标题
    header_row = 3
    for idx, (_attr, label, _w) in enumerate(OUTBOUND_COLUMNS, start=1):
        c = sheet.cell(row=header_row, column=idx, value=label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    # 列宽
    for idx, (_attr, _label, width) in enumerate(OUTBOUND_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width

    # 数据行
    start_row = header_row + 1
    for r_offset, data in enumerate(outbound_records or [], start=0):
        row = start_row + r_offset
        for c_idx, (attr, _label, _w) in enumerate(OUTBOUND_COLUMNS, start=1):
            val = data.get(attr, "")
            if val is None:
                val = ""
            cell = sheet.cell(row=row, column=c_idx, value=val)
            cell.border = border
            cell.alignment = left

    # 冻结表头
    sheet.freeze_panes = sheet.cell(row=start_row, column=1).coordinate

    # 自动筛选
    if outbound_records:
        last_row = start_row + len(outbound_records) - 1
        sheet.auto_filter.ref = f"A{header_row}:{last_col_letter}{last_row}"
    else:
        sheet.auto_filter.ref = f"A{header_row}:{last_col_letter}{header_row}"

    wb.save(save_path)
    return save_path
