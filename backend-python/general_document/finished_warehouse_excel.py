# services/finished_exports.py
# -*- coding: utf-8 -*-
from datetime import datetime, timedelta
from decimal import Decimal
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import desc, func

# === 按你的项目结构调整这些 import ===
from app_config import db
from models import (
    AccountingUnitConversionTable,
    Order,
    OrderShoe,
    OrderShoeType,
    Shoe,
    ShoeType,
    Color,
    Customer,
    FinishedShoeStorage,
    ShoeInboundRecord,
    ShoeInboundRecordDetail,
    ShoeOutboundRecord,
    ShoeOutboundRecordDetail,
)

# ================= 工具函数（通用） =================


def _parse_dt(s: str | None, end=False):
    if not s:
        return None
    s = s.strip()
    for f in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(s, f)
            if f == "%Y-%m-%d" and end:
                dt = dt + timedelta(days=1) - timedelta(seconds=1)
            return dt
        except ValueError:
            continue
    return None


def _format_dt(dt: datetime | None):
    return "" if not dt else dt.strftime("%Y-%m-%d %H:%M:%S")


def _safe_ilike(col, val):
    return col.ilike(f"%{val}%")


def _now_tag():
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _apply_common_filters(q, f):
    if f.get("order_rid"):
        q = q.filter(_safe_ilike(Order.order_rid, f["order_rid"]))
    if f.get("shoe_rid"):
        q = q.filter(_safe_ilike(Shoe.shoe_rid, f["shoe_rid"]))
    if f.get("customer_name"):
        q = q.filter(_safe_ilike(Customer.customer_name, f["customer_name"]))
    if f.get("customer_product_name"):
        q = q.filter(
            _safe_ilike(OrderShoe.customer_product_name, f["customer_product_name"])
        )
    if f.get("order_cid"):
        q = q.filter(_safe_ilike(Order.order_cid, f["order_cid"]))
    if f.get("customer_brand"):
        q = q.filter(_safe_ilike(Customer.customer_brand, f["customer_brand"]))
    return q


def _filters_summary(filters: dict) -> str:
    filters = filters or {}
    parts = []
    if filters.get("start_date") or filters.get("end_date"):
        parts.append(
            f"时间: {filters.get('start_date') or '-'} ~ {filters.get('end_date') or '-'}"
        )
    if filters.get("inbound_rid"):
        parts.append(f"入库单号: {filters['inbound_rid']}")
    if filters.get("outbound_rid"):
        parts.append(f"出库单号: {filters['outbound_rid']}")
    if filters.get("order_rid"):
        parts.append(f"订单号: {filters['order_rid']}")
    if filters.get("order_cid"):
        parts.append(f"客户订单号: {filters['order_cid']}")
    shoe_rid_val = filters.get("shoe_rid") or filters.get("shoeRid")
    if shoe_rid_val:
        parts.append(f"工厂型号: {shoe_rid_val}")
    if filters.get("customer_name"):
        parts.append(f"客户名称: {filters['customer_name']}")
    if filters.get("customer_brand"):
        parts.append(f"客户商标: {filters['customer_brand']}")
    if filters.get("customer_product_name"):
        parts.append(f"客户鞋型: {filters['customer_product_name']}")

    mode_val = (filters.get("mode") or "").lower()
    if mode_val == "month" and filters.get("month"):
        parts.append(f"月份: {filters['month']}")
    elif mode_val == "year" and filters.get("year"):
        parts.append(f"年份: {filters['year']}")

    direction = (filters.get("direction") or filters.get("Direction") or "").upper()
    if direction == "IN":
        parts.append("方向: 仅入库")
    elif direction == "OUT":
        parts.append("方向: 仅出库")

    keyword = filters.get("keyword")
    if keyword:
        parts.append(f"业务单号: {keyword}")

    color_kw = filters.get("color")
    if color_kw:
        parts.append(f"颜色: {color_kw}")

    category_kw = filters.get("category")
    if category_kw:
        parts.append(f"类别: {category_kw}")

    group_by = (filters.get("groupBy") or "model").lower()
    parts.append("分组: 型号+颜色" if group_by == "model_color" else "分组: 型号")

    return " | ".join(parts) if parts else "（无筛选条件）"


# —— 列宽估算（中英文混排，中文当作双倍宽度） ——
def _text_width(val) -> int:
    if val is None:
        return 0
    s = str(val)
    w = 0
    for ch in s:
        w += 2 if ord(ch) > 255 else 1
    return max(w, 1)


def _update_widths(widths: list[int], row_values: list):
    for i, v in enumerate(row_values):
        w = _text_width(v)
        if i >= len(widths):
            widths.append(w)
        else:
            widths[i] = max(widths[i], w)


def _apply_widths(
    ws, widths: list[int], padding: int = 2, min_w: int = 8, max_w: int = 48
):
    for idx, w in enumerate(widths, start=1):
        col_letter = get_column_letter(idx)
        # 大致换算：字符宽度 * 0.9，再加 padding
        width = min(max(min_w, (w * 0.9) + padding), max_w)
        ws.column_dimensions[col_letter].width = width


# —— 样式 ——
THIN = Side(style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", fgColor="F2F2F2")
HEAD_FONT = Font(bold=True)
TITLE_FONT = Font(size=16, bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def _write_title_filters(ws, title_text: str, header_len: int, filters: dict) -> int:
    """
    写：标题（合并、加粗居中） / 导出时间（右对齐，合并） / 筛选条件（左对齐，合并） / 空行
    返回下一行行号（即表头要写入的行号）
    """
    # 行1：大标题（合并到 header_len 列）
    ws.cell(row=1, column=1, value=title_text)
    if header_len > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=header_len)
    c1 = ws.cell(row=1, column=1)
    c1.font = TITLE_FONT
    c1.alignment = CENTER

    # 行2：导出时间（右对齐，合并整行）
    ws.cell(row=2, column=1, value=f"导出时间：{_format_dt(datetime.now())}")
    if header_len > 1:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=header_len)
    ws.cell(row=2, column=1).alignment = RIGHT

    # 行3：筛选条件（左对齐，合并整行）
    ws.cell(row=3, column=1, value=f"筛选条件：{_filters_summary(filters)}")
    if header_len > 1:
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=header_len)
    ws.cell(row=3, column=1).alignment = LEFT

    # 行4：空行
    return 5  # 表头从第5行开始写


def _write_header(ws, row_idx: int, header: list[str], widths: list[int]):
    for col_idx, val in enumerate(header, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = HEAD_FONT
        cell.alignment = CENTER
        cell.fill = HEAD_FILL
        cell.border = BORDER
    _update_widths(widths, header)


def _write_data_row(
    ws,
    row_idx: int,
    values: list,
    widths: list[int],
    center_cols: set[int] | None = None,
):
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border = BORDER
        if center_cols and col_idx in center_cols:
            cell.alignment = CENTER
    _update_widths(widths, values)


# ================== 导出：成品入库（仅自产） ==================


def build_finished_inbound_excel(filters: dict):
    start_dt = _parse_dt(filters.get("start_date"), end=False)
    end_dt = _parse_dt(filters.get("end_date"), end=True)

    # === 新增：汇总所有时间内的入库总量（不受页面时间筛选影响） ===
    inbound_sum_sq = (
        db.session.query(
            ShoeInboundRecordDetail.finished_shoe_storage_id.label("fsid"),
            func.sum(ShoeInboundRecordDetail.inbound_amount).label("total_inbound"),
        )
        .join(
            ShoeInboundRecord,
            ShoeInboundRecord.shoe_inbound_record_id
            == ShoeInboundRecordDetail.shoe_inbound_record_id,
        )
        .filter(ShoeInboundRecord.transaction_type == 1)
        .filter(ShoeInboundRecord.inbound_type == 0)  # 仅自产
        .group_by(ShoeInboundRecordDetail.finished_shoe_storage_id)
        .subquery()
    )

    q = (
        db.session.query(
            Order,
            Shoe.shoe_rid,
            OrderShoe.customer_product_name,
            Color.color_name,
            Customer,
            FinishedShoeStorage,
            ShoeInboundRecord,
            ShoeInboundRecordDetail,
            func.coalesce(inbound_sum_sq.c.total_inbound, 0).label(
                "all_time_inbound"
            ),  # 新增：全时段入库总和
        )
        .join(OrderShoe, Order.order_id == OrderShoe.order_id)
        .join(Shoe, Shoe.shoe_id == OrderShoe.shoe_id)
        .join(Customer, Customer.customer_id == Order.customer_id)
        .join(OrderShoeType, OrderShoeType.order_shoe_id == OrderShoe.order_shoe_id)
        .join(ShoeType, ShoeType.shoe_type_id == OrderShoeType.shoe_type_id)
        .join(Color, Color.color_id == ShoeType.color_id)
        .join(
            FinishedShoeStorage,
            FinishedShoeStorage.order_shoe_type_id == OrderShoeType.order_shoe_type_id,
        )
        .join(
            ShoeInboundRecordDetail,
            ShoeInboundRecordDetail.finished_shoe_storage_id
            == FinishedShoeStorage.finished_shoe_id,
        )
        .join(
            ShoeInboundRecord,
            ShoeInboundRecord.shoe_inbound_record_id
            == ShoeInboundRecordDetail.shoe_inbound_record_id,
        )
        # 报表行本身仍按筛选范围过滤与排序
        .filter(ShoeInboundRecord.transaction_type == 1)
        .filter(ShoeInboundRecordDetail.is_deleted == 0)
        .filter(ShoeInboundRecord.inbound_type == 0)  # 仅自产
        # 左连接全时段汇总（不受筛选影响）
        .outerjoin(
            inbound_sum_sq,
            inbound_sum_sq.c.fsid == FinishedShoeStorage.finished_shoe_id,
        )
        .order_by(desc(ShoeInboundRecord.inbound_datetime))
        .distinct(ShoeInboundRecordDetail.record_detail_id)
    )

    if start_dt:
        q = q.filter(ShoeInboundRecord.inbound_datetime >= start_dt)
    if end_dt:
        q = q.filter(ShoeInboundRecord.inbound_datetime <= end_dt)
    if filters.get("inbound_rid"):
        q = q.filter(
            _safe_ilike(ShoeInboundRecord.shoe_inbound_rid, filters["inbound_rid"])
        )
    q = _apply_common_filters(q, filters)

    rows = q.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "成品入库记录"

    header = [
        "订单号",
        "客户订单号",
        "客户名",
        "客户商标",
        "客户型号",
        "工厂型号",
        "颜色",
        "入库单号",
        "入库时间",
        "入库数量",
        "订单总数量",
        "未入库数量",
        "备注",
    ]
    widths: list[int] = []
    header_row = _write_title_filters(ws, "成品入库记录", len(header), filters)
    _write_header(ws, header_row, header, widths)

    data_start_row = header_row + 1
    center_cols = {9, 10}  # 入库时间、入库数量 居中
    r = data_start_row

    # 修改：使用 all_time_inbound 计算未入库数量（全时段总和）
    for (
        order,
        shoe_rid_v,
        cust_prod_name,
        color_name,
        customer,
        storage,
        record,
        detail,
        all_time_inbound,
    ) in rows:
        remaining = (storage.finished_estimated_amount or 0) - (all_time_inbound or 0)
        _write_data_row(
            ws,
            r,
            [
                order.order_rid,
                order.order_cid,
                customer.customer_name,
                customer.customer_brand,
                cust_prod_name,
                shoe_rid_v,
                color_name,
                record.shoe_inbound_rid,
                _format_dt(record.inbound_datetime),
                detail.inbound_amount,
                storage.finished_estimated_amount,
                remaining,  # ✅ 这里换成全时段入库总和后的差值
                detail.remark or "",
            ],
            widths,
            center_cols=center_cols,
        )
        r += 1

    ws.freeze_panes = ws[f"A{data_start_row}"]
    _apply_widths(ws, widths)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"成品入库记录_自产_{_now_tag()}.xlsx"
    return bio, filename


# ================== 导出：成品出库（仅自产） ==================


def build_finished_outbound_excel(filters: dict):
    start_dt = _parse_dt(filters.get("start_date"), end=False)
    end_dt = _parse_dt(filters.get("end_date"), end=True)

    # === 新增：全时段出库汇总（不受页面时间筛选影响） ===
    outbound_sum_sq = (
        db.session.query(
            ShoeOutboundRecordDetail.finished_shoe_storage_id.label("fsid"),
            func.sum(ShoeOutboundRecordDetail.outbound_amount).label("total_outbound"),
        )
        .join(
            ShoeOutboundRecord,
            ShoeOutboundRecord.shoe_outbound_record_id
            == ShoeOutboundRecordDetail.shoe_outbound_record_id,
        )
        .filter(ShoeOutboundRecord.outbound_type == 0)  # 仅自产出库
        .group_by(ShoeOutboundRecordDetail.finished_shoe_storage_id)
        .subquery()
    )

    q = (
        db.session.query(
            Order,
            Shoe.shoe_rid,
            Color.color_name,
            Customer,
            FinishedShoeStorage,
            OrderShoe.customer_product_name,
            OrderShoeType.unit_price,
            OrderShoeType.currency_type,
            ShoeOutboundRecord,
            ShoeOutboundRecordDetail,
            func.coalesce(outbound_sum_sq.c.total_outbound, 0).label(
                "all_time_outbound"
            ),  # 新增字段
        )
        .join(Customer, Customer.customer_id == Order.customer_id)
        .join(OrderShoe, Order.order_id == OrderShoe.order_id)
        .join(Shoe, Shoe.shoe_id == OrderShoe.shoe_id)
        .join(OrderShoeType, OrderShoeType.order_shoe_id == OrderShoe.order_shoe_id)
        .join(ShoeType, ShoeType.shoe_type_id == OrderShoeType.shoe_type_id)
        .join(Color, Color.color_id == ShoeType.color_id)
        .join(
            FinishedShoeStorage,
            FinishedShoeStorage.order_shoe_type_id == OrderShoeType.order_shoe_type_id
        )
        .join(
            ShoeOutboundRecordDetail,
            ShoeOutboundRecordDetail.finished_shoe_storage_id
            == FinishedShoeStorage.finished_shoe_id,
        )
        .join(
            ShoeOutboundRecord,
            ShoeOutboundRecord.shoe_outbound_record_id
            == ShoeOutboundRecordDetail.shoe_outbound_record_id,
        )
        .filter(ShoeOutboundRecord.outbound_type == 0)  # 仅自产出库
        .outerjoin(
            outbound_sum_sq,
            outbound_sum_sq.c.fsid == FinishedShoeStorage.finished_shoe_id,
        )  # ⬅️ 左连接全时段汇总
        .order_by(desc(ShoeOutboundRecord.outbound_datetime))
        .distinct(ShoeOutboundRecordDetail.record_detail_id)
    )

    if start_dt:
        q = q.filter(ShoeOutboundRecord.outbound_datetime >= start_dt)
    if end_dt:
        q = q.filter(ShoeOutboundRecord.outbound_datetime <= end_dt)
    if filters.get("outbound_rid"):
        q = q.filter(
            _safe_ilike(ShoeOutboundRecord.shoe_outbound_rid, filters["outbound_rid"])
        )
    q = _apply_common_filters(q, filters)

    rows = q.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "成品出库记录"

    header = [
        "订单号",
        "客户订单号",
        "客户名",
        "客户商标",
        "客户型号",
        "工厂型号",
        "颜色",
        "出库单号",
        "出库时间",
        "出库数量",
        "订单总数量",
        "未出库数量",
        "单价",
        "发货金额",
        "币种",
        "汇率",
        "人民币金额",
        "备注",
    ]
    widths: list[int] = []
    header_row = _write_title_filters(ws, "成品出库记录", len(header), filters)
    _write_header(ws, header_row, header, widths)

    data_start_row = header_row + 1
    center_cols = {9, 10}  # 出库时间、出库数量
    r = data_start_row
    exchange_USD = db.session.query(
        AccountingUnitConversionTable
    ).filter(AccountingUnitConversionTable.unit_from == 1, AccountingUnitConversionTable.unit_to == 4).first()
    exchange_EUR = db.session.query(
        AccountingUnitConversionTable
    ).filter(AccountingUnitConversionTable.unit_from == 1, AccountingUnitConversionTable.unit_to == 2).first()
    exchange_USD_rate = Decimal(exchange_USD.rate) if exchange_USD else Decimal("0")
    exchange_EUR_rate = Decimal(exchange_EUR.rate) if exchange_EUR else Decimal("0")
    exchange_CNY_rate = Decimal("1.0")
    for (
        order,
        shoe_rid_v,
        color_name,
        customer,
        storage,
        cust_prod_name,
        unit_price,
        currency_type,
        record,
        detail,
        all_time_outbound,
    ) in rows:
        # 兼容两种字段名
        qty = getattr(detail, "outbound_amount", None) or getattr(
            detail, "amount", None
        )
        
        remaining = (storage.finished_estimated_amount or 0) - (all_time_outbound or 0)
        if currency_type in ["CNY", "人民币", "RMB"]:
            exchange_rate = exchange_CNY_rate
        elif currency_type in ["USD", "USA", "美元"]:
            exchange_rate = exchange_USD_rate
        elif currency_type in ["EUR", "欧元"]:
            exchange_rate = exchange_EUR_rate
        else:
            exchange_rate = Decimal("1.0")

        _write_data_row(
            ws,
            r,
            [
                order.order_rid,
                order.order_cid,
                customer.customer_name,
                customer.customer_brand,
                cust_prod_name,
                shoe_rid_v,
                color_name,
                record.shoe_outbound_rid,
                _format_dt(record.outbound_datetime),
                qty,
                storage.finished_estimated_amount,
                remaining,  # ✅ 用全时段累计出库量计算的“未出库数量”
                unit_price,
                (unit_price or 0) * (qty or 0),
                currency_type,
                exchange_rate,
                ((unit_price or 0) * (qty or 0)) * exchange_rate,
                getattr(detail, "remark", "") or "",
            ],
            widths,
            center_cols=center_cols,
        )
        r += 1

    ws.freeze_panes = ws[f"A{data_start_row}"]
    _apply_widths(ws, widths)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"成品出库记录_自产_{_now_tag()}.xlsx"
    return bio, filename


# ============== 导出：成品出入库合并（仅自产） ==============


def build_finished_inout_excel(filters: dict):
    start_dt = _parse_dt(filters.get("start_date"), end=False)
    end_dt = _parse_dt(filters.get("end_date"), end=True)

    # 入库（仅自产）
    in_q = (
        db.session.query(
            Order,
            Shoe.shoe_rid,
            OrderShoe.customer_product_name,
            Color.color_name,
            Customer,
            ShoeInboundRecord,
            ShoeInboundRecordDetail,
        )
        .join(OrderShoe, Order.order_id == OrderShoe.order_id)
        .join(Shoe, Shoe.shoe_id == OrderShoe.shoe_id)
        .join(Customer, Customer.customer_id == Order.customer_id)
        .join(OrderShoeType, OrderShoeType.order_shoe_id == OrderShoe.order_shoe_id)
        .join(ShoeType, ShoeType.shoe_type_id == OrderShoeType.shoe_type_id)
        .join(Color, Color.color_id == ShoeType.color_id)
        .join(
            FinishedShoeStorage,
            FinishedShoeStorage.order_shoe_type_id == OrderShoeType.order_shoe_type_id,
        )
        .join(
            ShoeInboundRecordDetail,
            ShoeInboundRecordDetail.finished_shoe_storage_id
            == FinishedShoeStorage.finished_shoe_id,
        )
        .join(
            ShoeInboundRecord,
            ShoeInboundRecord.shoe_inbound_record_id
            == ShoeInboundRecordDetail.shoe_inbound_record_id,
        )
        .filter(ShoeInboundRecord.transaction_type == 1)
        .filter(ShoeInboundRecordDetail.is_deleted == 0)
        .filter(ShoeInboundRecord.inbound_type == 0)  # 仅自产入库
        .order_by(desc(ShoeInboundRecord.inbound_datetime))
        .distinct(ShoeInboundRecordDetail.record_detail_id)
    )
    if start_dt:
        in_q = in_q.filter(ShoeInboundRecord.inbound_datetime >= start_dt)
    if end_dt:
        in_q = in_q.filter(ShoeInboundRecord.inbound_datetime <= end_dt)
    if filters.get("inbound_rid"):
        in_q = in_q.filter(
            _safe_ilike(ShoeInboundRecord.shoe_inbound_rid, filters["inbound_rid"])
        )
    in_q = _apply_common_filters(in_q, filters)

    # 出库（仅自产）
    out_q = (
        db.session.query(
            Order,
            Shoe.shoe_rid,
            OrderShoe.customer_product_name,
            Color.color_name,
            Customer,
            ShoeOutboundRecord,
            ShoeOutboundRecordDetail,
        )
        .join(OrderShoe, Order.order_id == OrderShoe.order_id)
        .join(Shoe, Shoe.shoe_id == OrderShoe.shoe_id)
        .join(Customer, Customer.customer_id == Order.customer_id)
        .join(OrderShoeType, OrderShoeType.order_shoe_id == OrderShoe.order_shoe_id)
        .join(ShoeType, ShoeType.shoe_type_id == OrderShoeType.shoe_type_id)
        .join(Color, Color.color_id == ShoeType.color_id)
        .join(
            FinishedShoeStorage,
            FinishedShoeStorage.order_shoe_type_id == OrderShoeType.order_shoe_type_id,
        )
        .join(
            ShoeOutboundRecordDetail,
            ShoeOutboundRecordDetail.finished_shoe_storage_id
            == FinishedShoeStorage.finished_shoe_id,
        )
        .join(
            ShoeOutboundRecord,
            ShoeOutboundRecord.shoe_outbound_record_id
            == ShoeOutboundRecordDetail.shoe_outbound_record_id,
        )
        .filter(ShoeOutboundRecord.outbound_type == 0)  # 仅自产出库
        .order_by(desc(ShoeOutboundRecord.outbound_datetime))
        .distinct(ShoeOutboundRecordDetail.record_detail_id)
    )
    if start_dt:
        out_q = out_q.filter(ShoeOutboundRecord.outbound_datetime >= start_dt)
    if end_dt:
        out_q = out_q.filter(ShoeOutboundRecord.outbound_datetime <= end_dt)
    if filters.get("outbound_rid"):
        out_q = out_q.filter(
            _safe_ilike(ShoeOutboundRecord.shoe_outbound_rid, filters["outbound_rid"])
        )
    out_q = _apply_common_filters(out_q, filters)

    in_rows = in_q.all()
    out_rows = out_q.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "成品出入库合并"

    header = [
        "方向",
        "订单号",
        "客户订单号",
        "客户名",
        "客户商标",
        "客户型号",
        "工厂型号",
        "颜色",
        "单号",
        "时间",
        "数量",
        "备注",
    ]
    widths: list[int] = []
    header_row = _write_title_filters(ws, "成品出入库合并", len(header), filters)
    _write_header(ws, header_row, header, widths)

    data_start_row = header_row + 1
    center_cols = {10, 11}  # 时间、数量
    r = data_start_row

    for (
        order,
        shoe_rid_v,
        cust_prod_name,
        color_name,
        customer,
        record,
        detail,
    ) in in_rows:
        _write_data_row(
            ws,
            r,
            [
                "入库",
                order.order_rid,
                order.order_cid,
                customer.customer_name,
                customer.customer_brand,
                cust_prod_name,
                shoe_rid_v,
                color_name,
                record.shoe_inbound_rid,
                _format_dt(record.inbound_datetime),
                detail.inbound_amount,
                
                detail.remark or "",
            ],
            widths,
            center_cols=center_cols,
        )
        r += 1

    for (
        order,
        shoe_rid_v,
        cust_prod_name,
        color_name,
        customer,
        record,
        detail,
    ) in out_rows:
        qty = getattr(detail, "outbound_amount", None)
        if qty is None:
            qty = getattr(detail, "amount", None)
        _write_data_row(
            ws,
            r,
            [
                "出库",
                order.order_rid,
                order.order_cid,
                customer.customer_name,
                customer.customer_brand,
                cust_prod_name,
                shoe_rid_v,
                color_name,
                record.shoe_outbound_rid,
                _format_dt(record.outbound_datetime),
                qty,
                getattr(detail, "remark", "") or "",
            ],
            widths,
            center_cols=center_cols,
        )
        r += 1

    ws.freeze_panes = ws[f"A{data_start_row}"]
    _apply_widths(ws, widths)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"成品出入库合并_自产_{_now_tag()}.xlsx"
    return bio, filename

def _format_amount_map(amount_map: dict | None) -> str:
    if not amount_map:
        return ""
    parts = []
    for currency in sorted(amount_map.keys()):
        raw_val = amount_map[currency]
        try:
            dec_val = Decimal(str(raw_val))
        except Exception:
            dec_val = Decimal(0)
        parts.append(f"{currency} {dec_val.quantize(Decimal('0.00'))}")
    return " / ".join(parts)


def _format_unit_price(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, Decimal):
        dec_val = value
    else:
        try:
            dec_val = Decimal(str(value))
        except Exception:
            dec_val = Decimal(0)
    return float(dec_val.quantize(Decimal("0.00")))


def build_finished_inout_summary_by_model_excel(
    rows: list[dict], stat: dict, filters: dict
):
    wb = Workbook()
    ws = wb.active
    ws.title = "成品仓库存出入库明细"

    header = [
        "存货名称",
        "型号",
        "客户货号",
        "期初数量",
        "本期入库",
        "金额",
        "本期出库",
        "金额",
        "结存数量",
        "单价",
        "结存金额",
    ]

    widths: list[int] = []
    header_row = _write_title_filters(ws, "成品仓库存出入库明细", len(header), filters)
    _write_header(ws, header_row, header, widths)

    center_cols = {4, 5, 7, 9, 10}
    row_idx = header_row + 1
    for row in rows:
        inventory_name = row.get("shoeRid") or "-"
        color = row.get("color") or ""
        if color:
            inventory_name = f"{inventory_name}-{color}"
        values = [
            inventory_name,
            row.get("shoeRid") or "-",
            row.get("category") or "-",
            int(row.get("openingQty") or 0),
            int(row.get("inQty") or 0),
            _format_amount_map(row.get("inAmountByCurrency")),
            int(row.get("outQty") or 0),
            _format_amount_map(row.get("outAmountByCurrency")),
            int(row.get("closingQty") or 0),
            _format_unit_price(row.get("unitPrice")),
            _format_amount_map(row.get("closingAmountByCurrency")),
        ]
        _write_data_row(ws, row_idx, values, widths, center_cols=center_cols)
        row_idx += 1

    summary_row = [
        "合计",
        "",
        "",
        int(stat.get("openingQty", 0)),
        int(stat.get("inQty", 0)),
        _format_amount_map(stat.get("inAmountByCurrency")),
        int(stat.get("outQty", 0)),
        _format_amount_map(stat.get("outAmountByCurrency")),
        int(stat.get("closingQty", 0)),
        "",
        _format_amount_map(stat.get("closingAmountByCurrency")),
    ]
    _write_data_row(ws, row_idx, summary_row, widths, center_cols=center_cols)
    ws.cell(row=row_idx, column=1).font = HEAD_FONT

    data_start_row = header_row + 1
    ws.freeze_panes = ws[f"A{data_start_row}"]
    _apply_widths(ws, widths)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"成品仓库存出入库明细_{_now_tag()}.xlsx"
    return bio, filename
