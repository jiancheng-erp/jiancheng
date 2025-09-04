# services/inventory_excel.py
from io import BytesIO
from datetime import datetime
from decimal import Decimal
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def build_inventory_excel(*, rows, size_range, inventory_date: str = "", inventory_reason: str = ""):
    """
    传入 query_inventory_summary 的 rows 与 size_range，返回 BytesIO（二进制Excel）
    - 新增一列「订单号」：采用 r.order_rid（可为空）
    """

    # 本地数值转换，兼容字符串与 None
    def _to_float(v):
        if v is None:
            return 0.0
        if isinstance(v, (int, float, Decimal)):
            return float(v)
        try:
            s = str(v).strip().replace(",", "").replace("，", "").replace("\u3000", "")
            return float(s) if s else 0.0
        except Exception:
            return 0.0

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "材料盘库"

    # 样式
    title_font = Font(size=14, bold=True)
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    header_fill = PatternFill("solid", fgColor="E9EEF3")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    # 抬头
    title = "材料盘库清单"
    # base_headers 在下方定义；此处先占位，等定义后计算列数
    # 表头（含尺码列）列总数 = len(base_headers) + len(size_range)
    # 为了方便，这里先定义 base_headers
    base_headers = [
        "SPU编号", "订单号",  # ← 新增“订单号”列，紧随 SPU编号 后更直观
        "厂家名称", "材料类型", "材料名称",
        "材料型号", "材料规格", "颜色", "单位",
        "总库存", "总价格", "平均单价"
    ]
    size_headers = [f"尺码{sz}" for sz in size_range]
    headers = base_headers + size_headers

    total_cols = len(headers)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.cell(row=1, column=1, value=title).font = title_font
    ws.cell(row=1, column=1).alignment = center

    ws.cell(row=2, column=1, value=f"盘库日期：{inventory_date or '-'}")
    # 第二行右侧信息的位置可保持不变；若想更靠右，也可把列号调大
    ws.cell(row=2, column=6, value=f"盘库原因：{inventory_reason or '-'}")
    ws.cell(row=3, column=1, value=f"导出时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    # 表头
    header_row = 5
    for idx, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=idx, value=h)
        c.font = bold
        c.alignment = center
        c.fill = header_fill
        c.border = thin_border

    # 数据
    start_row = header_row + 1
    for i, r in enumerate(rows, start=start_row):
        
        order_rid = getattr(r, "order_rid", None) or ""

        data_row = [
            getattr(r, "spu_rid", "") or "",
            order_rid,  # 新增列：订单号
            getattr(r, "material_supplier", "") or "",
            getattr(r, "material_type_name", "") or "",
            getattr(r, "material_name", "") or "",
            getattr(r, "material_model", "") or "",
            getattr(r, "material_specification", "") or "",
            getattr(r, "color", "") or "",
            getattr(r, "unit", "") or "",
            _to_float(getattr(r, "total_current_amount", 0)),
            _to_float(getattr(r, "total_value_amount", 0)),
            _to_float(getattr(r, "avg_unit_price", 0)),
        ]
        for sz in size_range:
            data_row.append(_to_float(getattr(r, f"size_{sz}", 0)))

        for j, val in enumerate(data_row, 1):
            c = ws.cell(row=i, column=j, value=val)
            c.border = thin_border
            # 从“总库存”开始为数值列（这里是第 10 列起），右对齐
            if j >= (base_headers.index("总库存") + 1):  # 1-based
                c.alignment = right

    # 列宽（根据新列顺序调整；你也可按需微调）
    # 1:SPU编号 2:订单号 3:厂家名称 4:材料类型 5:材料名称 6:材料型号 7:材料规格 8:颜色 9:单位 10:总库存 11:总价格 12:平均单价
    width_map = {
        1: 16,  # SPU编号
        2: 16,  # 订单号
        3: 16,  # 厂家名称
        4: 14,  # 材料类型
        5: 16,  # 材料名称
        6: 16,  # 材料型号
        7: 18,  # 材料规格
        8: 10,  # 颜色
        9: 8,   # 单位
        10: 12, # 总库存
        11: 14, # 总价格
        12: 12, # 平均单价
    }
    for col_idx, w in width_map.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    # 尺码列统一宽度
    for k in range(len(base_headers) + 1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(k)].width = 10

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

