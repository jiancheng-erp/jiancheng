from openpyxl.utils import range_boundaries

from openpyxl import load_workbook
from logger import logger
from warehouse import material_storage


def generate_accounting_warehouse_excel(
    template_path,
    save_path,
    warehouse_name,
    supplier_name,
    time_range,
    materials_data,
):

    # Load the workbook and select the first sheet
    workbook = load_workbook(template_path)
    sheet = workbook["Sheet1"]

    # Insert metadata
    sheet["C2"] = warehouse_name if warehouse_name else "全部"
    sheet["F2"] = supplier_name if supplier_name else "全部"
    sheet["K2"] = time_range if time_range else "全部"

    quantity_sum_map = {
        "pendingInbound": 0,
        "pendingOutbound": 0,
        "inboundAmount": 0,
        "outboundAmount": 0,
        "makeInventoryInbound": 0,
        "makeInventoryOutbound": 0,
        "currentAmount": 0,
        "currentItemTotalPrice": 0,
    }

    # Insert materials data starting from row 8
    start_row = 8
    current_row = start_row
    for index, data in enumerate(materials_data, start=start_row):
        for j, col in enumerate(
            [
                "materialWarehouse",
                "supplierName",
                "materialType",
                "materialName",
                "materialModel",
                "materialSpecification",
                "color",
                "orderRId",
                "customerProductName",
                "shoeRId",
                "pendingInbound",
                "pendingOutbound",
                "inboundAmount",
                "outboundAmount",
                "makeInventoryInbound",
                "makeInventoryOutbound",
                "currentAmount",
                "unitPrice",
                "actualInboundUnit",
                "averagePrice",
                "currentItemTotalPrice",
            ]
        ):
            # determine the column letter
            col_letter = chr(65 + j)  # 65 is 'A'
            cell = f"{col_letter}{index}"
            sheet[cell] = data.get(col, "")

        # Update sums
        quantity_sum_map["pendingInbound"] += data.get("pendingInbound", 0)
        quantity_sum_map["pendingOutbound"] += data.get("pendingOutbound", 0)
        quantity_sum_map["inboundAmount"] += data.get("inboundAmount", 0)
        quantity_sum_map["outboundAmount"] += data.get("outboundAmount", 0)
        quantity_sum_map["makeInventoryInbound"] += data.get("makeInventoryInbound", 0)
        quantity_sum_map["makeInventoryOutbound"] += data.get("makeInventoryOutbound", 0)
        quantity_sum_map["currentAmount"] += data.get("currentAmount", 0)
        quantity_sum_map["currentItemTotalPrice"] += data.get("currentItemTotalPrice", 0)

        current_row += 1
    
    # Insert totals in the row after the last data row
    total_row = current_row
    sheet[f"J{total_row}"] = "合计"
    sheet[f"K{total_row}"] = quantity_sum_map["pendingInbound"]
    sheet[f"L{total_row}"] = quantity_sum_map["pendingOutbound"]
    sheet[f"M{total_row}"] = quantity_sum_map["inboundAmount"]
    sheet[f"N{total_row}"] = quantity_sum_map["outboundAmount"]
    sheet[f"O{total_row}"] = quantity_sum_map["makeInventoryInbound"]
    sheet[f"P{total_row}"] = quantity_sum_map["makeInventoryOutbound"]
    sheet[f"Q{total_row}"] = quantity_sum_map["currentAmount"]
    sheet[f"U{total_row}"] = quantity_sum_map["currentItemTotalPrice"]

    # Save the modified file
    workbook.save(save_path)
    return save_path


def generate_accounting_warehouse_grouped_excel(
    save_path,
    warehouse_name,
    supplier_name,
    time_range,
    materials_data,
):
    """Generate Excel for inventory grouped by material name with weighted average price."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "按名称汇总"

    # Header rows
    ws["A1"] = "财务部历史库存 - 按材料名称汇总"
    ws["A2"] = "仓库"
    ws["B2"] = warehouse_name if warehouse_name else "全部"
    ws["C2"] = "供应商"
    ws["D2"] = supplier_name if supplier_name else "全部"
    ws["E2"] = "日期"
    ws["F2"] = time_range if time_range else "全部"

    headers = [
        "材料名称",
        "未审核入库数",
        "未审核出库数",
        "采购入库数",
        "生产出库数",
        "盘库入库数",
        "盘库出库数",
        "库存数量",
        "加权均价",
        "库存总金额",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=4, column=col_idx, value=header)

    quantity_sum_map = {
        "pendingInbound": 0,
        "pendingOutbound": 0,
        "inboundAmount": 0,
        "outboundAmount": 0,
        "makeInventoryInbound": 0,
        "makeInventoryOutbound": 0,
        "currentAmount": 0,
        "currentItemTotalPrice": 0,
    }

    col_keys = [
        "materialName",
        "pendingInbound",
        "pendingOutbound",
        "inboundAmount",
        "outboundAmount",
        "makeInventoryInbound",
        "makeInventoryOutbound",
        "currentAmount",
        "averagePrice",
        "currentItemTotalPrice",
    ]

    current_row = 5
    for data in materials_data:
        for col_idx, key in enumerate(col_keys, start=1):
            ws.cell(row=current_row, column=col_idx, value=data.get(key, ""))

        for sum_key in quantity_sum_map:
            quantity_sum_map[sum_key] += data.get(sum_key, 0) or 0

        current_row += 1

    # Totals row
    ws.cell(row=current_row, column=1, value="合计")
    ws.cell(row=current_row, column=2, value=quantity_sum_map["pendingInbound"])
    ws.cell(row=current_row, column=3, value=quantity_sum_map["pendingOutbound"])
    ws.cell(row=current_row, column=4, value=quantity_sum_map["inboundAmount"])
    ws.cell(row=current_row, column=5, value=quantity_sum_map["outboundAmount"])
    ws.cell(row=current_row, column=6, value=quantity_sum_map["makeInventoryInbound"])
    ws.cell(row=current_row, column=7, value=quantity_sum_map["makeInventoryOutbound"])
    ws.cell(row=current_row, column=8, value=quantity_sum_map["currentAmount"])
    ws.cell(row=current_row, column=10, value=quantity_sum_map["currentItemTotalPrice"])

    wb.save(save_path)
    return save_path
