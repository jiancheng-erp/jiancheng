import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from logger import logger


def generate_excel_file(template_path, new_file_path, order_data):
    """
    生成材料订购单 Excel（新格式，含采购明细侧边栏）。

    order_data keys:
        供应商, 客户名, 订单信息, 日期, 备注,
        环保要求, 发货地址, 交货期限,
        seriesData: list of {物品名称, 单位, 数量, 单价, 用途说明, 备注}
    """
    logger.debug(f"Generating Excel file for order {order_data.get('订单信息', '')}")

    wb = Workbook()
    ws = wb.active
    ws.title = "材料订购单"

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── Row 1: title ─────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    ws["A1"] = "材 料 订 购 单"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 36

    # ── Row 2: supplier / order_rid / date ───────────────────────────────────
    ws.merge_cells("A2:C2")
    ws["A2"] = "供应商：" + order_data.get("供应商", "")
    ws["A2"].alignment = left_align

    ws.merge_cells("D2:E2")
    ws["D2"] = order_data.get("订单信息", "")
    ws["D2"].alignment = center

    ws["F2"] = "日期："
    ws["F2"].alignment = center

    ws.merge_cells("G2:H2")
    ws["G2"] = order_data.get("日期", "")
    ws["G2"].alignment = center
    ws.row_dimensions[2].height = 22

    # ── Determine data range ──────────────────────────────────────────────────
    series = order_data.get("seriesData", [])
    min_rows = max(len(series), 5)
    # Row 3: header; Rows 4..(3+min_rows): data
    last_data_row = 3 + min_rows

    # ── Column A sidebar: 采购明细 (merged header + all data rows) ─────────────
    ws.merge_cells(f"A3:A{last_data_row}")
    ws["A3"] = "采购明细"
    ws["A3"].font = Font(bold=True, size=11)
    ws["A3"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True, text_rotation=255
    )
    ws["A3"].border = border
    # openpyxl only paints border on the top-left cell of a merge; set left/right
    # on every cell in the column so the sidebar outline is fully visible
    left_right = Border(left=thin, right=thin)
    for _r in range(4, last_data_row + 1):
        ws.cell(row=_r, column=1).border = left_right
    # ensure bottom border on last cell of merged range
    ws.cell(row=last_data_row, column=1).border = Border(left=thin, right=thin, bottom=thin)

    # ── Row 3: column headers ─────────────────────────────────────────────────
    headers = ["序号", "物品名称", "单位", "数量", "单价", "用途说明", "备注"]
    for col_idx, header in enumerate(headers, start=2):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[3].height = 20

    # ── Data rows ─────────────────────────────────────────────────────────────
    total_qty = 0
    for i in range(min_rows):
        row = 4 + i
        item = series[i] if i < len(series) else {}
        qty = item.get("数量", "") if item else ""
        if qty not in ("", None):
            try:
                total_qty += float(qty)
            except (TypeError, ValueError):
                pass
        values = [
            i + 1,
            item.get("物品名称", ""),
            item.get("单位", ""),
            qty,
            item.get("单价", ""),
            item.get("用途说明", ""),
            item.get("备注", ""),
        ]
        for col_idx, value in enumerate(values, start=2):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = border
            cell.alignment = left_align if col_idx == 3 else center
        ws.row_dimensions[row].height = 28

    # ── Footer rows ──────────────────────────────────────────────────────────
    r = last_data_row + 1

    # 合计 row
    ws.merge_cells(f"A{r}:D{r}")
    ws[f"A{r}"] = "合   计"
    ws[f"A{r}"].font = Font(bold=True)
    ws[f"A{r}"].alignment = center
    ws[f"A{r}"].border = border
    for col in range(2, 5):  # B, C, D — apply border so merged cell has full outline
        ws.cell(row=r, column=col).border = border
    ws[f"E{r}"] = int(total_qty) if total_qty else ""
    ws[f"E{r}"].alignment = center
    ws[f"E{r}"].border = border
    for col in range(6, 9):  # F, G, H
        ws.cell(row=r, column=col).border = border
    ws.row_dimensions[r].height = 28

    # 发货地址 / 联系人
    r += 1
    ws.merge_cells(f"A{r}:D{r}")
    ws[f"A{r}"] = "发货地址：" + order_data.get("发货地址", "")
    ws[f"A{r}"].alignment = left_align
    ws[f"A{r}"].font = Font(size=9)
    ws.merge_cells(f"E{r}:H{r}")
    ws[f"E{r}"] = "联系人：范先生-13868846816"
    ws[f"E{r}"].alignment = left_align
    ws[f"E{r}"].font = Font(size=9)
    ws.row_dimensions[r].height = 22

    # 交货周期 / 责任条款
    r += 1
    ws.merge_cells(f"A{r}:D{r}")
    ws[f"A{r}"] = "交货周期：" + order_data.get("交货期限", "")
    ws[f"A{r}"].alignment = left_align
    ws[f"A{r}"].font = Font(color="0070C0", size=9)
    ws.merge_cells(f"E{r}:H{r}")
    ws[f"E{r}"] = "如有特殊情况提前5天反馈，无故延期有贵公司承担后续责任。"
    ws[f"E{r}"].alignment = left_align
    ws[f"E{r}"].font = Font(size=9)
    ws.row_dimensions[r].height = 22

    # 制表 / 审核
    r += 1
    ws.merge_cells(f"A{r}:D{r}")
    ws[f"A{r}"] = "制表："
    ws[f"A{r}"].alignment = left_align
    ws[f"A{r}"].font = Font(size=9)
    ws.merge_cells(f"E{r}:H{r}")
    ws[f"E{r}"] = "审核："
    ws[f"E{r}"].alignment = left_align
    ws[f"E{r}"].font = Font(size=9)
    ws.row_dimensions[r].height = 22

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [5, 5, 30, 8, 10, 10, 18, 15]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    os.makedirs(os.path.dirname(new_file_path), exist_ok=True)
    wb.save(new_file_path)
    logger.debug(f"材料订购单 saved: {new_file_path}")


# def test_case_1():
#     template_path = "H:\git-projects\jiancheng\\backend-python\general_document/标准采购订单.xlsx"
#     new_file_path = "H:\git-projects\jiancheng\\backend-python\general_document/pur_test.xlsx"
#     order_data = {
#         "订单信息": "订单编号12345",
#         "供应商": "供应商A",
#         "日期": "2024-11-19",
#         "seriesData": [
#             {"物品名称": "商品1", "单位": "个", "数量": 10, "单价": 5.5, "用途说明": "用途1", "备注": "备注1"},
#             {"物品名称": "商品2", "单位": "箱", "数量": 20, "单价": 15.0, "用途说明": "用途2", "备注": "备注2"},
#         ],
#         "合计": 350,
#         "备注": "总备注",
#         "环保要求": "符合环保标准",
#         "发货地址": "测试地址",
#         "交货期限": "2024-12-01",
#     }
#     generate_excel_file(template_path, new_file_path, order_data)
#     logger.debug("Test Case 1: Basic functionality passed.")

