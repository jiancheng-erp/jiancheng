# services/finished_outbound_export.py
# -*- coding: utf-8 -*-
import os
import io
from datetime import datetime
from decimal import Decimal

from openpyxl import load_workbook

from app_config import db
from models import (
    OrderShoe,
    OrderShoeBatchInfo,
    PackagingInfo,
    ShoeOutboundRecord,
    FinishedShoeStorage,
    OrderShoeType,
    Order,
    Customer,
    Shoe,
    Color,
    ShoeType,
    ShoeOutboundApply,
    ShoeOutboundApplyDetail,
)


def _normalize_int_ids(values):
    normalized = []
    for value in values or []:
        if value in (None, ""):
            continue
        try:
            normalized.append(int(value))
        except (TypeError, ValueError):
            raise ValueError("ID 列表包含非法数字")
    return normalized


def _collect_outbound_record_list(
    outbound_record_ids=None,
    outbound_rids=None,
    apply_ids=None,
):
    outbound_record_ids = _normalize_int_ids(outbound_record_ids)
    outbound_rids = outbound_rids or []
    apply_ids = set(_normalize_int_ids(apply_ids))

    record_id_filter = set()
    record_id_list = []

    if outbound_record_ids or outbound_rids:
        q_record = db.session.query(ShoeOutboundRecord)
        if outbound_record_ids:
            q_record = q_record.filter(
                ShoeOutboundRecord.shoe_outbound_record_id.in_(outbound_record_ids)
            )
        if outbound_rids:
            q_record = q_record.filter(
                ShoeOutboundRecord.shoe_outbound_rid.in_(outbound_rids)
            )
        records = q_record.all()
        if not records:
            raise ValueError("未找到对应的出库记录")

        record_id_filter.update(
            record.shoe_outbound_record_id for record in records if record
        )
        record_id_list = list(record_id_filter)

        if record_id_list:
            apply_rows = (
                db.session.query(ShoeOutboundApply.apply_id)
                .filter(
                    ShoeOutboundApply.outbound_record_id.in_(record_id_list),
                    ShoeOutboundApply.apply_id.isnot(None),
                )
                .all()
            )
            apply_ids.update(row[0] for row in apply_rows if row and row[0])

        if not apply_ids:
            raise ValueError("这些出库记录没有关联任何申请单（未匹配到申请单）")

    if not apply_ids:
        raise ValueError("未找到对应的申请单")

    query = (
        db.session.query(
            ShoeOutboundApply,
            ShoeOutboundRecord,
            ShoeOutboundApplyDetail,
            FinishedShoeStorage,
            OrderShoeType,
            OrderShoe,
            Order,
            Customer,
            Shoe,
            ShoeType,
            Color,
            OrderShoeBatchInfo,
            PackagingInfo,
        )
        .outerjoin(
            ShoeOutboundRecord,
            ShoeOutboundRecord.shoe_outbound_record_id
            == ShoeOutboundApply.outbound_record_id,
        )
        .join(
            ShoeOutboundApplyDetail,
            ShoeOutboundApplyDetail.apply_id == ShoeOutboundApply.apply_id,
        )
        .outerjoin(
            FinishedShoeStorage,
            FinishedShoeStorage.finished_shoe_id
            == ShoeOutboundApplyDetail.finished_shoe_storage_id,
        )
        .outerjoin(
            OrderShoeType,
            OrderShoeType.order_shoe_type_id
            == ShoeOutboundApplyDetail.order_shoe_type_id,
        )
        .outerjoin(
            OrderShoe,
            OrderShoe.order_shoe_id == OrderShoeType.order_shoe_id,
        )
        .outerjoin(
            Order,
            Order.order_id == OrderShoe.order_id,
        )
        .outerjoin(
            Customer,
            Customer.customer_id == Order.customer_id,
        )
        .outerjoin(
            ShoeType,
            ShoeType.shoe_type_id == OrderShoeType.shoe_type_id,
        )
        .outerjoin(
            Shoe,
            Shoe.shoe_id == ShoeType.shoe_id,
        )
        .outerjoin(
            Color,
            Color.color_id == ShoeType.color_id,
        )
        .outerjoin(
            OrderShoeBatchInfo,
            OrderShoeBatchInfo.order_shoe_batch_info_id
            == ShoeOutboundApplyDetail.order_shoe_batch_info_id,
        )
        .outerjoin(
            PackagingInfo,
            PackagingInfo.packaging_info_id
            == ShoeOutboundApplyDetail.packaging_info_id,
        )
        .filter(ShoeOutboundApply.apply_id.in_(apply_ids))
    )

    if record_id_list:
        query = query.filter(
            ShoeOutboundRecord.shoe_outbound_record_id.in_(record_id_list)
        )
    elif outbound_rids:
        query = query.filter(
            ShoeOutboundRecord.shoe_outbound_rid.in_(outbound_rids)
        )

    rows = query.all()
    if not rows:
        raise ValueError("未找到对应的申请明细记录")

    record_list = []
    for (
        apply_obj,
        outbound,
        apply_detail,
        finished,
        ost,
        oss,
        order,
        customer,
        shoe,
        shoe_type,
        color,
        batch_info,
        pkg,
    ) in rows:
        customer_name = getattr(customer, "customer_name", "") or ""
        customer_brand = getattr(customer, "customer_brand", "") or ""

        expected_dt = None
        if apply_obj and apply_obj.expected_outbound_datetime:
            expected_dt = apply_obj.expected_outbound_datetime

        actual_dt = None
        if apply_obj and apply_obj.actual_outbound_datetime:
            actual_dt = apply_obj.actual_outbound_datetime
        elif outbound and outbound.outbound_datetime:
            actual_dt = outbound.outbound_datetime

        expected_date_str = expected_dt.strftime("%Y-%m-%d") if expected_dt else ""
        actual_date_str = actual_dt.strftime("%Y-%m-%d") if actual_dt else ""

        style_no_customer = ""
        if oss:
            style_no_customer = (
                getattr(oss, "customer_product_name", None)
                or getattr(oss, "customer_shoe_type", "")
                or ""
            )
        color_name = getattr(color, "color_name", "") or ""
        factory_style_no = getattr(shoe, "shoe_rid", "") or ""
        batch_name = getattr(batch_info, "name", "") if batch_info else ""

        pairs_per_carton = None
        carton_count = None
        if apply_detail:
            if apply_detail.pairs_per_carton is not None:
                pairs_per_carton = int(apply_detail.pairs_per_carton or 0)
            if apply_detail.carton_count is not None:
                carton_count = float(apply_detail.carton_count)

        qty_pairs = 0
        if apply_detail and apply_detail.total_pairs:
            qty_pairs = int(apply_detail.total_pairs or 0)
        elif pairs_per_carton is not None and carton_count is not None:
            qty_pairs = int(round(pairs_per_carton * carton_count))

        unit_price = Decimal("0.000")
        if ost and ost.unit_price is not None:
            unit_price = ost.unit_price
        amount = (Decimal(qty_pairs) * unit_price).quantize(Decimal("0.01"))

        factory_name = "浙江健诚鞋业集团有限公司"

        record_list.append(
            {
                "customer_name": customer_name,
                "customer_brand": customer_brand,
                "expected_date": expected_date_str,
                "actual_date": actual_date_str,
                "style_no_customer": style_no_customer,
                "color_name": color_name,
                "factory_style_no": factory_style_no,
                "batch_name": batch_name,
                "qty_pairs": qty_pairs,
                "pairs_per_carton": pairs_per_carton,
                "carton_count": carton_count,
                "unit_price": unit_price,
                "amount": amount,
                "factory_name": factory_name,
            }
        )

    return record_list


def _export_record_list_to_excel(record_list, template_path):
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"模板文件不存在: {template_path}")
    if not record_list:
        raise ValueError("未找到对应的申请明细记录")

    record_list.sort(
        key=lambda r: (
            r["customer_name"] or "",
            r["customer_brand"] or "",
            r["actual_date"] or "",
            r["factory_style_no"] or "",
            r["color_name"] or "",
            r["style_no_customer"] or "",
            r["batch_name"] or "",
        )
    )

    wb = load_workbook(template_path)
    ws = wb["出货单"]
    start_row = 3

    groups = []
    prev_customer = None
    group_start_idx = None

    for idx, rec in enumerate(record_list):
        cname = rec["customer_name"]
        if cname != prev_customer:
            if prev_customer is not None and group_start_idx is not None:
                groups.append(
                    (
                        prev_customer,
                        start_row + group_start_idx,
                        start_row + idx - 1,
                    )
                )
            prev_customer = cname
            group_start_idx = idx

    if prev_customer is not None and group_start_idx is not None:
        groups.append(
            (
                prev_customer,
                start_row + group_start_idx,
                start_row + len(record_list) - 1,
            )
        )

    for offset, rec in enumerate(record_list):
        row_idx = start_row + offset

        cname = rec["customer_name"]
        brand = rec["customer_brand"]
        expected_date = rec["expected_date"]
        actual_date = rec["actual_date"]
        style_no_customer = rec["style_no_customer"]
        color_name = rec["color_name"]
        factory_style_no = rec["factory_style_no"]
        batch_name = rec["batch_name"]
        qty_pairs = rec["qty_pairs"]
        pairs_per_carton = rec["pairs_per_carton"]
        carton_count = rec["carton_count"]
        unit_price = rec["unit_price"]
        amount = rec["amount"]

        is_group_start = False
        for gcname, g_start, g_end in groups:
            if gcname == cname and g_start == row_idx:
                is_group_start = True
                break

        if is_group_start:
            ws.cell(row=row_idx, column=1, value=cname)
        ws.cell(row=row_idx, column=2, value=brand)

        ws.cell(row=row_idx, column=3, value=expected_date)
        ws.cell(row=row_idx, column=4, value=actual_date)
        ws.cell(row=row_idx, column=5, value=style_no_customer)
        ws.cell(row=row_idx, column=6, value=color_name)
        ws.cell(row=row_idx, column=7, value=factory_style_no)
        ws.cell(row=row_idx, column=8, value=batch_name)
        ws.cell(row=row_idx, column=9, value=qty_pairs)

        if pairs_per_carton is not None:
            ws.cell(row=row_idx, column=10, value=int(pairs_per_carton))
        if carton_count is not None:
            ws.cell(row=row_idx, column=11, value=float(carton_count))

        ws.cell(row=row_idx, column=12, value=float(unit_price))
        ws.cell(row=row_idx, column=13, value=float(amount))

    for cname, g_start, g_end in groups:
        if g_start < g_end:
            ws.merge_cells(
                start_row=g_start,
                start_column=1,
                end_row=g_end,
                end_column=1,
            )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"出货清单_{now_str}.xlsx"

    return output, filename


def generate_finished_outbound_excel(
    template_path: str,
    outbound_record_ids=None,
    outbound_rids=None,
):
    record_list = _collect_outbound_record_list(
        outbound_record_ids=outbound_record_ids,
        outbound_rids=outbound_rids,
    )
    return _export_record_list_to_excel(record_list, template_path)


def generate_finished_outbound_apply_excel(template_path: str, apply_ids=None):
    record_list = _collect_outbound_record_list(apply_ids=apply_ids)
    return _export_record_list_to_excel(record_list, template_path)
