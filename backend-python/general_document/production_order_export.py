import shutil
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.drawing.image import Image
from constants import SHOESIZERANGE
from file_locations import IMAGE_UPLOAD_PATH
import os
from collections import defaultdict
from openpyxl.utils import units
from openpyxl.styles import Font
from logger import logger

merge_ranges = []

def add_totals_section(
    ws,
    data_start_row: int,
    data_end_row: int,
    currency_type: str,
    size_start_col_letter: str = "G",
    size_names=None,
    include_price: bool = True,
):
    """
    在数据区下方插入合计行：件数、双数、金额
    - 件数：count 列（每行的包装件数）
    - 双数：pairs 列（总双数 = total_quantity * count）
    - 金额：amount 列（单价 * 总双数）
    """
    if size_names is None:
        size_names = []

    # 计算尺码实际列数
    size_start_idx = column_index_from_string(size_start_col_letter)
    actual_size_cols = sum(1 for n in size_names if n not in (None, ""))

    # 基于尺码列后的固定相对位置定位列
    qty_col_idx    = size_start_idx + actual_size_cols          # 总双数(单套)列（非合计）
    count_col_idx  = qty_col_idx + 1                            # 件数列
    pairs_col_idx  = qty_col_idx + 2                            # 双数列（=总双数*件数）
    remark_col_idx = qty_col_idx + 3
    price_col_idx  = qty_col_idx + 4
    amount_col_idx = qty_col_idx + 5                            # 金额列

    count_col_letter  = get_column_letter(count_col_idx)
    pairs_col_letter  = get_column_letter(pairs_col_idx)
    amount_col_letter = get_column_letter(amount_col_idx)

    next_row = data_end_row + 1
    if next_row <= ws.max_row and is_empty_row(ws, next_row):
        totals_row = data_end_row + 2   # 下一行确实存在而且是空的，留一行空白
    else:
        totals_row = data_end_row + 1   # 否则紧跟着放（包括“下一行不存在”的情况）

    # 合计标题放在 “C:E” 合并
    ws.merge_cells(f"C{totals_row}:E{totals_row}")
    ws[f"C{totals_row}"] = "合计"
    ws[f"C{totals_row}"].font = Font(bold=True)

    # 件数合计
    ws[f"{count_col_letter}{totals_row}"] = f"=SUM({count_col_letter}{data_start_row}:{count_col_letter}{data_end_row})"
    ws[f"{count_col_letter}{totals_row}"].font = Font(bold=True)

    # 双数合计
    ws[f"{pairs_col_letter}{totals_row}"] = f"=SUM({pairs_col_letter}{data_start_row}:{pairs_col_letter}{data_end_row})"
    ws[f"{pairs_col_letter}{totals_row}"].font = Font(bold=True)

    if include_price:
        ws[f"{amount_col_letter}{totals_row}"] = f"=SUM({amount_col_letter}{data_start_row}:{amount_col_letter}{data_end_row})"
        ws[f"{amount_col_letter}{totals_row}"].number_format = get_currency_format(currency_type or "")
        ws[f"{amount_col_letter}{totals_row}"].font = Font(bold=True)
    else:
        ws[f"{amount_col_letter}{totals_row}"] = ""

    # 让合计行更显眼：行高、（可选）边框你也可以按需加
    ws.row_dimensions[totals_row].height = 22


def is_empty_row(ws, row_index: int) -> bool:
    """Return True when every cell in the given row is blank or None."""
    for cell in ws[row_index]:
        if cell.value not in (None, ""):
            return False
    return True

def trim_trailing_blank_after_data(ws, last_data_row: int):
    """
    删除数据块末尾紧跟的所有空行（通常是插入逻辑遗留的那一行）。
    注意：在插入“合计”之前调用。
    """
    next_row = last_data_row + 1
    # 连续空行都删掉，保证数据块正好以最后一行数据结束
    while next_row <= ws.max_row and is_empty_row(ws, next_row):
        ws.delete_rows(next_row, 1)

def reset_print_area(ws):
    """
    将打印区域设置为“真实内容”占用的最小矩形：
    - 普通单元格：仅当 cell.value 非空才纳入
    - 合并单元格：仅当左上角单元格有值才纳入（空标题等不会撑大）
    - 图片：把图片锚点所在单元格也纳入
    同时设置为按 1 页宽度缩放。
    """
    from openpyxl.utils import get_column_letter, coordinate_to_tuple

    def update_bounds(c, r):
        nonlocal min_col, min_row, max_col, max_row
        if min_col is None or c < min_col:
            min_col = c
        if max_col is None or c > max_col:
            max_col = c
        if min_row is None or r < min_row:
            min_row = r
        if max_row is None or r > max_row:
            max_row = r

    min_col = min_row = max_col = max_row = None

    # 1) 非空单元格
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value not in (None, ""):
                update_bounds(cell.column, cell.row)

    # 2) 有值的合并区域（只看左上角是否有值）
    for rng in ws.merged_cells.ranges:
        mc_min_col, mc_min_row, mc_max_col, mc_max_row = rng.bounds
        tl = ws.cell(row=mc_min_row, column=mc_min_col)
        if tl.value not in (None, ""):
            update_bounds(mc_min_col, mc_min_row)
            update_bounds(mc_max_col, mc_max_row)

    # 3) 图片锚点
    for img in getattr(ws, "_images", []):
        try:
            anc = img.anchor
            col = row = None
            # 可能是 "A1" 字符串
            if isinstance(anc, str):
                col, row = coordinate_to_tuple(anc)
            else:
                # 也可能是 OneCellAnchor / TwoCellAnchor
                from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, TwoCellAnchor
                if isinstance(anc, OneCellAnchor):
                    col = anc._from.col + 1
                    row = anc._from.row + 1
                elif isinstance(anc, TwoCellAnchor):
                    col = anc._to.col + 1
                    row = anc._to.row + 1
            if col and row:
                update_bounds(col, row)
        except Exception:
            pass

    # 没有内容就不设打印区
    if min_col is None:
        ws.print_area = None
        return

    ws.print_area = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"

    # 按一页宽度缩放，行数不限
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def get_dynamic_columns(size_names, size_start_col_letter="G"):
    """
    基于尺码列的实际数量，计算 qty/count/pairs/remark/price/amount 的列索引与字母。
    和 add_totals_section 的定位逻辑保持一致。
    """
    size_start_idx = column_index_from_string(size_start_col_letter)
    actual_size_cols = sum(1 for n in size_names if n not in (None, ""))

    qty_idx    = size_start_idx + actual_size_cols
    count_idx  = qty_idx + 1
    pairs_idx  = qty_idx + 2
    remark_idx = qty_idx + 3
    price_idx  = qty_idx + 4
    amount_idx = qty_idx + 5

    return {
        "qty_idx": qty_idx, "qty_letter": get_column_letter(qty_idx),
        "count_idx": count_idx, "count_letter": get_column_letter(count_idx),
        "pairs_idx": pairs_idx, "pairs_letter": get_column_letter(pairs_idx),
        "remark_idx": remark_idx, "remark_letter": get_column_letter(remark_idx),
        "price_idx": price_idx, "price_letter": get_column_letter(price_idx),
        "amount_idx": amount_idx, "amount_letter": get_column_letter(amount_idx),
    }


from openpyxl.styles import Alignment
def apply_remark_wrap_and_row_height(ws, remark_col_letter: str, start_row: int, end_row: int,
                                     long_len: int = 20, min_height_if_long: float = 40.0):
    """
    让备注列自动换行；对较长备注的行，若当前行高偏小则适当增高。
    注意：我们只“增高”，不回退已有的更高行高（例如含图片的行）。
    """
    for r in range(start_row, end_row + 1):
        cell = ws[f"{remark_col_letter}{r}"]
        # 开启自动换行，顶部对齐
        old = cell.alignment or Alignment()
        cell.alignment = Alignment(
            horizontal=old.horizontal, vertical="top",
            wrap_text=True, shrink_to_fit=False,
            text_rotation=old.text_rotation, indent=old.indent
        )
        # 简单按文本长度启发式增高
        v = cell.value
        if isinstance(v, str) and len(v.strip()) >= long_len:
            current = ws.row_dimensions[r].height
            if current is None or current < min_height_if_long:
                ws.row_dimensions[r].height = min_height_if_long


def ensure_min_width(ws, col_letters, min_width: float):
    """把指定列的宽度下限提升到 min_width，不会减小已有更大宽度。"""
    for letter in col_letters:
        cur = ws.column_dimensions[letter].width
        # 某些模板列宽可能是 None，这里统一保障下限
        if cur is None or cur < min_width:
            ws.column_dimensions[letter].width = float(min_width)


def remove_price_amount_columns(ws, cols):
    """删除价格与金额两列，包含表头与数据。"""
    amount_idx = column_index_from_string(cols["amount_letter"])
    price_idx = column_index_from_string(cols["price_letter"])
    # 先删金额列，避免影响价格列索引
    ws.delete_cols(amount_idx, 1)
    ws.delete_cols(price_idx, 1)


def get_currency_format(currency_type):
    mapping = {
        "USD": '"$"#,##0.00',
        "USA": '"$"#,##0.00',
        "RMB": '"￥"#,##0.00',
        "CNY": '"￥"#,##0.00',
        "EUR": '"€"#,##0.00'
    }
    return mapping.get(currency_type.upper(), '#,##0.00')  # 默认2位小数无符号
def find_remark_column(ws, header_row=5, remark_header="备注"):
    for col_idx in range(1, ws.max_column + 1):
        if str(ws[f"{get_column_letter(col_idx)}{header_row}"].value).strip() == remark_header:
            return col_idx
    return None

def adjust_title_merge(ws, title_row=1, start_col_letter="A"):
    """
    调整第一行标题合并单元格范围，使其覆盖下方有内容的最右列
    """
    # 找到下方数据区域的最右列索引
    max_col_idx = 1
    for row in ws.iter_rows(min_row=title_row + 1):
        for cell in reversed(row):  # 从右向左找
            if cell.value not in (None, ""):
                if cell.column > max_col_idx:
                    max_col_idx = cell.column
                break

    # 列字母
    end_col_letter = get_column_letter(max_col_idx)

    # 删除原有第一行的合并区域
    for merged_range in list(ws.merged_cells.ranges):
        if str(merged_range).startswith(f"{start_col_letter}{title_row}:"):
            ws.merged_cells.ranges.remove(merged_range)

    # 重新合并 A1 到最右列
    merge_range = f"{start_col_letter}{title_row}:{end_col_letter}{title_row}"
    ws.merge_cells(merge_range)
def set_global_font(ws, font_size=14):
    """
    为整个工作表设置统一字体和字号，但保留原有的粗体、颜色等属性，
    并跳过第一行。
    """
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        if row_idx == 1:  # 跳过第一行
            continue
        for cell in row:
            if cell.value is not None:
                old_font = cell.font or Font()
                cell.font = Font(
                    size=font_size,
                    bold=old_font.bold,
                    italic=old_font.italic,
                    vertAlign=old_font.vertAlign,
                    underline=old_font.underline,
                    strike=old_font.strike,
                    color=old_font.color
                )
def auto_adjust_column_width(ws, max_width=50, size_start_col_letter="G", size_max_width=6):
    """
    根据内容自动调整列宽，但不超过 max_width。
    尺码列从 size_start_col_letter 开始，单独限制为 size_max_width。
    """
    size_start_idx = column_index_from_string(size_start_col_letter)

    for col_cells in ws.columns:
        first_cell = next((c for c in col_cells if c is not None), None)
        if first_cell is None:
            continue

        col_idx = first_cell.column
        col_letter = get_column_letter(col_idx)

        # 判断是否是尺码列
        if col_idx >= size_start_idx:
            ws.column_dimensions[col_letter].width = size_max_width
            continue

        max_length = 0
        for cell in col_cells:
            if cell.value is not None:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length

        needed_width = min(max_length + 1, max_width)
        current_width = ws.column_dimensions[col_letter].width or 8.43
        if needed_width > current_width:
            ws.column_dimensions[col_letter].width = needed_width


# Function to load the Excel template and prepare for modification
def load_template(template_path, new_file_path):
    # Copy the template to a new file
    shutil.copy(template_path, new_file_path)
    # Load the new workbook
    wb = load_workbook(new_file_path)
    return wb


def get_next_column_name(current_column_name):
    # Convert column letter to index
    column_index = column_index_from_string(current_column_name)
    # Increment the index to get the next column
    next_column_index = column_index + 1
    # Convert the new index back to a column letter
    next_column_name = get_column_letter(next_column_index)
    return next_column_name


# Function to copy formatting and insert a new row
def insert_row_with_format(ws, row_to_copy, new_row_idx):
    # Insert a new row
    ws.insert_rows(new_row_idx)
    ws.row_dimensions[new_row_idx].height = ws.row_dimensions[row_to_copy].height
    # Copy the formatting
    for col_idx, cell in enumerate(ws[row_to_copy], start=1):
        new_cell = ws.cell(row=new_row_idx, column=col_idx)
        if cell.has_style:
            new_cell.border = cell.border.copy()
            new_cell.alignment = cell.alignment.copy()
            new_cell.number_format = cell.number_format

def delete_extra_size_columns(ws, size_name_list, start_col_letter="G", total_size_count=13):
    """
    删除从 start_col_letter 开始的尺码列，只保留有效列。
    并重新设置第7~8行的“尺码”标题区域（用于生产单模板）。
    """
    start_col_idx = column_index_from_string(start_col_letter)
    actual_size_cols = sum(1 for name in size_name_list if name not in ("", None))
    extra_count = total_size_count - actual_size_cols

    # 删除多余列（从 G + 有效数量 开始）
    if extra_count > 0:
        delete_start_col = start_col_idx + actual_size_cols
        ws.delete_cols(delete_start_col, extra_count)

    # 移除原 G7:G8 合并区域
    for merged_range in list(ws.merged_cells.ranges):
        if str(merged_range).startswith(f"{start_col_letter}4:"):
            ws.merged_cells.ranges.remove(merged_range)

    # 重新设置“尺码”标题区域合并 G7:??8
    if actual_size_cols > 0:
        end_col_letter = get_column_letter(start_col_idx + actual_size_cols - 1)
        merge_range = f"{start_col_letter}4:{end_col_letter}4"
        ws.merge_cells(merge_range)
        cell = ws[f"{start_col_letter}4"]
        cell.value = "尺码"
        cell.font = Font(bold=True)
        
def delete_zero_size_columns(ws, size_name_list, start_col_letter: str, data_start_row: int, data_end_row: int):
    """
    在尺码区间内，删除“数据区（data_start_row..data_end_row）里全为 0/空”的尺码列。
    同步返回更新后的 size_name_list（只保留未被删除的尺码名，顺序不变）。
    - 判定“为 0”：None、空串、'0'、数值 0 都算 0；能转成数字且==0 也算 0。
    - 只检查数据区，不看表头行。
    """
    start_idx = column_index_from_string(start_col_letter)

    # 只考虑有效尺码名（非空）
    names = [n for n in size_name_list if n not in (None, "")]
    delete_offsets = []

    for i, _name in enumerate(names):
        col_idx = start_idx + i
        all_zero = True
        for r in range(data_start_row, data_end_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is None:
                continue
            if isinstance(v, str):
                s = v.strip()
                if s == "" or s == "0":
                    continue
                # 可转数的字符串
                try:
                    if float(s) == 0:
                        continue
                    else:
                        all_zero = False
                        break
                except:
                    # 有非数值文本，认为“非零/有效”
                    all_zero = False
                    break
            else:
                # 数值或其他可转型对象
                try:
                    if float(v) == 0:
                        continue
                    else:
                        all_zero = False
                        break
                except:
                    all_zero = False
                    break
        if all_zero:
            delete_offsets.append(i)

    # 先右到左删除列，避免索引位移
    for i in reversed(delete_offsets):
        ws.delete_cols(start_idx + i, 1)
        del names[i]

    return names  # 返回剩余的尺码名列表（作为后续新的 sizeNames 使用）

        
def fix_header_merges_after_size_columns(ws, size_start_col_letter="G", size_name_list=None, total_size_count=13):
    """
    删除尺码列后，从其下一列开始，将所有列的第7行和第8行重新合并，修复双行表头。
    用于处理“生成生产单”类模板（尺码列从 G 列起）。
    """
    if size_name_list is None:
        size_name_list = []

    size_start_idx = column_index_from_string(size_start_col_letter)
    actual_size_cols = sum(1 for name in size_name_list if name not in ("", None))
    size_end_idx = size_start_idx + actual_size_cols - 1
    merge_start_idx = size_end_idx + 1
    max_col_idx = ws.max_column

    for col_idx in range(merge_start_idx, max_col_idx + 1):
        col_letter = get_column_letter(col_idx)

        # 删除原合并区域（如果还存在）
        for merged_range in list(ws.merged_cells.ranges):
            if f"{col_letter}4:{col_letter}5" in str(merged_range):
                ws.merged_cells.ranges.remove(merged_range)

        # 重新合并第7~8行
        ws.merge_cells(f"{col_letter}4:{col_letter}5")
        cell = ws[f"{col_letter}4"]
        cell.font = cell.font.copy(bold=True)
        
def insert_series_data(wb: Workbook, series_data, col, row, include_price: bool = True):
    ws = wb.active
    grouped_data = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    NORMAL_ROW_HEIGHT = 20
    IMAGE_ROW_HEIGHT = 120

    # Group data: shoe_rid → customer_product_name → color
    for order_shoe_id in series_data:
        title = series_data[order_shoe_id].get("title")
        ws['A1'] = title
        order_rid = series_data[order_shoe_id].get("orderRId")
        ws["B3"] = order_rid
        start_date = series_data[order_shoe_id].get("orderStartDate")
        end_date = series_data[order_shoe_id].get("orderEndDate")
        customer_rid = series_data[order_shoe_id].get("orderCId")
        ws["D3"] = start_date
        ws["I3"] = end_date
        ws["N3"] = customer_rid
        metadata = series_data[order_shoe_id]
        shoe_rid = metadata.get("shoeRId")
        customer_product_name = metadata.get("customerProductName")
        for color_shoe in metadata["shoes"]:
            color = color_shoe.get("color")
            grouped_data[shoe_rid][customer_product_name][color].append({
                "color_name": color_shoe.get("colorName"),
                "img_url": color_shoe.get("imgUrl"),
                "unit_price": color_shoe.get("unitPrice"),
                "packaging_info": color_shoe.get("packagingInfo"),
                "remark": metadata.get("remark")
            })

    for shoe_rid, cust_group in grouped_data.items():
        shoe_start = row
        for cust_name, color_group in cust_group.items():
            cust_start = row
            for color, color_entries in color_group.items():
                color_start = row
                color_name = color_entries[0].get("color_name")
                img_url = color_entries[0].get("img_url")

                row_count_for_color = sum(len(entry["packaging_info"]) for entry in color_entries)
                color_end = row + row_count_for_color - 1

                if row_count_for_color < 6:
                    per_row_height = IMAGE_ROW_HEIGHT / row_count_for_color
                else:
                    per_row_height = NORMAL_ROW_HEIGHT

                color_rows = []  # store rows to set height and image anchor later

                for entry in color_entries:
                    for packaging in entry["packaging_info"]:
                        insert_row_with_format(ws, row, row + 1)
                        ws.row_dimensions[row].height = per_row_height
                        color_rows.append(row)

                        col_idx = column_index_from_string("C")
                        ws[f"{get_column_letter(col_idx)}{row}"] = cust_name
                        col_idx += 1
                        ws[f"{get_column_letter(col_idx)}{row}"] = color
                        col_idx += 1
                        ws[f"{get_column_letter(col_idx)}{row}"] = color_name
                        col_idx += 1
                        ws[f"{get_column_letter(col_idx)}{row}"] = packaging.get("packagingInfoName")
                        col_idx += 1

                        for i in SHOESIZERANGE:
                            if packaging.get(f"size{i}Ratio") == 0:
                                ws[f"{get_column_letter(col_idx)}{row}"] = ""
                            else:
                                # Multiply by count to get the total for this packaging
                                ws[f"{get_column_letter(col_idx)}{row}"] = packaging.get(f"size{i}Ratio")
                            col_idx += 1

                        total_quantity = packaging.get("totalQuantityRatio", 0)
                        count = packaging.get("count", 0)
                        unit_price = entry.get("unit_price", 0)

                        ws[f"{get_column_letter(col_idx)}{row}"] = total_quantity
                        col_idx += 1
                        ws[f"{get_column_letter(col_idx)}{row}"] = count
                        col_idx += 1
                        ws[f"{get_column_letter(col_idx)}{row}"] = total_quantity * count
                        col_idx += 1
                        remark_col_idx = col_idx
                        ws[f"{get_column_letter(col_idx)}{row}"] = entry.get("remark")
                        col_idx += 1
                        currency_format = get_currency_format(metadata.get("currencyType", ""))

                        if include_price:
                            cell_price = ws[f"{get_column_letter(col_idx)}{row}"]
                            cell_price.value = unit_price
                            cell_price.number_format = currency_format
                            col_idx += 1

                            cell_amount = ws[f"{get_column_letter(col_idx)}{row}"]
                            cell_amount.value = unit_price * total_quantity * count
                            cell_amount.number_format = currency_format
                        else:
                            ws[f"{get_column_letter(col_idx)}{row}"] = ""
                            col_idx += 1
                            ws[f"{get_column_letter(col_idx)}{row}"] = ""

                        row += 1

                # Merge columns A (image), D, E
                if row_count_for_color > 1:
                    ws.merge_cells(f"A{color_start}:A{color_end}")
                    ws.merge_cells(f"D{color_start}:D{color_end}")
                    ws.merge_cells(f"E{color_start}:E{color_end}")

                # Insert image only if provided
                if img_url:
                    img_path = os.path.join(IMAGE_UPLOAD_PATH, img_url)
                    if os.path.exists(img_path):
                        img = Image(img_path)
                        img.width = 120
                        img.height = 120

                        # Set height if needed
                        if row_count_for_color < 6:
                            ws.row_dimensions[color_end].height = IMAGE_ROW_HEIGHT

                        # Anchor to top-left of merged cell (A{color_start})
                        cell = f"A{color_start}"
                        img.anchor = cell

                        # Centering: simulate padding manually (only horizontal in openpyxl)
                        # Approximate offsets: depending on cell size
                        try:
                            from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor
                            img.anchor = cell
                            img.drawing.left = units.pixels_to_EMU(30)   # Horizontal center offset
                            img.drawing.top = units.pixels_to_EMU(10)    # Vertical offset (if row height < 120)
                        except Exception:
                            pass  # fallback if the openpyxl version doesn't expose drawing attributes

                        ws.add_image(img)

            # Merge column C (customer_product_name)
            if row - cust_start > 1:
                ws.merge_cells(f"C{cust_start}:C{row - 1}")

        # Merge column B (shoe_rid)
        if row - shoe_start > 1:
            ws.merge_cells(f"B{shoe_start}:B{row - 1}")
            ws.merge_cells(f"{get_column_letter(remark_col_idx)}{shoe_start}:{get_column_letter(remark_col_idx)}{row - 1}")
            merge_ranges.append((shoe_start, row - 1))
        ws[f"B{shoe_start}"] = shoe_rid
        return row - 1
                
def insert_series_data_amount(wb: Workbook, series_data, col, row, include_price: bool = True):
    ws = wb.active
    grouped_data = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    NORMAL_ROW_HEIGHT = 20
    IMAGE_ROW_HEIGHT = 120

    # Group data: shoe_rid → customer_product_name → color
    for order_shoe_id in series_data:
        title = series_data[order_shoe_id].get("title")
        ws['A1'] = title
        order_rid = series_data[order_shoe_id].get("orderRId")
        ws["B3"] = order_rid
        start_date = series_data[order_shoe_id].get("orderStartDate")
        end_date = series_data[order_shoe_id].get("orderEndDate")
        customer_rid = series_data[order_shoe_id].get("orderCId")
        ws["D3"] = start_date
        ws["I3"] = end_date
        ws["N3"] = customer_rid
        metadata = series_data[order_shoe_id]
        shoe_rid = metadata.get("shoeRId")
        customer_product_name = metadata.get("customerProductName")
        for color_shoe in metadata["shoes"]:
            color = color_shoe.get("color")
            grouped_data[shoe_rid][customer_product_name][color].append({
                "color_name": color_shoe.get("colorName"),
                "img_url": color_shoe.get("imgUrl"),
                "unit_price": color_shoe.get("unitPrice"),
                "packaging_info": color_shoe.get("packagingInfo"),
                "remark": metadata.get("remark")
            })

    for shoe_rid, cust_group in grouped_data.items():
        shoe_start = row
        for cust_name, color_group in cust_group.items():
            cust_start = row
            for color, color_entries in color_group.items():
                color_start = row
                color_name = color_entries[0].get("color_name")
                img_url = color_entries[0].get("img_url")

                row_count_for_color = sum(len(entry["packaging_info"]) for entry in color_entries)
                color_end = row + row_count_for_color - 1

                if row_count_for_color < 6:
                    per_row_height = IMAGE_ROW_HEIGHT / row_count_for_color
                else:
                    per_row_height = NORMAL_ROW_HEIGHT

                color_rows = []  # store rows to set height and image anchor later

                for entry in color_entries:
                    for packaging in entry["packaging_info"]:
                        insert_row_with_format(ws, row, row + 1)
                        ws.row_dimensions[row].height = per_row_height
                        color_rows.append(row)
                        count = packaging.get("count", 0)
                        col_idx = column_index_from_string("C")
                        ws[f"{get_column_letter(col_idx)}{row}"] = cust_name
                        col_idx += 1
                        ws[f"{get_column_letter(col_idx)}{row}"] = color
                        col_idx += 1
                        ws[f"{get_column_letter(col_idx)}{row}"] = color_name
                        col_idx += 1
                        ws[f"{get_column_letter(col_idx)}{row}"] = packaging.get("packagingInfoName")
                        col_idx += 1

                        for i in SHOESIZERANGE:
                            if packaging.get(f"size{i}Ratio") == 0:
                                ws[f"{get_column_letter(col_idx)}{row}"] = ""
                            else:
                                # Multiply by count to get the total for this packaging
                                ws[f"{get_column_letter(col_idx)}{row}"] = packaging.get(f"size{i}Ratio") * count
                            col_idx += 1

                        total_quantity = packaging.get("totalQuantityRatio", 0)
                        
                        unit_price = entry.get("unit_price", 0)

                        ws[f"{get_column_letter(col_idx)}{row}"] = total_quantity
                        col_idx += 1
                        ws[f"{get_column_letter(col_idx)}{row}"] = count
                        col_idx += 1
                        ws[f"{get_column_letter(col_idx)}{row}"] = total_quantity * count
                        col_idx += 1
                        remark_col_idx = col_idx
                        ws[f"{get_column_letter(col_idx)}{row}"] = entry.get("remark")
                        col_idx += 1
                        currency_format = get_currency_format(metadata.get("currencyType", ""))

                        if include_price:
                            cell_price = ws[f"{get_column_letter(col_idx)}{row}"]
                            cell_price.value = unit_price
                            cell_price.number_format = currency_format
                            col_idx += 1

                            cell_amount = ws[f"{get_column_letter(col_idx)}{row}"]
                            cell_amount.value = unit_price * total_quantity * count
                            cell_amount.number_format = currency_format
                        else:
                            ws[f"{get_column_letter(col_idx)}{row}"] = ""
                            col_idx += 1
                            ws[f"{get_column_letter(col_idx)}{row}"] = ""

                        row += 1

                # Merge columns A (image), D, E
                if row_count_for_color > 1:
                    ws.merge_cells(f"A{color_start}:A{color_end}")
                    ws.merge_cells(f"D{color_start}:D{color_end}")
                    ws.merge_cells(f"E{color_start}:E{color_end}")

                # Insert image only if provided
                if img_url:
                    img_path = os.path.join(IMAGE_UPLOAD_PATH, img_url)
                    if os.path.exists(img_path):
                        img = Image(img_path)
                        img.width = 120
                        img.height = 120

                        # Set height if needed
                        if row_count_for_color < 6:
                            ws.row_dimensions[color_end].height = IMAGE_ROW_HEIGHT

                        # Anchor to top-left of merged cell (A{color_start})
                        cell = f"A{color_start}"
                        img.anchor = cell

                        # Centering: simulate padding manually (only horizontal in openpyxl)
                        # Approximate offsets: depending on cell size
                        try:
                            from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor
                            img.anchor = cell
                            img.drawing.left = units.pixels_to_EMU(30)   # Horizontal center offset
                            img.drawing.top = units.pixels_to_EMU(10)    # Vertical offset (if row height < 120)
                        except Exception:
                            pass  # fallback if the openpyxl version doesn't expose drawing attributes

                        ws.add_image(img)

            # Merge column C (customer_product_name)
            if row - cust_start > 1:
                ws.merge_cells(f"C{cust_start}:C{row - 1}")

        # Merge column B (shoe_rid)
        
        if row - shoe_start > 1:
            ws.merge_cells(f"B{shoe_start}:B{row - 1}")
            ws.merge_cells(f"{get_column_letter(remark_col_idx)}{shoe_start}:{get_column_letter(remark_col_idx)}{row - 1}")
            merge_ranges.append((shoe_start, row - 1))

        ws[f"B{shoe_start}"] = shoe_rid
        return row - 1


# Function to save the workbook after modification
def save_workbook(wb, new_file_path):
    wb.save(new_file_path)


# Main function to generate the Excel file
def generate_production_excel_file(
    template_path,
    new_file_path,
    order_data: dict,
    metadata: dict,
    include_price: bool = True,
):
    logger.debug(f"Generating Excel file")
    # Load template
    wb = load_template(template_path, new_file_path)
    ws = wb.active
    # Insert the series data
    data_end_row = insert_series_data(wb, order_data, "A", 6, include_price=include_price)
    data_start_row = 6
    trim_trailing_blank_after_data(ws, data_end_row)

    # insert shoe size name
    column = "G"
    row = 5
    for i in range(len(SHOESIZERANGE)):
        ws[f"{column}{row}"] = metadata["sizeNames"][i]
        column = get_next_column_name(column)
    q3_value = ws["Q3"].value
    delete_extra_size_columns(ws, metadata["sizeNames"], start_col_letter="G", total_size_count=len(SHOESIZERANGE))
    fix_header_merges_after_size_columns(ws, size_start_col_letter="G", size_name_list=metadata["sizeNames"])
    add_totals_section(
        ws=ws,
        data_start_row=data_start_row,
        data_end_row=data_end_row,
        currency_type=order_data[next(iter(order_data))].get("currencyType", ""),  # 若你的 currencyType 在 metadata，也可以用 metadata.get("currencyType")
        size_start_col_letter="G",
        size_names=metadata["sizeNames"],
        include_price=include_price,
    )
    cols = get_dynamic_columns(metadata["sizeNames"], "G")
    if not include_price:
        remove_price_amount_columns(ws, cols)
    apply_remark_wrap_and_row_height(ws, cols["remark_letter"], data_start_row, data_end_row)
    remark_col_idx = find_remark_column(ws, header_row=4)
    if remark_col_idx:
        col_letter = get_column_letter(remark_col_idx)
        for start_row, end_row in merge_ranges:
            ws.merge_cells(f"{col_letter}{start_row}:{col_letter}{end_row}")
    merge_ranges.clear()  # 清除已处理的合并范围
    order_shoe_id = next(iter(order_data))  # Get the first key from order_data
    start_date = order_data[order_shoe_id].get("orderStartDate")
    end_date = order_data[order_shoe_id].get("orderEndDate")
    customer_rid = order_data[order_shoe_id].get("orderCId")
    ws["D3"] = start_date
    ws["I3"] = end_date
    ws["N3"] = customer_rid
    ws["Q3"] = q3_value
    set_global_font(ws, font_size=14)  # 这里设置全局字体
    auto_adjust_column_width(ws, max_width=30)
    if include_price:
        ensure_min_width(ws, [cols["price_letter"], cols["amount_letter"]], 14)
    ensure_min_width(ws, [cols["remark_letter"]], 18)  # 备注列稍宽些（可按需改）
    adjust_title_merge(ws, title_row=1, start_col_letter="A")
    adjust_title_merge(ws, title_row=1, start_col_letter="A")  # 调整标题合并范围
    #reset_print_area(ws)
    # Save the workbook
    save_workbook(wb, new_file_path)
    logger.debug(f"Workbook saved as {new_file_path}")
    
def generate_production_amount_excel_file(
    template_path,
    new_file_path,
    order_data: dict,
    metadata: dict,
    include_price: bool = True,
):
    logger.debug(f"Generating Excel file")
    # Load template
    wb = load_template(template_path, new_file_path)
    ws = wb.active
    # Insert the series data
    data_end_row = insert_series_data_amount(wb, order_data, "A", 6, include_price=include_price)
    data_start_row = 6
    trim_trailing_blank_after_data(ws, data_end_row)
    # insert shoe size name
    column = "G"
    row = 5
    for i in range(len(SHOESIZERANGE)):
        ws[f"{column}{row}"] = metadata["sizeNames"][i]
        column = get_next_column_name(column)
    q3_value = ws["Q3"].value
    delete_extra_size_columns(ws, metadata["sizeNames"], start_col_letter="G", total_size_count=len(SHOESIZERANGE))
    fix_header_merges_after_size_columns(ws, size_start_col_letter="G", size_name_list=metadata["sizeNames"])
    add_totals_section(
        ws=ws,
        data_start_row=data_start_row,
        data_end_row=data_end_row,
        currency_type=order_data[next(iter(order_data))].get("currencyType", ""),
        size_start_col_letter="G",
        size_names=metadata["sizeNames"],
        include_price=include_price,
    )
    cols = get_dynamic_columns(metadata["sizeNames"], "G")
    if not include_price:
        remove_price_amount_columns(ws, cols)
    apply_remark_wrap_and_row_height(ws, cols["remark_letter"], data_start_row, data_end_row)
    remark_col_idx = find_remark_column(ws, header_row=4)
    if remark_col_idx:
        col_letter = get_column_letter(remark_col_idx)
        for start_row, end_row in merge_ranges:
            ws.merge_cells(f"{col_letter}{start_row}:{col_letter}{end_row}")
    merge_ranges.clear()  # Clear merge ranges for the next use
    order_shoe_id = next(iter(order_data))  # Get the first key from order_data
    start_date = order_data[order_shoe_id].get("orderStartDate")
    end_date = order_data[order_shoe_id].get("orderEndDate")
    customer_rid = order_data[order_shoe_id].get("orderCId")
    ws["D3"] = start_date
    ws["I3"] = end_date
    ws["N3"] = customer_rid
    ws["Q3"] = q3_value
    set_global_font(ws, font_size=14)  # 这里设置全局字体
    auto_adjust_column_width(ws, max_width=30)
    if include_price:
        ensure_min_width(ws, [cols["price_letter"], cols["amount_letter"]], 14)
    ensure_min_width(ws, [cols["remark_letter"]], 18)  # 备注列稍宽些（可按需改）
    adjust_title_merge(ws, title_row=1, start_col_letter="A")
    adjust_title_merge(ws, title_row=1, start_col_letter="A")  # 调整标题合并范围
    #reset_print_area(ws)
    # Save the workbook
    save_workbook(wb, new_file_path)
    logger.debug(f"Workbook saved as {new_file_path}")
