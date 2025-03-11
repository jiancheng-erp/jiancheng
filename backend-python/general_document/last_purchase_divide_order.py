from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl import load_workbook
import shutil


def load_template(template_path, new_file_path):
    shutil.copy(template_path, new_file_path)
    wb = load_workbook(new_file_path)
    ws = wb.active
    return wb, ws


def add_borders(ws, start_cell, end_cell):
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws[start_cell:end_cell]:
        for cell in row:
            cell.border = border


def format_cells(ws, range_start, range_end, center=True, bold_cells=None):
    """
    Apply formatting to cells:
    - Center alignment for all cells in the range.
    - Auto wrap text.
    - Bold for specific cells.
    """
    for row in ws[range_start:range_end]:
        for cell in row:
            if center:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if bold_cells and cell.coordinate in bold_cells:
                cell.font = Font(bold=True)


def insert_series_data(ws, series_data, start_row=4):
    """
    Insert series data into the worksheet.
    - C4-J4: First 8 sizes for the first item (bold).
    - C5-J5: Corresponding amounts for the first 8 sizes (and sum in K5).
    - C6-J6: Remaining sizes if size count > 8 (bold).
    - C7-J7: Corresponding amounts for the remaining sizes (and sum in K7).
    - Automatically calculate `合计` for each amount row.
    - Merge `备注` field for each item.
    """
    current_row = start_row  # Start at row 4

    for item in series_data:
        sizes = [key for key in item.keys() if key not in ("物品名称", "型号", "类别", "合计", "备注")]
        size_chunks = [sizes[x:x+8] for x in range(0, len(sizes), 8)]  # Break sizes into chunks of 8

        # Insert sizes and amounts
        for chunk in size_chunks:
            # Insert sizes in the current row
            column = "C"
            for size in chunk:
                ws[f"{column}{current_row}"] = size
                ws[f"{column}{current_row}"].font = Font(bold=True)  # Make sizes bold
                column = get_next_column_name(column)

            # Insert corresponding amounts in the next row
            current_row += 1  # Move to the next row for amounts
            column = "C"
            row_total = 0  # Track total for this row

            for size in chunk:
                value = item.get(size, 0)  # Get quantity, default to 0
                ws[f"{column}{current_row}"] = value
                row_total += int(value)  # Sum row values
                column = get_next_column_name(column)

            # Insert row total in column K
            ws[f"K{current_row}"] = row_total
            ws[f"K{current_row}"].font = Font(bold=True)  # Bold the total

            current_row += 1  # Prepare for the next chunk

        # Merge `备注` field
        merge_start_row = current_row - len(size_chunks) * 2  # Calculate merge start
        ws.merge_cells(start_row=merge_start_row, start_column=1, end_row=current_row - 1, end_column=1)
        ws.merge_cells(start_row=merge_start_row, start_column=2, end_row=current_row - 1, end_column=2)
        ws[f"A{merge_start_row}"] = item.get("型号", "")  # Set value in merged cell
        ws[f"B{merge_start_row}"] = item.get("类别", "")  # Set value in merged cell

        ws.merge_cells(start_row=merge_start_row + 1, start_column=12, end_row=current_row - 1, end_column=12)
        ws[f"L{merge_start_row + 1}"] = item.get("备注", "")  # Set `备注` value in merged cell

    return current_row



def get_next_column_name(current_column_name):
    column_index = column_index_from_string(current_column_name)
    next_column_index = column_index + 1
    next_column_name = get_column_letter(next_column_index)
    return next_column_name


def generate_last_excel_file(template_path, new_file_path, order_data):
    wb, ws = load_template(template_path, new_file_path)

    # Insert order details
    ws["C2"] = order_data.get("订单信息", "")
    ws["B2"] = order_data.get("供应商", "")
    ws["H2"] = order_data.get("日期", "")
    ws["K2"] = order_data.get("客户名", "")

    # Insert series data
    row = insert_series_data(ws, order_data.get("seriesData", []))

    # Fill summary fields (below data table)
    ws[f"A{row + 1}"] = "合计"
    # 合计 = sum of all `合计` values insert computed in the insert_series_data function
    total_sum = sum(ws[f"K{r}"].value for r in range(5, row, 2) if ws[f"K{r}"].value)
    ws[f"K{row + 1}"] = total_sum
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    ws[f"A{row}"] = order_data.get("备注", "")
    ws[f"A{row + 2}"] = "环境要求:"
    ws[f"B{row + 2}"] = order_data.get("环保要求", "")
    ws[f"A{row + 3}"] = "发货地址:"
    ws[f"B{row + 3}"] = order_data.get("发货地址", "")
    ws[f"A{row + 4}"] = "交货期限:"
    ws[f"B{row + 4}"] = order_data.get("交货期限", "")
    ws[f"G{row + 4}"] = "如有特殊情况提前5天反馈，无故延期有贵公司承担后续责任。"
    ws[f"A{row + 5}"] = "制表:"
    ws[f"G{row + 5}"] = "审核:"

    # Apply formatting to only data region
    bold_cells = {"A2", "B2", "F2", "A3", "C3", "K3", "L3"}
    format_cells(ws, "A1", f"L{row - 1}", center=True, bold_cells=bold_cells)  # 只居中数据部分

    # Add borders
    add_borders(ws, "A1", f"L{row - 1}")  # 只添加边框到数据部分

    wb.save(new_file_path)
    print(f"Workbook saved as {new_file_path}")


# template_path = "H:\git-projects\jiancheng\\backend-python\general_document\新标准采购订单尺码版.xlsx"
# new_file_path = "H:\git-projects\jiancheng\\backend-python\general_document\size_test.xlsx"

# order_data = {
#     "订单信息": "订单编号12345",
#     "供应商": "供应商A",
#     "日期": "2024-11-19",
#     "seriesData": [
#         {"物品名称": "商品2", "型号":"type1", "类别":"type1" ,"7": 10, "7.5": 20, "8": 30, "8.5": 40, "9": 50, "10": 60, "11": 70, "12":100,"13":150, "合计": 280, "备注": "无"},
#         {"物品名称": "商品2", "型号":"type1", "类别":"type1", "7": 10, "7.5": 20, "8": 30, "8.5": 40, "9": 50, "10": 60, "11": 70, "12":100,"13":150, "合计": 280, "备注": "无"},
#         # {"物品名称": "商品2", "35": 15, "36": 25, "37": 35, "38": 45, "39": 55, "40": 65, "41": 75, "合计": 315, "备注": "无"},
#     ],
#     "合计": 595,
#     "备注": "测试备注",
#     "环保要求": "符合环保标准",
#     "发货地址": "测试地址",
#     "交货期限": "2024-12-01",
# }

# generate_size_excel_file(template_path, new_file_path, order_data)
