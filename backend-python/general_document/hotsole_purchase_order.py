from calendar import c
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl import load_workbook
import shutil
import math
from logger import logger
def set_wrapped_cell(ws, cell_address, text, center=True, column_width_chars=40, line_height=15, min_height=30):
    """
    设置自动换行、估算行高的单元格。

    参数：
    - ws: 工作表对象
    - cell_address: 如 'A10'
    - text: 填入内容
    - center: 是否居中（默认居中）
    - column_width_chars: 每行大概容纳多少字符（默认 40）
    - line_height: 每行对应的高度（默认 15）
    - min_height: 最小行高
    """
    ws[cell_address] = text

    # 设置对齐方式
    alignment = Alignment(wrap_text=True)
    if center:
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[cell_address].alignment = alignment

    # 获取行号
    row = int(''.join(filter(str.isdigit, cell_address)))

    # 估算换行行数（实际无换行符）
    num_lines = math.ceil(len(str(text)) / column_width_chars)
    height = max(min_height, num_lines * line_height)

    # 设置行高
    ws.row_dimensions[row].height = height + 10


craft_whole_name = "1."
craft_series_number = 1

def load_template(template_path, new_file_path):
    shutil.copy(template_path, new_file_path)
    wb = load_workbook(new_file_path)
    ws = wb.active or wb.worksheets[0]
    wb.active = ws
    return wb, ws


def add_borders(ws, start_cell, end_cell):
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws[start_cell:end_cell]:
        for cell in row:
            cell.border = border


def format_cells(ws, range_start, range_end, center=True, bold_cells=None):
    """
    Apply formatting to cells:
    - Center alignment for all cells in the range.
    - Auto wrap text.
    - Bold for specific cells.
    """
    for row in ws[range_start:range_end]:
        for cell in row:
            if center:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if bold_cells and cell.coordinate in bold_cells:
                    cell.font = Font(bold=True, size=12)

def insert_series_data(ws, series_data, start_row=7):
    """
    Insert series data into the worksheet.
    - C4-J4: First 8 sizes for the first item (bold).
    - C5-J5: Corresponding amounts for the first 8 sizes (and sum in K5).
    - C6-J6: Remaining sizes if size count > 8 (bold).
    - C7-J7: Corresponding amounts for the remaining sizes (and sum in K7).
    - Automatically calculate `合计` for each amount row.
    - Merge `备注` field for each item.
    """
    global craft_whole_name, craft_series_number
    current_row = start_row  # Start at row 4

    for item in series_data:
        sizes = [key for key in item.keys() if key not in ("物品名称", "型号", "类别", "合计", "备注", "工厂型号", "鞋面颜色", "工艺说明", "使用材料")]
        size_chunks = [sizes[x:x+8] for x in range(0, len(sizes), 8)]  # Break sizes into chunks of 8
        # 去除各字段内部回车，编号间回车保留
        _mat = (item.get("使用材料") or "").replace("\r\n", " ").replace("\n", " ").replace("\r", " ").strip()
        _craft = (item.get("工艺说明") or "").replace("\r\n", " ").replace("\n", " ").replace("\r", " ").strip()
        _entry = _mat + _craft
        if craft_whole_name == "1.":
            craft_whole_name = "1." + _entry
            craft_series_number = craft_series_number + 1
        else:
            craft_whole_name = craft_whole_name + "\n" + str(craft_series_number) + "." + _entry
            craft_series_number = craft_series_number + 1
            

        # Insert sizes and amounts
        for chunk in size_chunks:
            # Insert sizes in the current row
            column = "C"
            for size in chunk:
                ws[f"{column}{current_row}"] = size
                ws[f"{column}{current_row}"].font = Font(bold=True, size=12)  # Make sizes bold
                column = get_next_column_name(column)

            # Insert corresponding amounts in the next row
            current_row += 1  # Move to the next row for amounts
            column = "C"
            row_total = 0  # Track total for this row

            for size in chunk:
                value = item.get(size, 0)  # Get quantity, default to 0
                ws[f"{column}{current_row}"] = value if value else None  # 数量为0时留空
                row_total += int(value)  # Sum row values
                column = get_next_column_name(column)

            # Insert row total in column K
            ws[f"K{current_row}"] = row_total if row_total else None
            ws[f"K{current_row}"].font = Font(size=12)

            current_row += 1  # Prepare for the next chunk

        # Merge `备注` field
        merge_start_row = current_row - len(size_chunks) * 2  # Calculate merge start
        ws.merge_cells(start_row=merge_start_row, start_column=1, end_row=current_row - 1, end_column=1)
        ws.merge_cells(start_row=merge_start_row, start_column=2, end_row=current_row - 1, end_column=2)
        ws[f"A{merge_start_row}"] = item.get("工厂型号", "")  # Set value in merged cell
        ws[f"B{merge_start_row}"] = item.get("鞋面颜色", "")  # Set value in merged cell

        ws.merge_cells(start_row=merge_start_row + 1, start_column=12, end_row=current_row - 1, end_column=12)
        ws[f"L{merge_start_row + 1}"] = item.get("备注", "")  # Set `备注` value in merged cell

    return current_row



def get_next_column_name(current_column_name):
    column_index = column_index_from_string(current_column_name)
    next_column_index = column_index + 1
    next_column_name = get_column_letter(next_column_index)
    return next_column_name


def generate_hotsole_excel_file(template_path, new_file_path, order_data):
    global craft_whole_name, craft_series_number
    craft_whole_name = "1."
    craft_series_number = 1
    logger.debug("start generate_last_excel_file")
    wb, ws = load_template(template_path, new_file_path)

    # 删除模板中无用的第2、3行（删后行号整体-2）
    ws.delete_rows(2, 2)

    # 清除第4行（表头行）M列及以后的多余内容，隐藏 M-Z 列
    for col in range(13, 27):
        ws.cell(row=4, column=col).value = None
        ws.column_dimensions[get_column_letter(col)].hidden = True

    # 重置第4行表头：工厂型号 | 鞋面颜色 | 尺码(C-J合并) | 合计 | 备注
    thin = Side(border_style="thin", color="000000")
    hdr_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    hdr_bold = Font(bold=True, size=12)

    # 先清除A4-L4所有合并和内容
    for col in range(1, 13):
        cell = ws.cell(row=4, column=col)
        cell.value = None
        cell.border = hdr_border
        cell.alignment = hdr_center
        cell.font = hdr_bold

    # 取消可能存在的合并
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row == 4 and merged.max_row == 4:
            ws.unmerge_cells(str(merged))

    ws["A4"] = "工厂型号"
    ws["B4"] = "鞋面颜色"
    # C4:J4 合并为"尺码"
    ws.merge_cells("C4:J4")
    ws["C4"] = "尺码"
    ws["C4"].border = hdr_border
    ws["C4"].alignment = hdr_center
    ws["C4"].font = hdr_bold
    ws["K4"] = "合计"
    ws["K4"].border = hdr_border
    ws["K4"].alignment = hdr_center
    ws["K4"].font = hdr_bold
    ws["L4"] = "备注"
    ws["L4"].border = hdr_border
    ws["L4"].alignment = hdr_center
    ws["L4"].font = hdr_bold

    # Insert order details（原第4行 → 现第2行）
    ws["H2"] = order_data.get("订单信息", "") + " " + order_data.get("客户名", "") + " " + order_data.get("商标", "")
    ws["B2"] = order_data.get("供应商", "")

    # Insert series data（原start_row=7 → 现5）
    row = insert_series_data(ws, order_data.get("seriesData", []), start_row=5)

    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Fill summary fields (below data table)
    ws[f"A{row + 1}"] = "合计"
    total_sum = sum(ws[f"K{r}"].value for r in range(6, row, 2) if ws[f"K{r}"].value)
    ws[f"K{row + 1}"] = total_sum
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    ws[f"A{row}"] = order_data.get("备注", "")

    ws.merge_cells(start_row=row + 2, start_column=1, end_row=row + 2, end_column=12)
    set_wrapped_cell(ws, f"A{row + 2}", craft_whole_name, center=False)

    ws[f"A{row + 3}"] = "环境要求:"
    ws.merge_cells(start_row=row + 3, start_column=2, end_row=row + 3, end_column=12)
    ws[f"B{row + 3}"] = order_data.get("环保要求", "")
    ws[f"B{row + 3}"].alignment = left_wrap

    ws[f"A{row + 4}"] = "发货地址:"
    ws.merge_cells(start_row=row + 4, start_column=2, end_row=row + 4, end_column=12)
    ws[f"B{row + 4}"] = order_data.get("发货地址", "")
    ws[f"B{row + 4}"].alignment = left_wrap

    ws[f"A{row + 5}"] = "交货期限:"
    ws.merge_cells(start_row=row + 5, start_column=2, end_row=row + 5, end_column=12)
    deadline_text = (order_data.get("交货期限", "") + "    如有特殊情况提前5天反馈，无故延期有贵公司承担后续责任。").strip()
    ws[f"B{row + 5}"] = deadline_text
    ws[f"B{row + 5}"].alignment = left_wrap
    ws.row_dimensions[row + 5].height = 30

    ws[f"A{row + 6}"] = "制表:"
    ws[f"G{row + 6}"] = "审核:"
    ws[f"G{row + 7}"] = order_data.get("日期", "")

    # Apply formatting to only data region（原A7→A5，原A6→A4）
    bold_cells = {"A2", "A4", "B4", "K4", "L4"}
    format_cells(ws, "A5", f"L{row - 1}", center=True, bold_cells=bold_cells)

    # Add borders（原A6→A4）
    add_borders(ws, "A4", f"L{row - 1}")

    # 列宽：A=工厂型号15，B=鞋面颜色15，C-J=数量列6，K=合计8，L=备注15
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    for col in range(3, 11):  # C~J
        ws.column_dimensions[get_column_letter(col)].width = 6
    ws.column_dimensions["K"].width = 8
    ws.column_dimensions["L"].width = 15

    # 页面设置：横向打印，宽度适应1页
    if ws.sheet_properties is None:
        ws.sheet_properties = WorksheetProperties()
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # 全局字号 12
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            if cell.font and cell.font.size and cell.font.size != 12:
                cell.font = cell.font.copy(size=12)
            elif not cell.font or not cell.font.size:
                cell.font = cell.font.copy(size=12) if cell.font else Font(size=12)

    wb.save(new_file_path)
    logger.debug(f"Workbook saved as {new_file_path}")


# template_path = "D:\catSupermarket\jiancheng\\backend-python\general_document\烫底标准采购订单.xlsx"
# new_file_path = "D:\catSupermarket\jiancheng\\backend-python\general_document\hotsole_test.xlsx"

# order_data = {
#     "订单信息": "K25-301",
#     "供应商": "金乡加工",
#     "客户名": "37",
#     "商标": "REFRESH",
#     "日期": "2024-11-19",
#     "seriesData": [
#         {"工厂型号":"C21072M1","鞋面颜色":"米色" ,"7": 10, "7.5": 20, "8": 30, "8.5": 40, "9": 50, "10": 60, "11": 70, "12":100,"13":150, "合计": 280, "备注": "无","工艺说明": "1501-10卡其色复90密度5*9蓝色成型印白色Refresh客人商标"},
#         {"工厂型号": "C21072M1", "鞋面颜色":"黑色", "7": 10, "7.5": 20, "8": 30, "8.5": 40, "9": 50, "10": 60, "11": 70, "12":100,"13":150, "合计": 280, "备注": "无","工艺说明": "1501-10卡其色复90密度5*9蓝色成型印白色Refresh客人商标"},
#         # {"物品名称": "商品2", "35": 15, "36": 25, "37": 35, "38": 45, "39": 55, "40": 65, "41": 75, "合计": 315, "备注": "无"},
#     ],
#     "合计": 595,
#     "备注": "测试备注",
#     "环保要求": "符合环保标准",
#     "发货地址": "测试地址",
#     "交货期限": "2024-12-01",
# }

# generate_hotsole_excel_file(template_path, new_file_path, order_data)
