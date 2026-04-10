import shutil
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.drawing.image import Image
from constants import SHOESIZERANGE
from file_locations import IMAGE_UPLOAD_PATH
import os
from collections import defaultdict
from openpyxl.utils import units
from logger import logger
from openpyxl.styles import Font

def get_currency_format(currency_type: str, decimals: int = 3) -> str:
    """
    返回 openpyxl 的 number_format，例如 '"$"#,##0.000'
    未匹配到币种时不加符号，仅保留小数位。
    """
    symbol_map = {
        "USD": "$",
        "USA": "$",
        "CNY": "￥",
        "RMB": "￥",
        "EUR": "€",
    }
    symbol = symbol_map.get((currency_type or "").upper(), "")
    dec = "0" * decimals
    return f'"{symbol}"#,##0.{dec}' if symbol else f'#,##0.{dec}'

# Function to load the Excel template and prepare for modification
def load_template(template_path, new_file_path):
    # Copy the template to a new file
    shutil.copy(template_path, new_file_path)
    # Load the new workbook
    wb = load_workbook(new_file_path)
    return wb


def get_next_column_name(current_column_name):
    # Convert column letter to index
    column_index = column_index_from_string(current_column_name)
    # Increment the index to get the next column
    next_column_index = column_index + 1
    # Convert the new index back to a column letter
    next_column_name = get_column_letter(next_column_index)
    return next_column_name


# Function to copy formatting and insert a new row
def insert_row_with_format(ws, row_to_copy, new_row_idx):
    # Insert a new row
    ws.insert_rows(new_row_idx)
    ws.row_dimensions[new_row_idx].height = ws.row_dimensions[row_to_copy].height
    # Copy the formatting
    for col_idx, cell in enumerate(ws[row_to_copy], start=1):
        new_cell = ws.cell(row=new_row_idx, column=col_idx)
        if cell.has_style:
            new_cell.border = cell.border.copy()
            new_cell.alignment = cell.alignment.copy()
            new_cell.number_format = cell.number_format
            
def delete_extra_size_columns(ws, size_name_list, start_col_letter="E", total_size_count=13):
    """
    删除从 start_col_letter 开始的尺码列，只保留非空名称对应的列。
    size_name_list: 尺码名称列表（可能含 '', None）
    total_size_count: 模板中总共预留了多少尺码列（默认 13）
    """
    start_col_idx = column_index_from_string(start_col_letter)
    actual_size_cols = sum(1 for name in size_name_list if name not in ("", None))
    extra_count = total_size_count - actual_size_cols
    if extra_count > 0:
        delete_start_col = start_col_idx + actual_size_cols
        ws.delete_cols(delete_start_col, extra_count)
    for merged_range in list(ws.merged_cells.ranges):
        if f"{start_col_letter}7" in str(merged_range):
            ws.merged_cells.ranges.remove(merged_range)

    if actual_size_cols > 0:
        end_col_letter = get_column_letter(start_col_idx + actual_size_cols - 1)
        merge_range = f"{start_col_letter}7:{end_col_letter}7"
        ws.merge_cells(merge_range)
        title_cell = ws[f"{start_col_letter}7"]
        title_cell.value = "尺码"
        title_cell.font = Font(bold=True)
def fix_header_merges_after_size_columns(ws, size_start_col_letter="E", size_name_list=None, total_size_count=13):
    """
    删除尺码列后，从其下一列开始，将所有列的第7行和第8行重新合并，修复双行表头。
    """
    if size_name_list is None:
        size_name_list = []

    size_start_idx = column_index_from_string(size_start_col_letter)
    actual_size_cols = sum(1 for name in size_name_list if name not in ("", None))
    size_end_idx = size_start_idx + actual_size_cols - 1

    # 下一列就是合并起始列
    merge_start_idx = size_end_idx + 1
    max_col_idx = ws.max_column

    for col_idx in range(merge_start_idx, max_col_idx + 1):
        col_letter = get_column_letter(col_idx)

        # 删除原本在这一列的 7~8 行合并区域（如果存在）
        for merged_range in list(ws.merged_cells.ranges):
            if f"{col_letter}7:{col_letter}8" in str(merged_range):
                ws.merged_cells.ranges.remove(merged_range)

        # 合并新的 7~8 行
        ws.merge_cells(f"{col_letter}7:{col_letter}8")
        cell = ws[f"{col_letter}7"]
        cell.font = Font(bold=True)

        # 可选：继承对齐样式（假设 E7 是模板参考）
        if "E7" in ws:
            cell.alignment = ws["E7"].alignment

def insert_series_data(wb: Workbook, series_data, col, row):
    all_size_names = []
    ws = wb.active
    grouped_data = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    NORMAL_ROW_HEIGHT = 20
    IMAGE_ROW_HEIGHT = 120

    # Group data: shoe_rid → customer_product_name → color
    for order_shoe_id in series_data:
        metadata = series_data[order_shoe_id]
        shoe_rid = metadata.get("shoeRId")
        customer_product_name = metadata.get("customerProductName")
        for color_shoe in metadata["shoes"]:
            color = color_shoe.get("colorName")
            grouped_data[shoe_rid][customer_product_name][color].append({
                "color_name": color_shoe.get("colorName"),
                "img_url": color_shoe.get("imgUrl"),
                "unit_price": color_shoe.get("unitPrice"),
                "packaging_info": color_shoe.get("packagingInfo"),
                "remark": metadata.get("remark"),
                "currency_type": metadata.get("currencyType"),
            })

    first_customer_shoe_written = False  # 标记是否写过第一个鞋型

    for shoe_rid, cust_group in grouped_data.items():
        shoe_start = row
        for cust_name, color_group in cust_group.items():
            cust_start = row

            # 🟩 获取当前鞋型的尺码名（从任意颜色-包装里取即可）
            size_names = []
            found = False
            for color_entries in color_group.values():
                for entry in color_entries:
                    for packaging in entry["packaging_info"]:
                        size_names = packaging.get("sizeNames", [])
                        found = True
                        break
                    if found:
                        break
                if found:
                    break
            if len([s for s in size_names if s not in ("", None)]) > len([s for s in all_size_names if s not in ("", None)]):
                all_size_names = size_names  # 记录下最长的那组有效尺码名

            # 🟨 写入尺码行（第一个鞋型写到第 8 行，其余插入新行）
            if not first_customer_shoe_written:
                temp_column = column_index_from_string("E")
                for name in size_names:
                    cell = ws[f"{get_column_letter(temp_column)}8"]
                    cell.value = name
                    cell.font = Font(bold=True)  # 🟩 加粗
                    temp_column += 1
                merge_start_row = cust_start
                first_customer_shoe_written = True
            else:
                insert_row_with_format(ws, row, row + 1)
                ws.row_dimensions[row].height = NORMAL_ROW_HEIGHT
                ws[f"D{row}"] = "尺码"
                ws[f"D{row}"].font = Font(bold=True)  # 🟩 “尺码” 也加粗
                temp_column = column_index_from_string("E")
                for name in size_names:
                    cell = ws[f"{get_column_letter(temp_column)}{row}"]
                    cell.value = name
                    cell.font = Font(bold=True)  # 🟩 加粗
                    temp_column += 1
                merge_start_row = row + 1  # 记录尺码行位置
                row += 1  # advance to next row
            for color, color_entries in color_group.items():
                color_start = row
                color_name = color_entries[0].get("color_name")
                img_url = color_entries[0].get("img_url")

                row_count_for_color = sum(len(entry["packaging_info"]) for entry in color_entries)
                color_end = row + row_count_for_color - 1
                if row_count_for_color < 6:
                    per_row_height = IMAGE_ROW_HEIGHT / row_count_for_color
                else:
                    per_row_height = NORMAL_ROW_HEIGHT

                color_rows = []  # store rows to set height and image anchor later

                for entry in color_entries:
                    for packaging in entry["packaging_info"]:
                        insert_row_with_format(ws, row, row + 1)
                        ws.row_dimensions[row].height = per_row_height
                        color_rows.append(row)

                        col_idx = column_index_from_string("B")
                        ws[f"{get_column_letter(col_idx)}{row}"] = cust_name
                        col_idx += 1
                        ws[f"{get_column_letter(col_idx)}{row}"] = color
                        col_idx += 1
                        ws[f"{get_column_letter(col_idx)}{row}"] = packaging.get("packagingInfoName")
                        col_idx += 1

                        for i in SHOESIZERANGE:
                            if packaging.get(f"size{i}Ratio") == 0:
                                ws[f"{get_column_letter(col_idx)}{row}"] = ""
                            else:
                                # Multiply by count to get the total for this packaging
                                ws[f"{get_column_letter(col_idx)}{row}"] = packaging.get(f"size{i}Ratio")
                            col_idx += 1

                        total_quantity = packaging.get("totalQuantityRatio", 0)
                        count = packaging.get("count", 0)
                        unit_price = entry.get("unit_price", 0)

                        ws[f"{get_column_letter(col_idx)}{row}"] = total_quantity
                        col_idx += 1
                        ws[f"{get_column_letter(col_idx)}{row}"] = count
                        col_idx += 1
                        ws[f"{get_column_letter(col_idx)}{row}"] = total_quantity * count
                        col_idx += 1
                        currency_fmt = get_currency_format(entry.get("currency_type"), decimals=2)

                        cell_price = ws[f"{get_column_letter(col_idx)}{row}"]
                        cell_price.value = unit_price
                        cell_price.number_format = currency_fmt
                        col_idx += 1

                        cell_amount = ws[f"{get_column_letter(col_idx)}{row}"]
                        cell_amount.value = unit_price * total_quantity * count
                        cell_amount.number_format = currency_fmt

                        row += 1

                # Merge columns A (image), C
                if row_count_for_color > 1:
                    ws.merge_cells(f"A{color_start}:A{color_end}")
                    ws.merge_cells(f"C{color_start}:C{color_end}")

                # Insert image only if provided
                if img_url:
                    img_path = os.path.join(IMAGE_UPLOAD_PATH, img_url)
                    if os.path.exists(img_path):
                        img = Image(img_path)
                        img.width = 120
                        img.height = 120

                        # Set height if needed
                        if row_count_for_color < 6:
                            ws.row_dimensions[color_end].height = IMAGE_ROW_HEIGHT

                        # Anchor to top-left of merged cell (A{color_start})
                        cell = f"A{color_start}"
                        img.anchor = cell

                        # Centering: simulate padding manually (only horizontal in openpyxl)
                        # Approximate offsets: depending on cell size
                        try:
                            from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor
                            img.anchor = cell
                            img.drawing.left = units.pixels_to_EMU(30)   # Horizontal center offset
                            img.drawing.top = units.pixels_to_EMU(10)    # Vertical offset (if row height < 120)
                        except Exception:
                            pass  # fallback if the openpyxl version doesn't expose drawing attributes

                        ws.add_image(img)

            # Merge column C (customer_product_name)
            if row - merge_start_row > 1:
                ws.merge_cells(f"B{merge_start_row}:B{row - 1}")
    delete_extra_size_columns(ws, all_size_names)
    fix_header_merges_after_size_columns(ws, size_start_col_letter="E", size_name_list=all_size_names)
                
def insert_series_data_amount(wb: Workbook, series_data, col, row):
    all_size_names = []
    ws = wb.active
    grouped_data = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    NORMAL_ROW_HEIGHT = 20
    IMAGE_ROW_HEIGHT = 120

    # Group data: shoe_rid → customer_product_name → color
    for order_shoe_id in series_data:
        metadata = series_data[order_shoe_id]
        shoe_rid = metadata.get("shoeRId")
        customer_product_name = metadata.get("customerProductName")
        for color_shoe in metadata["shoes"]:
            color = color_shoe.get("colorName")
            grouped_data[shoe_rid][customer_product_name][color].append({
                "color_name": color_shoe.get("colorName"),
                "img_url": color_shoe.get("imgUrl"),
                "unit_price": color_shoe.get("unitPrice"),
                "packaging_info": color_shoe.get("packagingInfo"),
                "remark": metadata.get("remark"),
                "currency_type": metadata.get("currencyType"),
            })

    first_customer_shoe_written = False  # 标记是否写过第一个鞋型

    for shoe_rid, cust_group in grouped_data.items():
        shoe_start = row
        for cust_name, color_group in cust_group.items():
            cust_start = row

            # 🟩 获取当前鞋型的尺码名（从任意颜色-包装里取即可）
            size_names = []
            found = False
            for color_entries in color_group.values():
                for entry in color_entries:
                    for packaging in entry["packaging_info"]:
                        size_names = packaging.get("sizeNames", [])
                        found = True
                        break
                    if found:
                        break
                if found:
                    break
            if len([s for s in size_names if s not in ("", None)]) > len([s for s in all_size_names if s not in ("", None)]):
                all_size_names = size_names  # 记录下最长的那组有效尺码名

            # 🟨 写入尺码行（第一个鞋型写到第 8 行，其余插入新行）
            if not first_customer_shoe_written:
                temp_column = column_index_from_string("E")
                for name in size_names:
                    cell = ws[f"{get_column_letter(temp_column)}8"]
                    cell.value = name
                    cell.font = Font(bold=True)  # 🟩 加粗
                    temp_column += 1
                merge_start_row = cust_start
                first_customer_shoe_written = True
            else:
                insert_row_with_format(ws, row, row + 1)
                ws.row_dimensions[row].height = NORMAL_ROW_HEIGHT
                ws[f"D{row}"] = "尺码"
                ws[f"D{row}"].font = Font(bold=True)  # 🟩 “尺码” 也加粗
                temp_column = column_index_from_string("E")
                for name in size_names:
                    cell = ws[f"{get_column_letter(temp_column)}{row}"]
                    cell.value = name
                    cell.font = Font(bold=True)  # 🟩 加粗
                    temp_column += 1
                merge_start_row = row + 1  # 记录尺码行位置
                row += 1  # advance to next row
            for color, color_entries in color_group.items():
                color_start = row
                color_name = color_entries[0].get("color_name")
                img_url = color_entries[0].get("img_url")

                row_count_for_color = sum(len(entry["packaging_info"]) for entry in color_entries)
                color_end = row + row_count_for_color - 1
                if row_count_for_color < 6:
                    per_row_height = IMAGE_ROW_HEIGHT / row_count_for_color
                else:
                    per_row_height = NORMAL_ROW_HEIGHT

                color_rows = []  # store rows to set height and image anchor later

                for entry in color_entries:
                    for packaging in entry["packaging_info"]:
                        insert_row_with_format(ws, row, row + 1)
                        ws.row_dimensions[row].height = per_row_height
                        color_rows.append(row)
                        count = packaging.get("count", 0)

                        col_idx = column_index_from_string("B")
                        ws[f"{get_column_letter(col_idx)}{row}"] = cust_name
                        col_idx += 1
                        ws[f"{get_column_letter(col_idx)}{row}"] = color
                        col_idx += 1
                        ws[f"{get_column_letter(col_idx)}{row}"] = packaging.get("packagingInfoName")
                        col_idx += 1

                        for i in SHOESIZERANGE:
                            if packaging.get(f"size{i}Ratio") == 0:
                                ws[f"{get_column_letter(col_idx)}{row}"] = ""
                            else:
                                ws[f"{get_column_letter(col_idx)}{row}"] = packaging.get(f"size{i}Ratio") * count
                            col_idx += 1

                        total_quantity = packaging.get("totalQuantityRatio", 0)
                        
                        unit_price = entry.get("unit_price", 0)

                        ws[f"{get_column_letter(col_idx)}{row}"] = total_quantity
                        col_idx += 1
                        ws[f"{get_column_letter(col_idx)}{row}"] = count
                        col_idx += 1
                        ws[f"{get_column_letter(col_idx)}{row}"] = total_quantity * count
                        col_idx += 1
                        currency_fmt = get_currency_format(entry.get("currency_type"), decimals=2)

                        cell_price = ws[f"{get_column_letter(col_idx)}{row}"]
                        cell_price.value = unit_price
                        cell_price.number_format = currency_fmt
                        col_idx += 1

                        cell_amount = ws[f"{get_column_letter(col_idx)}{row}"]
                        cell_amount.value = unit_price * total_quantity * count
                        cell_amount.number_format = currency_fmt

                        row += 1

                # Merge columns A (image), C
                if row_count_for_color > 1:
                    ws.merge_cells(f"A{color_start}:A{color_end}")
                    ws.merge_cells(f"C{color_start}:C{color_end}")

                # Insert image only if provided
                if img_url:
                    img_path = os.path.join(IMAGE_UPLOAD_PATH, img_url)
                    if os.path.exists(img_path):
                        img = Image(img_path)
                        img.width = 120
                        img.height = 120

                        # Set height if needed
                        if row_count_for_color < 6:
                            ws.row_dimensions[color_end].height = IMAGE_ROW_HEIGHT

                        # Anchor to top-left of merged cell (A{color_start})
                        cell = f"A{color_start}"
                        img.anchor = cell

                        # Centering: simulate padding manually (only horizontal in openpyxl)
                        # Approximate offsets: depending on cell size
                        try:
                            from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor
                            img.anchor = cell
                            img.drawing.left = units.pixels_to_EMU(30)   # Horizontal center offset
                            img.drawing.top = units.pixels_to_EMU(10)    # Vertical offset (if row height < 120)
                        except Exception:
                            pass  # fallback if the openpyxl version doesn't expose drawing attributes

                        ws.add_image(img)

            # Merge column C (customer_product_name)
            if row - merge_start_row > 1:
                ws.merge_cells(f"B{merge_start_row}:B{row - 1}")
    delete_extra_size_columns(ws, all_size_names)
    fix_header_merges_after_size_columns(ws, size_start_col_letter="E", size_name_list=all_size_names)


# Function to save the workbook after modification
def save_workbook(wb, new_file_path):
    wb.save(new_file_path)


# Main function to generate the Excel file
def generate_excel_file(template_path, new_file_path, order_data: dict, metadata: dict):
    logger.debug(f"Generating Excel file")
    # Load template
    wb = load_template(template_path, new_file_path)
    ws = wb.active
    # Insert the series data
    insert_series_data(wb, order_data, "A", 9)

    # insert shoe size name
    # column = "E"
    # row = 8
    # for i in range(len(SHOESIZERANGE)):
    #     ws[f"{column}{row}"] = metadata["sizeNames"][i]
    #     column = get_next_column_name(column)

    # Save the workbook
    save_workbook(wb, new_file_path)
    logger.debug(f"Workbook saved as {new_file_path}")
    
def generate_amount_excel_file(template_path, new_file_path, order_data: dict, metadata: dict):
    logger.debug(f"Generating Excel file")
    # Load template
    wb = load_template(template_path, new_file_path)
    ws = wb.active
    # Insert the series data
    insert_series_data_amount(wb, order_data, "A", 9)

    # insert shoe size name
    # column = "E"
    # row = 8
    # for i in range(len(SHOESIZERANGE)):
    #     ws[f"{column}{row}"] = metadata["sizeNames"][i]
    #     column = get_next_column_name(column)

    # Save the workbook
    save_workbook(wb, new_file_path)
    logger.debug(f"Workbook saved as {new_file_path}")
