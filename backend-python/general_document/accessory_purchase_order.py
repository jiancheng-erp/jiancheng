"""
生成辅料订购单（拉链 + 拉链头合并格式）

列: 编号 | 工厂货号 | 材料货号 | 拉头颜色 | 单位 | 数量 | 备注
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from logger import logger


def generate_accessory_purchase_order(file_path, order_data):
    """
    生成辅料订购单 Excel 文件。

    order_data keys:
        供应商, 客户名, 订单信息, 备注, 发货地址, 交货期限,
        seriesData: list of {工厂货号, 材料货号, 拉头颜色, 单位, 数量, 备注}
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "辅料订购单"

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Row 1: title ─────────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    ws["A1"] = "辅料订购单"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center_align
    ws.row_dimensions[1].height = 28

    # ── Row 2: supplier / order info ─────────────────────────────────────────
    ws.merge_cells("A2:C2")
    ws["A2"] = "供应商：" + order_data.get("供应商", "")
    ws["A2"].alignment = center_align

    ws.merge_cells("D2:G2")
    ws["D2"] = (
        order_data.get("客户名", "")
        + "  " + order_data.get("订单信息", "")
        + "  " + order_data.get("商标", "")
    ).strip()
    ws["D2"].alignment = center_align
    ws.row_dimensions[2].height = 20

    # ── Row 3: column headers ─────────────────────────────────────────────────
    headers = ["编号", "工厂货号", "材料货号", order_data.get("颜色列名", "颜色"), "单位", "数量", "备注"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = center_align
        cell.border = border
    ws.row_dimensions[3].height = 20

    # ── Data rows ─────────────────────────────────────────────────────────────
    # Merge rows with same 工厂货号 + 材料货号 + 颜色 by summing 数量
    raw_series = order_data.get("seriesData", [])
    merged: dict = {}
    for item in raw_series:
        key = (item.get("工厂货号", ""), item.get("材料货号", ""), item.get("颜色", ""))
        if key in merged:
            try:
                merged[key]["数量"] = (merged[key]["数量"] or 0) + (item.get("数量") or 0)
            except TypeError:
                pass
        else:
            merged[key] = dict(item)
    series = list(merged.values())
    for i, item in enumerate(series):
        row = 4 + i
        values = [
            i + 1,
            item.get("工厂货号", ""),
            item.get("材料货号", ""),
            item.get("颜色", ""),
            item.get("单位", ""),
            item.get("数量", ""),
            item.get("备注", ""),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = border
            cell.alignment = center_align
        ws.row_dimensions[row].height = 20

    # ── Footer ────────────────────────────────────────────────────────────────
    note_row = 4 + len(series)
    ws.merge_cells(f"A{note_row}:G{note_row}")
    ws[f"A{note_row}"] = (
        "注：下单之日超5天内交货，逾期一天将处罚500元/天，"
        "如有异议接单之日起两天内答复沟通！"
    )
    ws[f"A{note_row}"].font = Font(color="FF0000")
    ws[f"A{note_row}"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    ws.row_dimensions[note_row].height = 35

    # Footer: 订购 / 核准 labels
    sign_row = note_row + 1
    ws.merge_cells(f"A{sign_row}:C{sign_row}")
    ws[f"A{sign_row}"] = "订购："
    ws.merge_cells(f"E{sign_row}:G{sign_row}")
    ws[f"E{sign_row}"] = "核准"
    ws.row_dimensions[sign_row].height = 40

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [6, 15, 42, 16, 8, 8, 15]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(file_path)
    logger.debug(f"辅料订购单 saved: {file_path}")


def _item_display_name(item):
    """返回 item 的实际显示名称：优先用 _material_name，否则用 物品名称。"""
    return item.get("_material_name", "") or item.get("物品名称", "")


def _item_color_desc(item):
    """组合 型号+规格+颜色 填入拉头颜色栏。"""
    parts = [
        item.get("_model", ""),
        item.get("_spec", ""),
        item.get("_material_color", ""),
    ]
    return " ".join(p for p in parts if p)


def _is_zipper(item):
    """判断是否为拉链（含"拉链"但不含"拉链头"或"拉头"）。"""
    # Check _material_name first, then fall back to 物品名称
    for field in ("_material_name", "物品名称"):
        name = item.get(field, "")
        if name:
            return "拉链" in name and "拉链头" not in name and "拉头" not in name
    return False


def _is_head(item):
    """判断是否为拉链头（含"拉链头"或"拉头"，如"B字拉头"、"装饰拉头"等）。"""
    for field in ("_material_name", "物品名称"):
        name = item.get(field, "")
        if name:
            return "拉链头" in name or "拉头" in name
    return False


def _is_eyelet(item):
    """判断是否为鞋眼（含"鞋眼"但不含"垫片"）。"""
    for field in ("_material_name", "物品名称"):
        name = item.get(field, "")
        if name:
            return "鞋眼" in name and "垫片" not in name
    return False


def _is_washer(item):
    """判断是否为垫片。"""
    for field in ("_material_name", "物品名称"):
        name = item.get(field, "")
        if name:
            return "垫片" in name
    return False


def _build_color_map(item_list):
    """按 _shoe_color 建立映射，同色取第一个。"""
    m = {}
    for it in item_list:
        sc = it.get("_shoe_color", "")
        if sc not in m:
            m[sc] = it
    return m


def _build_pair_map(item_list):
    """
    建立多层查找映射，用于拉链-拉头配对。
    key = (pair_id, shoe_color, material_color)，每条目注册所有回退键。
    先注册者优先，保证精确键命中正确条目。
    """
    m = {}
    for it in item_list:
        pid = it.get("_zipper_pair_id")
        sc = it.get("_shoe_color", "")
        mc = it.get("_material_color", "")
        for key in [
            (pid, sc, mc),
            (pid, sc, ""),
            (pid, "", mc),
            (pid, "", ""),
            (None, sc, mc),
            (None, sc, ""),
            (None, "", mc),
            (None, "", ""),
        ]:
            if key not in m:
                m[key] = it
    return m


def _find_matching_head(z, head_pair_map, first_head):
    """
    找到与拉链 z 配对的拉头。
    优先顺序（精确→宽松）：
      pair+鞋色+材料色 → pair+鞋色 → pair+材料色 → pair
      → 鞋色+材料色 → 鞋色 → 材料色 → 任意 → first_head
    """
    pid = z.get("_zipper_pair_id")
    sc = z.get("_shoe_color", "")
    mc = z.get("_material_color", "")
    for key in [
        (pid, sc, mc),
        (pid, sc, ""),
        (pid, "", mc),
        (pid, "", ""),
        (None, sc, mc),
        (None, sc, ""),
        (None, "", mc),
        (None, "", ""),
    ]:
        h = head_pair_map.get(key)
        if h is not None:
            return h
    return first_head


def split_zipper_orders(purchase_divide_order_dict):
    """
    识别分采购订单中需要合并打印为「辅料订购单」的条目：
      - 拉链 + 拉链头：按鞋色匹配，输出含「拉头颜色」列
      - 鞋眼 + 垫片：按鞋色匹配，材料货号追加"+垫片"，拉头颜色列留空

    返回:
        standard_dict : pdo_rid → 标准格式数据（已移除 _ 内部字段）
        zipper_dict   : pdo_rid → 辅料订购单格式数据
    """
    standard_dict = {}
    zipper_dict = {}

    for pdo_rid, data in purchase_divide_order_dict.items():
        items = data["seriesData"]

        zipper_items  = [i for i in items if _is_zipper(i)]
        head_items    = [i for i in items if _is_head(i)]
        eyelet_items  = [i for i in items if _is_eyelet(i)]
        washer_items  = [i for i in items if _is_washer(i)]
        other_items   = [
            i for i in items
            if not _is_zipper(i) and not _is_head(i)
            and not _is_eyelet(i) and not _is_washer(i)
        ]

        needs_accessory = (zipper_items and head_items) or (eyelet_items and washer_items)

        if needs_accessory:
            accessory_series = []

            # ── 拉链 + 拉链头 ─────────────────────────────────────────────
            if zipper_items and head_items:
                heads_by_color: dict = {}
                for h in head_items:
                    sc = h.get("_shoe_color", "")
                    heads_by_color.setdefault(sc, []).append(h)
                head_pair_map = _build_pair_map(head_items)
                first_head = head_items[0]

                def _pick_head_for_zipper_z(z):
                    sc = z.get("_shoe_color", "")
                    sc_heads = heads_by_color.get(sc) or heads_by_color.get("") or []
                    if not sc_heads:
                        return first_head
                    unique_head_names = {_item_display_name(h) for h in sc_heads}
                    if len(unique_head_names) == 1:
                        return sc_heads[0]
                    sc_head_map = _build_pair_map(sc_heads)
                    return _find_matching_head(z, sc_head_map, sc_heads[0])

                for z in sorted(zipper_items, key=lambda x: (x.get("_zipper_pair_id") or 0, x.get("_shoe_color", ""))):
                    head = _pick_head_for_zipper_z(z)
                    accessory_series.append({
                        "工厂货号": (z.get("_factory_no", "") + " " + z.get("_shoe_color", "")).strip(),
                        "材料货号": z.get("物品名称", "") or _item_display_name(z),
                        "颜色": _item_color_desc(head),
                        "单位": z.get("单位", ""),
                        "数量": z.get("数量", ""),
                        "备注": z.get("备注", ""),
                    })
            elif zipper_items:
                # 拉链无对应拉链头 — 直接输出
                for z in sorted(zipper_items, key=lambda x: x.get("_shoe_color", "")):
                    accessory_series.append({
                        "工厂货号": (z.get("_factory_no", "") + " " + z.get("_shoe_color", "")).strip(),
                        "材料货号": z.get("物品名称", "") or _item_display_name(z),
                        "颜色": "",  # 拉链头缺失时无拉头颜色
                        "单位": z.get("单位", ""),
                        "数量": z.get("数量", ""),
                        "备注": z.get("备注", ""),
                    })

            # ── 鞋眼 + 垫片 ───────────────────────────────────────────────
            if eyelet_items and washer_items:
                washer_map   = _build_color_map(washer_items)
                first_washer = washer_items[0]
                for e in eyelet_items:
                    sc = e.get("_shoe_color", "")
                    # washer matched by shoe color (fallback to first)
                    washer_map.get(sc) or washer_map.get("") or first_washer
                    mat_desc = (e.get("物品名称", "") or _item_display_name(e)) + "+垫片"
                    accessory_series.append({
                        "工厂货号": (e.get("_factory_no", "") + " " + e.get("_shoe_color", "")).strip(),
                        "材料货号": mat_desc,
                        "颜色": e.get("_material_color", ""),
                        "单位": e.get("单位", ""),
                        "数量": e.get("数量", ""),
                        "备注": e.get("备注", ""),
                    })
            elif eyelet_items:
                # 鞋眼无对应垫片 — 作为普通物品处理
                other_items = list(eyelet_items) + list(other_items)

            zipper_dict[pdo_rid] = {
                **{k: v for k, v in data.items() if k != "seriesData"},
                "seriesData": accessory_series,
            }

            # 剩余普通物品保留到标准格式
            clean_other = [
                {k: v for k, v in i.items() if not k.startswith("_")}
                for i in other_items
            ]
            if clean_other:
                standard_dict[pdo_rid] = {
                    **{k: v for k, v in data.items() if k != "seriesData"},
                    "seriesData": clean_other,
                }
        else:
            # 无需合并 — 去除内部字段，全部走标准格式
            clean_all = [
                {k: v for k, v in i.items() if not k.startswith("_")}
                for i in items
            ]
            standard_dict[pdo_rid] = {
                **{k: v for k, v in data.items() if k != "seriesData"},
                "seriesData": clean_all,
            }

    return standard_dict, zipper_dict


# Material type IDs that stay in standard (fabric/lining/chemical) format
# 1=面料, 2=里料, 4=化工 (热熔胶 etc.)
_STANDARD_TYPE_IDS = {1, 2, 4}


def split_second_purchase_orders(purchase_divide_order_dict):
    """
    二次采购专用分类函数：
      - 面料(1)、里料(2)、化工(4, 热熔胶等) → 标准采购订单格式
      - 其余所有辅料 → 辅料订购单格式，其中：
          * 拉链 + 拉链头：按鞋色匹配，拉头颜色列填拉链头材料颜色
          * 鞋眼 + 垫片：按鞋色匹配，材料货号追加 "+垫片"
          * 其他辅料：直接输出，拉头颜色留空

    返回:
        standard_dict : pdo_rid → 标准格式数据
        accessory_dict: pdo_rid → 辅料订购单格式数据
    """
    standard_dict = {}
    accessory_dict = {}

    for pdo_rid, data in purchase_divide_order_dict.items():
        items = data["seriesData"]

        std_items = [i for i in items if i.get("_material_type_id") in _STANDARD_TYPE_IDS]
        acc_items = [i for i in items if i.get("_material_type_id") not in _STANDARD_TYPE_IDS]

        # ── Standard items ────────────────────────────────────────────────────
        clean_std = [
            {k: v for k, v in i.items() if not k.startswith("_")}
            for i in std_items
        ]
        if clean_std:
            standard_dict[pdo_rid] = {
                **{k: v for k, v in data.items() if k != "seriesData"},
                "seriesData": clean_std,
            }

        if not acc_items:
            continue

        # ── Accessory items: apply pairing then output ────────────────────────
        zipper_items = [i for i in acc_items if _is_zipper(i)]
        head_items   = [i for i in acc_items if _is_head(i)]
        eyelet_items = [i for i in acc_items if _is_eyelet(i)]
        washer_items = [i for i in acc_items if _is_washer(i)]
        other_items  = [
            i for i in acc_items
            if not _is_zipper(i) and not _is_head(i)
            and not _is_eyelet(i) and not _is_washer(i)
        ]

        has_zipper_pair = bool(zipper_items and head_items)
        # Column label: "拉头颜色" only for zipper orders, else "颜色"
        color_col_name = "拉头颜色" if has_zipper_pair else "颜色"

        accessory_series = []

        # 拉链 + 拉链头
        if zipper_items:
            # 按鞋色建立拉头索引
            heads_by_color: dict = {}
            for h in head_items:
                sc = h.get("_shoe_color", "")
                heads_by_color.setdefault(sc, []).append(h)

            head_pair_map = _build_pair_map(head_items) if head_items else {}
            first_head = head_items[0] if head_items else {}

            def _pick_head_for_zipper(z):
                """
                为拉链选取配对拉头：
                  1. 同鞋色只有一种拉头 → 直接使用（无需 pair_id）
                  2. 同鞋色有多种拉头   → 用 pair_id + material_color 精确匹配
                  3. 兜底              → 全局 first_head
                """
                sc = z.get("_shoe_color", "")
                sc_heads = heads_by_color.get(sc) or heads_by_color.get("") or []
                if not sc_heads:
                    return first_head
                # 同色只有一种拉头：直接返回
                unique_head_names = {_item_display_name(h) for h in sc_heads}
                if len(unique_head_names) == 1:
                    return sc_heads[0]
                # 同色多种拉头：按 pair_id + material_color 匹配
                sc_head_map = _build_pair_map(sc_heads)
                matched = _find_matching_head(z, sc_head_map, sc_heads[0])
                return matched

            for z in sorted(zipper_items, key=lambda x: (x.get("_zipper_pair_id") or 0, x.get("_shoe_color", ""))):
                head = _pick_head_for_zipper(z) if head_items else {}
                accessory_series.append({
                    "工厂货号": (z.get("_factory_no", "") + " " + z.get("_shoe_color", "")).strip(),
                    "材料货号": z.get("物品名称", "") or _item_display_name(z),
                    "颜色": _item_color_desc(head) if head else "",
                    "单位": z.get("单位", ""),
                    "数量": z.get("数量", ""),
                    "备注": z.get("备注", ""),
                })

        # 鞋眼 + 垫片
        if eyelet_items:
            washer_map   = _build_color_map(washer_items) if washer_items else {}
            first_washer = washer_items[0] if washer_items else None
            for e in eyelet_items:
                sc = e.get("_shoe_color", "")
                mat_desc = e.get("物品名称", "") or _item_display_name(e)
                if washer_map or first_washer:
                    mat_desc += "+垫片"
                accessory_series.append({
                    "工厂货号": (e.get("_factory_no", "") + " " + e.get("_shoe_color", "")).strip(),
                    "材料货号": mat_desc,
                    "颜色": e.get("_material_color", ""),
                    "单位": e.get("单位", ""),
                    "数量": e.get("数量", ""),
                    "备注": e.get("备注", ""),
                })

        # All other accessory items (饰品, 底材, 包材, etc.)
        for item in other_items:
            accessory_series.append({
                "工厂货号": (item.get("_factory_no", "") + " " + item.get("_shoe_color", "")).strip(),
                "材料货号": item.get("物品名称", "") or _item_display_name(item),
                "颜色": item.get("_material_color", ""),
                "单位": item.get("单位", ""),
                "数量": item.get("数量", ""),
                "备注": item.get("备注", ""),
            })

        if accessory_series:
            accessory_series.sort(key=lambda x: x.get("工厂货号", ""))
            accessory_dict[pdo_rid] = {
                **{k: v for k, v in data.items() if k != "seriesData"},
                "颜色列名": color_col_name,
                "seriesData": accessory_series,
            }

    return standard_dict, accessory_dict
