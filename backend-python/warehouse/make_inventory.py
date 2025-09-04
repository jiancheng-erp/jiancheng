# routes/material_inventory.py
from cProfile import label
from flask import Blueprint, current_app, request, jsonify, send_file
from sqlalchemy import func, or_, text, update
from decimal import Decimal
from constants import SHOESIZERANGE
from app_config import db
from models import *
from datetime import datetime
from general_document.make_inventory_excel import build_inventory_excel
from file_locations import FILE_STORAGE_PATH
import os, glob, json
from werkzeug.utils import secure_filename
import logger
import time
from sqlalchemy.orm import load_only
from sqlalchemy.exc import IntegrityError
from sqlalchemy.dialects.mysql import insert as mysql_insert
import math
from datetime import date, datetime
from script.refresh_spu_rid import generate_spu_rid
# ====== 目录常量 ======
EXPORT_DIR = os.path.join(FILE_STORAGE_PATH, "财务部文件", "盘库文件")
IMPORT_DIR = os.path.join(FILE_STORAGE_PATH, "财务部文件", "盘库文件回传")
DIFF_DIR   = os.path.join(IMPORT_DIR, "diff")
for _d in (EXPORT_DIR, IMPORT_DIR, DIFF_DIR):
    os.makedirs(_d, exist_ok=True)

# ====== Excel 解析配置 ======
# 可见工作表名（与你导出时的一致）
VISIBLE_SHEET_NAME = "材料盘库"
# 主键列（必须在导出中包含）
KEY_COL = "SPU_MATERIAL_ID"
# 主要汇总列
COL_TOTAL_QTY   = "总库存"
COL_TOTAL_VALUE = "总价格"
COL_AVG_PRICE   = "平均单价"
# 其他可显示列（用于回显，不参与加减）
DISPLAY_COLS = ["SPU编号", "材料类型", "材料名称", "材料型号", "材料规格", "颜色", "单位"]

make_inventory_bp = Blueprint("material_inventory_bp", __name__)

# ====== 工具函数 ======
def _ensure_dir(p: str):
    os.makedirs(p, exist_ok=True)

def _sanitize_filename_part(s: str) -> str:
    s = (s or "").strip()
    return s.translate(str.maketrans('', '', '\\/:*?"<>|')).replace("\n", " ").replace("\r", " ")

def _inventory_export_filename(*, record: MakeInventoryRecord | None, inventory_date: str, inventory_reason: str) -> str:
    # 保持你现在的实现（带 rid 的会是：材料盘库_<RID>_日期_原因.xlsx；没 rid 就是 材料盘库_日期_原因.xlsx）
    safe_reason = _sanitize_filename_part(inventory_reason)
    if record and record.make_inventory_rid:
        date_part = record.record_date.isoformat() if record.record_date else (inventory_date or "")
        base = f"{record.make_inventory_rid}_{date_part}{'_'+safe_reason if safe_reason else ''}"
    else:
        base = f"{inventory_date or ''}{'_'+safe_reason if safe_reason else ''}"
    return f"材料盘库_{base}.xlsx"

def _find_baseline_export_file(record: MakeInventoryRecord) -> str | None:
    """
    基线文件查找优先级（按 record.record_date 精确匹配）：
      1) 材料盘库_<RID>_<DATE>_*.xlsx
      2) 材料盘库_<DATE>_*.xlsx              （无RID历史文件兜底）
      3) 材料盘库_<RID>_*.xlsx               （只匹配RID，不限日期）
      4) 材料盘库_*.xlsx 中按修改时间最新     （最终兜底，不建议长期依赖）
    """
    if not record or not record.make_inventory_rid:
        return None

    _ensure_dir(EXPORT_DIR)
    rid = record.make_inventory_rid
    date_str = ""
    try:
        if record.record_date:
            # 保证是 YYYY-MM-DD 形式
            date_str = record.record_date.isoformat()
    except Exception:
        date_str = ""

    # 1) RID + DATE 精确匹配
    if date_str:
        pat1 = os.path.join(EXPORT_DIR, f"材料盘库_{rid}_{date_str}_*.xlsx")
        cand1 = sorted(glob.glob(pat1), key=os.path.getmtime, reverse=True)
        if cand1:
            return cand1[0]

    # 2) 仅日期匹配（历史无rid导出兜底）
    if date_str:
        pat2 = os.path.join(EXPORT_DIR, f"材料盘库_{date_str}_*.xlsx")
        cand2 = sorted(glob.glob(pat2), key=os.path.getmtime, reverse=True)
        if cand2:
            return cand2[0]

    # 3) 仅 RID 匹配（不限日期）
    pat3a = os.path.join(EXPORT_DIR, f"材料盘库_{rid}_*.xlsx")
    pat3b = os.path.join(EXPORT_DIR, f"{rid}.xlsx")  # 极早期命名
    cand3 = glob.glob(pat3a) + glob.glob(pat3b)
    if cand3:
        cand3.sort(key=os.path.getmtime, reverse=True)
        return cand3[0]

    # 4) 最终兜底：目录下任何“材料盘库_*.xlsx”的最新一个
    files = [p for p in glob.glob(os.path.join(EXPORT_DIR, "材料盘库_*.xlsx"))
             if os.path.basename(p).startswith("材料盘库_")]
    return max(files, key=os.path.getmtime) if files else None

def _sum_current_amount_cols():
    exprs = []
    for sz in SHOESIZERANGE:
        col = getattr(MaterialStorage, f"size_{sz}_current_amount")
        exprs.append(func.coalesce(func.sum(col), 0).label(f"size_{sz}"))
    return exprs

def _to_float(v):
    if isinstance(v, Decimal):
        return float(v)
    try:
        return float(v or 0)
    except Exception:
        return 0.0

def _read_visible_sheet_to_map(xlsx_path: str) -> dict:
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, data_only=True)

    # 1) 选工作表：优先“材料盘库”
    sheet_names = [VISIBLE_SHEET_NAME] + [n for n in wb.sheetnames if n != VISIBLE_SHEET_NAME]
    ws = None
    for name in sheet_names:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        raise ValueError(f"未找到工作表：{VISIBLE_SHEET_NAME} 或任何工作表")

    # 2) 动态定位表头行（前 20 行里找到同时包含“SPU编号/总库存”的行）
    header_row_idx = None
    header_map = {}
    MUST_HEADERS = ["SPU编号", "总库存"]

    def build_header_map_at_row(r: int):
        m = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip():
                m[v.strip()] = c
        return m

    for r in range(1, min(ws.max_row, 20) + 1):
        hm = build_header_map_at_row(r)
        if all(h in hm for h in MUST_HEADERS):
            header_row_idx = r
            header_map = hm
            break

    if header_row_idx is None:
        raise ValueError(f"未在前20行找到表头（至少应包含 {MUST_HEADERS}）")

    def cell(r, header_name, default=None):
        cidx = header_map.get(header_name)
        if not cidx:
            return default
        v = ws.cell(row=r, column=cidx).value
        return v if v is not None else default

    # 3) 列名映射（与导出一致）
    KEY_NAME       = "SPU编号"   # 作为 key
    COL_SUPPLIER   = "厂家名称"
    COL_TYPE       = "材料类型"
    COL_NAME       = "材料名称"
    COL_MODEL      = "材料型号"
    COL_SPEC       = "材料规格"
    COL_COLOR      = "颜色"
    COL_UNIT       = "单位"
    COL_TOTAL_QTY  = "总库存"
    COL_TOTAL_VAL  = "总价格"
    COL_AVG_PRICE  = "平均单价"
    COL_ORDER_RID  = "订单号"

    # 动态识别尺码列（尺码34..46）——即使你后面不展示，保留兼容（不影响性能）
    size_headers = []
    for sz in SHOESIZERANGE:
        colname = f"尺码{sz}"
        if colname in header_map:
            size_headers.append((str(sz), colname))

    # 4) 逐行读取数据（从 header 下一行开始）
    res = {}
    new_idx = 0  # 生成“未解析新材料”的占位 id
    for r in range(header_row_idx + 1, ws.max_row + 1):
        spu_rid_raw = cell(r, KEY_NAME, None)
        order_rid_val = cell(r, COL_ORDER_RID, None)
        supplier    = cell(r, COL_SUPPLIER, "") or ""
        type_name   = cell(r, COL_TYPE, "") or ""
        name        = cell(r, COL_NAME, "") or ""
        model       = cell(r, COL_MODEL, "") or ""
        spec        = cell(r, COL_SPEC, "") or ""
        color       = cell(r, COL_COLOR, "") or ""
        unit        = cell(r, COL_UNIT, "") or ""


        # 如果该行完全空，就跳过
        if not any([spu_rid_raw, supplier, type_name, name, model, spec, color, unit]):
            continue

        # 先构造对象（后续可能补齐 spuRid）
        row_obj = {
            "spuRid":        (str(spu_rid_raw).strip() if spu_rid_raw not in (None, "") else ""),
            "supplier":      supplier,
            "type":          type_name,
            "name":          name,
            "model":         model,
            "specification": spec,
            "color":         color,
            "unit":          unit,
            "totalCurrentAmount": _to_float(cell(r, COL_TOTAL_QTY, 0)),
            "totalValueAmount":   _to_float(cell(r, COL_TOTAL_VAL, 0)),
            "avgUnitPrice":       _to_float(cell(r, COL_AVG_PRICE, 0)),
            "orderRid":           order_rid_val,
            "sizes": {},
            # 标记：该行是否通过属性成功解析为已有 SPU
            "_resolvedToSpu": False
        }
        for sz_str, colname in size_headers:
            row_obj["sizes"][sz_str] = _to_float(cell(r, colname, 0))

        # 关键：如果缺少 SPU编号，尝试通过属性匹配已有 SPU
        key = row_obj["spuRid"]
        if not key:
            resolved_rid, _resolved_id = _resolve_spu_by_attrs(
                supplier_name=supplier, type_name=type_name, material_name=name,
                model=model, specification=spec, color=color
            )
            if resolved_rid:
                key = resolved_rid
                row_obj["spuRid"] = resolved_rid
                row_obj["_resolvedToSpu"] = True
            else:
                # 仍无法解析——作为“新材料”占位 key
                new_idx += 1
                key = f"__NEW__#{new_idx}"  # 不会与任何真实 SPU 冲突

        res[key] = row_obj

    return res

def _norm_str(s, *, allow_none: bool = False) -> str | None:
    """
    把任意类型的单元格值安全转成去空格后的字符串。
    - None -> "" (或 allow_none=True 时返回 None)
    - bytes -> utf-8 解码
    - int/float/Decimal -> 字符串（float 的 NaN/Inf 视为空）
    - date/datetime -> ISO 格式
    - 其他 -> str(...)
    """
    if s is None:
        return None if allow_none else ""

    if isinstance(s, bytes):
        s = s.decode("utf-8", errors="ignore")
    elif isinstance(s, (int, Decimal)):
        s = str(s)
    elif isinstance(s, float):
        if math.isnan(s) or math.isinf(s):
            return None if allow_none else ""
        # 对类似 123.0 的值，去掉小数点零更友好
        s = str(int(s)) if s.is_integer() else str(s)
    elif isinstance(s, (date, datetime)):
        s = s.isoformat()
    else:
        s = str(s)

    return s.strip()

def _resolve_spu_by_attrs(*, supplier_name: str, type_name: str, material_name: str,
                          model: str, specification: str, color: str) -> tuple[str | None, int | None]:
    """
    尝试用“厂家/类型/名称/型号/规格/颜色”去匹配一个已存在的 SPU。
    返回 (spu_rid, spu_material_id)；匹配不到则返回 (None, None)。

    匹配策略：所有传入且非空的字段做“等值（忽略大小写/空格差异）”过滤。
    """
    sup = _norm_str(supplier_name)
    typ = _norm_str(type_name)
    nam = _norm_str(material_name)
    mdl = _norm_str(model)
    spc = _norm_str(specification)
    clr = _norm_str(color)

    q = (db.session.query(SPUMaterial.spu_rid, SPUMaterial.spu_material_id)
         .join(Material, Material.material_id == SPUMaterial.material_id)
         .join(MaterialType, MaterialType.material_type_id == Material.material_type_id)
         .join(Supplier, Supplier.supplier_id == Material.material_supplier))

    # 动态等值（大小写不敏感）
    def eq(col, val):
        return func.lower(func.trim(col)) == func.lower(func.trim(func.cast(val, col.type)))  # 统一大小写与空格

    filters = []
    if sup: filters.append(func.lower(func.trim(Supplier.supplier_name)) == func.lower(func.trim(sup)))
    if typ: filters.append(func.lower(func.trim(MaterialType.material_type_name)) == func.lower(func.trim(typ)))
    if nam: filters.append(func.lower(func.trim(Material.material_name)) == func.lower(func.trim(nam)))
    if mdl: filters.append(func.lower(func.trim(SPUMaterial.material_model)) == func.lower(func.trim(mdl)))
    if spc: filters.append(func.lower(func.trim(SPUMaterial.material_specification)) == func.lower(func.trim(spc)))
    if clr: filters.append(func.lower(func.trim(SPUMaterial.color)) == func.lower(func.trim(clr)))

    if filters:
        q = q.filter(*filters)
    else:
        # 一个有效字段都没有，无法匹配
        return None, None

    # 这里简单取第一个匹配；如需更严格可加唯一性校验
    row = q.first()
    return (row.spu_rid, row.spu_material_id) if row else (None, None)


# ====== 业务：库存汇总 ======
@make_inventory_bp.route("/warehouse/inventorysummary", methods=["GET"])
def material_inventory_summary():
    """
    盘库汇总（材料库）
    - 维度：SPU + order_id（订单可为空）
    - totalValueAmount = SUM(average_price * current_amount)
    - avgUnitPrice     = totalValueAmount / totalCurrentAmount
    - 过滤 totalCurrentAmount = 0
    """

    def to_float(v):
        if v is None:
            return 0.0
        if isinstance(v, (int, float, Decimal)):
            return float(v)
        try:
            s = str(v).strip().replace(",", "").replace("，", "").replace("\u3000", "")
            return float(s) if s else 0.0
        except Exception:
            return 0.0

    page = max(int(request.args.get("page", 1)), 1)
    page_size = max(int(request.args.get("pageSize", 20)), 1)
    keyword = (request.args.get("keyword") or "").strip()

    sort_by = (request.args.get("sortBy") or "total_current_amount").strip()
    sort_order = (request.args.get("sortOrder") or "desc").strip().lower()
    sort_order = "desc" if sort_order not in ("asc", "desc") else sort_order

    size_sum_exprs = _sum_current_amount_cols()
    qty_sum_expr   = func.coalesce(func.sum(MaterialStorage.current_amount), 0)
    value_sum_expr = func.coalesce(func.sum(MaterialStorage.average_price * MaterialStorage.current_amount), 0)
    avg_price_expr = (value_sum_expr / func.nullif(qty_sum_expr, 0))

    base_q = (
        db.session.query(
            # 关键：选择 SPU + order_id 两个维度（order_id 可能为空）
            SPUMaterial.spu_material_id.label("spu_material_id"),
            SPUMaterial.spu_rid.label("spu_rid"),
            MaterialType.material_type_name.label("material_type_name"),
            Material.material_name.label("material_name"),
            Supplier.supplier_name.label("material_supplier"),
            SPUMaterial.material_model.label("material_model"),
            SPUMaterial.material_specification.label("material_specification"),
            SPUMaterial.color.label("color"),
            MaterialStorage.actual_inbound_unit.label("unit"),

            # 新增：把存储里的 order_id 直接选出来（用于分组 & 返回）
            MaterialStorage.order_id.label("order_id"),
            # 订单号（人读）：左连接可能为 None
            Order.order_rid.label("order_rid"),

            qty_sum_expr.label("total_current_amount"),
            value_sum_expr.label("total_value_amount"),
            avg_price_expr.label("avg_unit_price"),
            *size_sum_exprs,
        )
        .select_from(MaterialStorage)
        .join(SPUMaterial, SPUMaterial.spu_material_id == MaterialStorage.spu_material_id)
        .join(Material, Material.material_id == SPUMaterial.material_id)
        .join(Order, Order.order_id == MaterialStorage.order_id, isouter=True)
        .join(MaterialType, MaterialType.material_type_id == Material.material_type_id)
        .join(Supplier, Supplier.supplier_id == Material.material_supplier)
    )

    if keyword:
        like = f"%{keyword}%"
        base_q = base_q.filter(
            or_(
                SPUMaterial.material_model.ilike(like),
                SPUMaterial.material_specification.ilike(like),
                SPUMaterial.color.ilike(like),
                SPUMaterial.spu_rid.ilike(like),
                Material.material_name.ilike(like),
                MaterialType.material_type_name.ilike(like),
                # 允许按订单号搜索（左连接下可为 NULL，不影响）
                Order.order_rid.ilike(like),
            )
        )

    # 分组：加入 MaterialStorage.order_id 与 Order.order_rid
    base_q = base_q.group_by(
        SPUMaterial.spu_material_id,
        SPUMaterial.spu_rid,
        MaterialType.material_type_name,
        Material.material_name,
        Supplier.supplier_name,
        MaterialStorage.actual_inbound_unit,
        SPUMaterial.material_model,
        SPUMaterial.material_specification,
        SPUMaterial.color,
        MaterialStorage.order_id,  # 新增：真正用于聚合的 order_id 维度
        Order.order_rid,           # 显示用的人类可读订单号
    ).having(qty_sum_expr > 0)

    sub = base_q.subquery()

    SORTABLE = {
        "total_current_amount": sub.c.total_current_amount,
        "total_value_amount": sub.c.total_value_amount,
        "avg_unit_price": sub.c.avg_unit_price,
        "spu_material_id": sub.c.spu_material_id,
        "spu_rid": sub.c.spu_rid,
        "material_type_name": sub.c.material_type_name,
        "material_name": sub.c.material_name,
        "material_model": sub.c.material_model,
        "material_specification": sub.c.material_specification,
        "color": sub.c.color,
        "unit": sub.c.unit,
        # 新增：支持按 order_id / order_rid 排序
        "order_id": sub.c.order_id,
        "order_rid": sub.c.order_rid,
    }
    for sz in SHOESIZERANGE:
        SORTABLE[f"size_{sz}"] = getattr(sub.c, f"size_{sz}")

    sort_col = SORTABLE.get(sort_by, sub.c.total_current_amount)
    order_by_expr = sort_col.desc() if sort_order == "desc" else sort_col.asc()

    total = db.session.query(func.count(text("1"))).select_from(sub).scalar() or 0
    rows = (
        db.session.query(sub)
        .order_by(order_by_expr)
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    result_list = []
    for r in rows:
        sizes = {str(sz): to_float(getattr(r, f"size_{sz}", 0)) for sz in SHOESIZERANGE}
        result_list.append({
            "spuMaterialId": r.spu_material_id,
            "spuRid": r.spu_rid or "",
            "type": r.material_type_name or "",
            "name": r.material_name or "",
            "model": r.material_model or "",
            "supplier": r.material_supplier or "",
            "specification": r.material_specification or "",
            "unit": r.unit or "",
            "color": r.color or "",
            "totalCurrentAmount": to_float(r.total_current_amount),
            "totalValueAmount": to_float(r.total_value_amount),
            "avgUnitPrice": to_float(r.avg_unit_price),
            "sizes": sizes,

            # 返回两个字段：数值型 id（便于后端流转）+ 人类可读单号（便于展示 & 导出）
            "orderId": r.order_id,           # 可能为 None
            "orderRid": r.order_rid or "",   # 可能为 ""
        })

    return jsonify({
        "code": 200,
        "message": "ok",
        "list": result_list,
        "total": int(total),
        "sizeColumns": SHOESIZERANGE,
    })


# ====== 盘库记录 CRUD ======
def _gen_rid(prefix="MI"):
    today = datetime.now().strftime("%Y%m%d")
    like = f"{prefix}{today}-%"
    seq = (db.session.query(func.count(MakeInventoryRecord.make_inventory_record_id))
           .filter(MakeInventoryRecord.make_inventory_rid.like(like)).scalar() or 0) + 1
    return f"{prefix}{today}-{seq:04d}"

@make_inventory_bp.route("/warehouse/makeinventory/list", methods=["GET"])
def list_make_inventory():
    page = max(int(request.args.get("page", 1)), 1)
    page_size = max(int(request.args.get("pageSize", 20)), 1)
    keyword = (request.args.get("keyword") or "").strip()
    date_from = request.args.get("dateFrom")
    date_to = request.args.get("dateTo")

    q = MakeInventoryRecord.query
    if keyword:
        like = f"%{keyword}%"
        q = q.filter(
            (MakeInventoryRecord.make_inventory_rid.ilike(like)) |
            (MakeInventoryRecord.make_inventory_reason.ilike(like))
        )
    if date_from:
        q = q.filter(MakeInventoryRecord.record_date >= date_from)
    if date_to:
        q = q.filter(MakeInventoryRecord.record_date <= date_to)

    total = q.count()
    rows = (q.order_by(MakeInventoryRecord.make_inventory_record_id.desc())
              .offset((page-1)*page_size).limit(page_size).all())

    def to_obj(r):
        return {
            "makeInventoryRecordId": r.make_inventory_record_id,
            "makeInventoryRid": r.make_inventory_rid,
            "recordDate": r.record_date.isoformat() if r.record_date else "",
            "reason": r.make_inventory_reason or "",
            "inboundRecordId": r.inbound_record_id,
            "outboundRecordId": r.outbound_record_id,
            "status": r.make_inventory_status,             # 0未回传 1已回传 2等待确认
            "excelReuploadStatus": r.excel_reupload_status # 0/1
        }

    return jsonify({"code": 200, "message": "ok", "list": [to_obj(r) for r in rows], "total": total})

@make_inventory_bp.route("/warehouse/makeinventory/update", methods=["PUT"])
def update_make_inventory():
    data = request.get_json(force=True) or {}
    rec_id = data.get("makeInventoryRecordId")
    if not rec_id:
        return jsonify({"code": 400, "message": "makeInventoryRecordId 必填"}), 400
    rec = MakeInventoryRecord.query.get(rec_id)
    if not rec:
        return jsonify({"code": 404, "message": "记录不存在"}), 404

    if "recordDate" in data:
        rd = data.get("recordDate")
        rec.record_date = datetime.strptime(rd, "%Y-%m-%d").date() if rd else None
    if "reason" in data:
        rec.make_inventory_reason = (data.get("reason") or "").strip()
    db.session.commit()
    return jsonify({"code": 200, "message": "ok"})



@make_inventory_bp.route("/warehouse/makeinventory/confirm", methods=["POST"])
def confirm_make_inventory():
    """
    盘库确认（按 supplier_id + warehouse_id 拆分）：
      A. 清库出库：仅对“首入库时间 <= 盘库日期”的 storage 清零并建出库单
      B. 回填入库：以 (SPU + 单位 + order_id) 聚合，优先复用已存在 storage（避免唯一键冲突）
         - 按行严校验，禁止创建“空 SPU”
         - 供应商缺失自动创建；材料类型必须存在（不自动建）
         - 有订单号必须同时填入 order_id / order_shoe_id（查不到则都为 None）
      C. 若材料类型未配置 warehouse_id，返回 400 并附带可定位到 Excel 的 missingRows 清单
    """

    # ---------- 小工具 ----------
    def _s(v) -> str:
        return str(v).strip() if v is not None else ""

    def _nonempty(*vals) -> bool:
        return any(_s(v) != "" for v in vals)

    def _validate_after_row(row: dict) -> tuple[bool, str]:
        """
        有 spuRid -> 直接有效；否则必须：
        - type/material_name 非空
        - model/spec/color 至少一个
        - unit 非空
        """
        if _s(row.get("spuRid")):
            return True, ""
        type_name     = _s(row.get("type"))
        material_name = _s(row.get("name"))
        model         = _s(row.get("model"))
        spec          = _s(row.get("specification"))
        color         = _s(row.get("color"))
        unit          = _s(row.get("unit"))

        if not type_name or not material_name:
            return False, "材料类型或材料名称为空"
        if not _nonempty(model, spec, color):
            return False, "型号/规格/颜色至少应有一个"
        if not unit:
            return False, "单位为空"
        return True, ""

    def to_dec(x):
        try:
            if isinstance(x, Decimal):
                return x
            if x is None:
                return Decimal("0")
            return Decimal(str(x).replace(",", "").strip())
        except Exception:
            return Decimal("0")

    def _supplier_and_wh_by_spu(spu_id: int):
        supplier_id, warehouse_id = None, None
        spu = db.session.get(SPUMaterial, spu_id)
        if not spu:
            print(f"SPU not found: {spu_id}")
            return None, None

        material = db.session.get(Material, spu.material_id)
        if material:
            supplier_id = material.material_supplier
            mt = db.session.get(MaterialType, material.material_type_id)
            if mt:
                # 如果你的列名不是 warehouse_id，这里改成真实的列名
                wh_id = getattr(mt, "warehouse_id", None)
                warehouse_id = wh_id

                if wh_id is None:
                    # 仅在缺仓库时打诊断日志
                    try:
                        same_named = (
                            db.session.query(MaterialType.material_type_id, MaterialType.material_type_name,
                                            getattr(MaterialType, "warehouse_id").label("warehouse_id"))
                            .filter(MaterialType.material_type_name == mt.material_type_name)
                            .all()
                        )
                    except Exception:
                        same_named = []
                    logger.logger.error(
                        "【诊断/缺仓库】spu_id=%s material_id=%s material_type_id=%s "
                        "type_name=%s warehouse_id(None) 同名类型候选=%s",
                        spu_id, material.material_id, mt.material_type_id,
                        getattr(mt, "material_type_name", None),
                        [{"id": x.material_type_id, "name": x.material_type_name, "warehouse_id": x.warehouse_id} for x in same_named]
                    )
            else:
                logger.logger.error("【诊断】spu_id=%s 指向的 material_type_id=%s 不存在", spu_id, material.material_type_id)
        return supplier_id, warehouse_id

    def _ensure_spu_only(after_row: dict) -> int:
        """
        仅确保 SPU 存在；严格防空 & 不制造“空 SPU”
        - 有 spuRid：精确找；若无，按属性尝试找“已有”SPU，不存在则报错
        - 无 spuRid：先通过 _validate_after_row；类型必须已存在；供应商可自动创建；字段充分才创建 Material/SPU
        """
        spu_rid = _s(after_row.get("spuRid"))
        if spu_rid:
            spu_obj = db.session.query(SPUMaterial).filter_by(spu_rid=spu_rid).first()
            if spu_obj:
                return spu_obj.spu_material_id

            # 用属性找“已有”SPU（不够信息或找不到 -> 抛错，不创建）
            type_name     = _s(after_row.get("type"))
            material_name = _s(after_row.get("name"))
            model         = _s(after_row.get("model"))
            spec          = _s(after_row.get("specification"))
            color         = _s(after_row.get("color"))

            mat = (db.session.query(Material)
                   .join(MaterialType, MaterialType.material_type_id == Material.material_type_id)
                   .filter(Material.material_name == material_name,
                           MaterialType.material_type_name == type_name)
                   .first()) if (material_name and type_name) else None
            if mat:
                spu = (db.session.query(SPUMaterial)
                       .filter(SPUMaterial.material_id == mat.material_id,
                               SPUMaterial.material_model == model,
                               SPUMaterial.material_specification == spec,
                               SPUMaterial.color == color).first())
                if spu:
                    return spu.spu_material_id
            raise ValueError("按 SPU编号/属性 未找到已有 SPU，且不允许自动创建")

        # 无 spuRid：严格校验
        ok, reason = _validate_after_row(after_row)
        if not ok:
            raise ValueError(reason)

        type_name     = _s(after_row.get("type"))
        material_name = _s(after_row.get("name"))
        model         = _s(after_row.get("model"))
        spec          = _s(after_row.get("specification"))
        color         = _s(after_row.get("color"))
        supplier_name = _s(after_row.get("supplier"))
        unit          = _s(after_row.get("unit"))

        # 供应商：不存在则创建
        supplier = None
        if supplier_name:
            supplier = db.session.query(Supplier).filter_by(supplier_name=supplier_name).first()
            if not supplier:
                supplier = Supplier(supplier_name=supplier_name)
                db.session.add(supplier)
                db.session.flush()

        # 类型必须存在（不自动建，避免入错仓库）
        mtype = db.session.query(MaterialType).filter_by(material_type_name=type_name).first()
        if not mtype:
            raise ValueError(f"材料类型不存在：{type_name}")

        # Material：名称+类型+供应商 唯一
        material = (db.session.query(Material)
                    .filter(Material.material_name == material_name,
                            Material.material_type_id == mtype.material_type_id,
                            Material.material_supplier == (supplier.supplier_id if supplier else None))
                    .first())
        if not material:
            material_category = 0
            if type_name == "底材":
                material_category = 1
            material = Material(
                material_name=material_name,
                material_type_id=mtype.material_type_id,
                material_supplier=(supplier.supplier_id if supplier else None),
                material_unit=unit,
                material_creation_date=datetime.now().date(),
                material_category=material_category
            )
            db.session.add(material); db.session.flush()

        # SPU：Material + 型号 + 规格 + 颜色 唯一
        spu = (db.session.query(SPUMaterial)
               .filter(SPUMaterial.material_id == material.material_id,
                       SPUMaterial.material_model == model,
                       SPUMaterial.material_specification == spec,
                       SPUMaterial.color == color).first())
        if not spu:
            new_rid = generate_spu_rid(material.material_id)
            spu = SPUMaterial(
                material_id=material.material_id,
                material_model=model,
                material_specification=spec,
                color=color,
                spu_rid=new_rid,
            )
            db.session.add(spu); db.session.flush()

        return spu.spu_material_id

    def _chunk(iterable, n=500):
        buf = []
        for x in iterable:
            buf.append(x)
            if len(buf) >= n:
                yield buf; buf = []
        if buf:
            yield buf

    def _get_or_upsert_storage(spu_id: int,
                            unit: str,
                            order_id: int | None,
                            order_shoe_id: int | None,
                            unit_price: Decimal) -> MaterialStorage:
        """
        幂等获取/创建 MaterialStorage（MySQL）：
        - 先查：若已存在直接复用，必要时补 order_shoe_id
        - 否则：INSERT IGNORE 插入；若因唯一键已存在 -> 不报错，再查复用
        绝不触发 IntegrityError，因此不会把上游 Supplier/Material/SPU 回滚掉。
        """

        # 1) 先查（多数情况下直接复用）
        q = (db.session.query(MaterialStorage)
            .filter(MaterialStorage.spu_material_id == spu_id,
                    MaterialStorage.actual_inbound_unit == unit))
        if order_id is None:
            q = q.filter(MaterialStorage.order_id.is_(None))
        else:
            q = q.filter(MaterialStorage.order_id == order_id)

        st = q.first()
        if st:
            # 仅在缺失时补齐 order_shoe_id（避免覆盖）
            if (not st.order_shoe_id) and order_shoe_id:
                st.order_shoe_id = order_shoe_id
                db.session.flush()
            return st

        # 2) 不存在 -> INSERT IGNORE（命中唯一键不会报错）
        stmt = mysql_insert(MaterialStorage).values(
            spu_material_id=spu_id,
            order_id=order_id if order_id else None,
            order_shoe_id=order_shoe_id if order_shoe_id else None,
            actual_inbound_unit=unit,
            inbound_amount=Decimal("0"),
            current_amount=Decimal("0"),
            unit_price=unit_price,
            average_price=unit_price,
            shoe_size_columns=[],
        ).prefix_with("IGNORE")   # 关键：让冲突静默
        db.session.execute(stmt)

        # 3) 再查一遍（插入成功或因唯一键已存在都会查到）
        st = q.first()
        if not st:
            # 极端并发下理论上也应该能查到；若没查到，抛 500 让上层看日志
            raise RuntimeError("INSERT IGNORE 后未读到 material_storage 记录")
        # 可选：若行已存在且缺少 order_shoe_id，这里再补一次
        if (not st.order_shoe_id) and order_shoe_id:
            st.order_shoe_id = order_shoe_id
            db.session.flush()
        return st

    # ---------- 主流程 ----------
    t0 = time.time()
    logger.logger.info("【盘库确认】开始")

    data = request.get_json(force=True) or {}
    rec_id = data.get("makeInventoryRecordId")
    if not rec_id:
        logger.logger.warning("【盘库确认】缺少 makeInventoryRecordId")
        return jsonify({"code": 400, "message": "makeInventoryRecordId 必填"}), 400

    rec = MakeInventoryRecord.query.get(rec_id)
    if not rec:
        logger.logger.warning(f"【盘库确认】记录不存在: rec_id={rec_id}")
        return jsonify({"code": 404, "message": "记录不存在"}), 404
    if rec.make_inventory_status != 1:
        logger.logger.warning(f"【盘库确认】状态非法(需=1已回传): status={rec.make_inventory_status}, rid={rec.make_inventory_rid}")
        return jsonify({"code": 400, "message": "未回传Excel，无法确认盘库"}), 400

    rid = rec.make_inventory_rid
    logger.logger.info(f"【盘库确认】记录匹配: rec_id={rec_id}, rid={rid}")

    # 盘库日期（出库过滤用）
    inv_date = rec.record_date
    if isinstance(inv_date, str):
        inv_date = datetime.strptime(inv_date[:10], "%Y-%m-%d").date()
    elif isinstance(inv_date, datetime):
        inv_date = inv_date.date()
    elif not isinstance(inv_date, date):
        inv_date = datetime.now().date()
    inv_dt_end = datetime.combine(inv_date, datetime.max.time())

    # 最近一次回传文件
    pattern = os.path.join(IMPORT_DIR, f"回传_{rid}_*.xlsx")
    candidates = sorted(glob.glob(pattern), key=os.path.getmtime, reverse=True)
    logger.logger.info(f"【盘库确认】查找回传Excel: pattern={pattern}, found={len(candidates)}")
    if not candidates:
        return jsonify({"code": 400, "message": "未找到该记录的回传Excel"}), 400
    uploaded_path = candidates[0]
    logger.logger.info(f"【盘库确认】使用回传文件: {uploaded_path}")

    # 读取回传表
    t1 = time.time()
    try:
        after_map = _read_visible_sheet_to_map(uploaded_path)
    except Exception as e:
        logger.logger.exception(f"【盘库确认】解析Excel失败: {e}")
        return jsonify({"code": 400, "message": f"解析Excel失败: {e}"}), 400
    logger.logger.info(f"【盘库确认】解析回传完成，行数={len(after_map)}，耗时={time.time() - t1:.2f}s")

    now = datetime.now()
    ts_tag = now.strftime("%Y%m%d%H%M%S")

    try:
        # ==========================
        # A) 出库（<= 盘库日期）
        # ==========================
        ta = time.time()
        logger.logger.info("【盘库确认】A.开始出库快照查询(首入库时间过滤)")

        first_inbound_dt_subq = (
            db.session.query(
                InboundRecordDetail.material_storage_id.label("sid"),
                func.min(InboundRecord.inbound_datetime).label("first_inbound_dt")
            )
            .join(InboundRecord, InboundRecord.inbound_record_id == InboundRecordDetail.inbound_record_id)
            .group_by(InboundRecordDetail.material_storage_id)
        ).subquery()

        snap_q = (
            db.session.query(
                MaterialStorage.material_storage_id.label("sid"),
                MaterialStorage.spu_material_id.label("spu_id"),
                MaterialStorage.current_amount.label("qty"),
                MaterialStorage.unit_price.label("uprice"),
                Material.material_supplier.label("supplier_id"),
                MaterialType.warehouse_id.label("warehouse_id"),
            )
            .join(SPUMaterial, SPUMaterial.spu_material_id == MaterialStorage.spu_material_id)
            .join(Material, Material.material_id == SPUMaterial.material_id)
            .join(MaterialType, MaterialType.material_type_id == Material.material_type_id)
            .outerjoin(first_inbound_dt_subq, first_inbound_dt_subq.c.sid == MaterialStorage.material_storage_id)
            .filter(MaterialStorage.current_amount > 0)
            .filter(func.coalesce(first_inbound_dt_subq.c.first_inbound_dt, datetime(1970, 1, 1)) <= inv_dt_end)
        )

        snapshots = []
        for row in snap_q.all():
            qty = to_dec(row.qty)
            if qty <= 0:
                continue
            snapshots.append({
                "sid": int(row.sid),
                "spu_id": int(row.spu_id),
                "qty": qty,
                "uprice": to_dec(row.uprice),
                "supplier_id": (int(row.supplier_id) if row.supplier_id is not None else None),
                "warehouse_id": (int(row.warehouse_id) if row.warehouse_id is not None else None),
            })
        logger.logger.info(f"【盘库确认】A.快照完成 count={len(snapshots)}，耗时={time.time() - ta:.2f}s")

        # 分组建出库单
        out_groups = {}
        for s in snapshots:
            out_groups.setdefault((s["supplier_id"], s["warehouse_id"]), []).append(s)
        logger.logger.info(f"【盘库确认】A.出库分组数={len(out_groups)}")

        created_out_records = []
        with db.session.no_autoflush:
            for idx, ((sup_id, wh_id), items) in enumerate(out_groups.items(), start=1):
                orid = f"OR{ts_tag}INV-{idx:02d}"
                outbound = OutboundRecord(
                    outbound_datetime=now,
                    outbound_type=4,
                    outbound_rid=orid,
                    supplier_id=sup_id,
                    picker=None,
                    remark=f"盘库[{rid}]清库出库",
                    outbound_department=None,
                    approval_status=1,
                    outsource_info_id=None,
                    total_price=Decimal("0"),
                    is_sized_material=0,
                )
                db.session.add(outbound); db.session.flush()
                if hasattr(outbound, "warehouse_id"):
                    setattr(outbound, "warehouse_id", wh_id)

                total_out_price = Decimal("0")
                details_payload = []
                for it in items:
                    item_total = it["uprice"] * it["qty"]
                    total_out_price += item_total
                    details_payload.append({
                        "outbound_record_id": outbound.outbound_record_id,
                        "outbound_amount": it["qty"],
                        "remark": "盘库清库",
                        "order_id": None,
                        "order_shoe_id": None,
                        "spu_material_id": it["spu_id"],
                        "material_storage_id": it["sid"],
                        "unit_price": it["uprice"],
                        "item_total_price": item_total,
                    })
                if details_payload:
                    db.session.bulk_insert_mappings(OutboundRecordDetail, details_payload)
                outbound.total_price = total_out_price
                db.session.flush()
                created_out_records.append(outbound)

            # 批量清零库存与尺码列
            if snapshots:
                id_list = [s["sid"] for s in snapshots]
                for ids in _chunk(id_list, 1000):
                    db.session.execute(
                        update(MaterialStorage)
                        .where(MaterialStorage.material_storage_id.in_(ids))
                        .values(current_amount=Decimal("0"))
                    )
                if SHOESIZERANGE:
                    set_frag = ", ".join([f"size_{sz}_current_amount=0" for sz in SHOESIZERANGE])
                    for ids in _chunk(id_list, 1000):
                        db.session.execute(
                            text(f"UPDATE material_storage SET {set_frag} WHERE material_storage_id IN :ids"),
                            {"ids": tuple(ids)}
                        )

        # ==========================
        # B) 回填入库（聚合+复用）
        # ==========================
        tb = time.time()
        logger.logger.info("【盘库确认】B.开始入库聚合（SPU+单位+订单[ID]）")

        agg = {}             # key=(spu_id, unit, order_id) -> {qty, value, order_shoe_id, __row}
        invalid_rows = []    # 字段不足/不允许创建
        missing_rows = []    # 类型无仓库，记录可定位信息

        # 逐行聚合
        for _, row in after_map.items():
            qty   = to_dec(row.get("totalCurrentAmount"))
            if qty <= 0:
                continue
            try:
                value = to_dec(row.get("totalValueAmount"))
                avg   = to_dec(row.get("avgUnitPrice"))
                spu_id = _ensure_spu_only(row)  # 可能抛 ValueError
                unit   = _s(row.get("unit"))

                # 先判仓库是否缺失（用于精确回报 Excel 行）
                sup_id, wh_id = _supplier_and_wh_by_spu(spu_id)
                if wh_id is None:
                    missing_rows.append({
                        "rowNo": row.get("_rowNo"),
                        "spuRid": _s(row.get("spuRid")),
                        "type": _s(row.get("type")),
                        "name": _s(row.get("name")),
                        "model": _s(row.get("model")),
                        "specification": _s(row.get("specification")),
                        "color": _s(row.get("color")),
                        "unit": unit,
                        "orderRid": _s(row.get("orderRid"))
                    })
                    # 不参与聚合，直接让后面统一返回 400
                    continue

                # 订单
                order_rid = _s(row.get("orderRid"))
                order_id = None
                order_shoe_id = None
                if order_rid:
                    o = db.session.query(Order).filter_by(order_rid=order_rid).first()
                    order_id = o.order_id if o else None
                    if order_id:
                        ost = db.session.query(OrderShoe).filter_by(order_id=order_id).first()
                        order_shoe_id = ost.order_shoe_id if ost else None

                key = (spu_id, unit, order_id)
                amt = value if value > 0 else (avg * qty)
                acc = agg.get(key)
                if not acc:
                    agg[key] = {
                        "qty": qty,
                        "value": amt,
                        "order_shoe_id": order_shoe_id,
                        "__row": row  # 用于必要时的追溯
                    }
                else:
                    acc["qty"]   += qty
                    acc["value"] += amt
                    if not acc.get("order_shoe_id") and order_shoe_id:
                        acc["order_shoe_id"] = order_shoe_id

            except ValueError as ve:
                invalid_rows.append({
                    "rowNo": row.get("_rowNo"),
                    "spuRid": _s(row.get("spuRid")),
                    "type": _s(row.get("type")),
                    "name": _s(row.get("name")),
                    "model": _s(row.get("model")),
                    "specification": _s(row.get("specification")),
                    "color": _s(row.get("color")),
                    "unit": _s(row.get("unit")),
                    "orderRid": _s(row.get("orderRid")),
                    "reason": str(ve),
                })

        if invalid_rows:
            logger.logger.error(f"【盘库确认】存在 {len(invalid_rows)} 条无效行，拒绝继续（防止创建空 SPU）")
            invalid_rows.sort(key=lambda x: (x["rowNo"] is None, x["rowNo"] or 0))
            return jsonify({
                "code": 400,
                "message": "回传表中存在字段不足或非法的行，已拒绝创建空SPU，请修正后重试",
                "invalidRows": invalid_rows[:500],
                "invalidCount": len(invalid_rows)
            }), 400

        if missing_rows:
            logger.logger.error(f"【盘库确认】存在 {len(missing_rows)} 行数据的材料类型未配置 warehouse_id，拒绝入库")
            missing_rows.sort(key=lambda x: (x["rowNo"] is None, x["rowNo"] or 0))
            return jsonify({
                "code": 400,
                "message": "有材料类型未配置仓库（warehouse_id），请先维护后重试",
                "missingRows": missing_rows[:500],
                "missingCount": len(missing_rows)
            }), 400

        logger.logger.info(f"【盘库确认】B.入库聚合完成 key数={len(agg)} 耗时={time.time() - tb:.2f}s")

        # 创建/复用 storage，并按 (supplier, warehouse) 分组
        in_groups = {}
        created_new_storages, reused_storages = [], []

        for (spu_id, unit, order_id), acc in agg.items():
            qty = acc["qty"]
            total_value = acc["value"]
            unit_price = (total_value / qty) if qty and qty != Decimal("0") else Decimal("0")
            order_shoe_id = acc.get("order_shoe_id")

            st = None
            if order_id is not None:
                st = (db.session.query(MaterialStorage)
                        .filter(MaterialStorage.spu_material_id == spu_id,
                                MaterialStorage.actual_inbound_unit == unit,
                                MaterialStorage.order_id == order_id)
                        .first())
                if st:
                    if (not st.order_shoe_id) and order_shoe_id:
                        st.order_shoe_id = order_shoe_id
                    reused_storages.append(st.material_storage_id)

            st = _get_or_upsert_storage(
                spu_id=spu_id,
                unit=unit,
                order_id=order_id,
                order_shoe_id=order_shoe_id,
                unit_price=unit_price
            )
            # 后续照旧：把 st 丢进 in_groups / 构造明细 / 批量加量
            sup_id, wh_id = _supplier_and_wh_by_spu(spu_id)
            in_groups.setdefault((sup_id, wh_id), []).append({
                "storage": st,
                "qty": qty,
                "unit_price": unit_price,
                "total_price": total_value,
                "row": acc.get("__row"),
            })

            sup_id, wh_id = _supplier_and_wh_by_spu(spu_id)
            in_groups.setdefault((sup_id, wh_id), []).append({
                "storage": st,
                "qty": qty,
                "unit_price": unit_price,
                "total_price": total_value,
                "row": acc.get("__row"),   # 追溯 Excel 原始行
            })

        logger.logger.info(f"【盘库确认】B.storage 创建={len(created_new_storages)} 复用={len(reused_storages)}")

        # 入库单生成
        created_in_records = []
        with db.session.no_autoflush:
            bad_rows = []
            for (sup_id, wh_id), items in in_groups.items():
                logger.logger.info(f"【盘库确认】B.入库分组 supplier_id={sup_id} warehouse_id={wh_id} 行数={len(items)}")
                if wh_id is None:
                    logger.logger.warning(f"【盘库确认】B.入库分组 supplier_id={sup_id} warehouse_id={wh_id} 行数={len(items)}有误")
                    for it in items:
                        r = (it.get("row") or {})  # 原始 Excel 行
                        bad_rows.append({
                            "rowNo": r.get("_rowNo"),
                            "spuRid": _s(r.get("spuRid")),
                            "type": _s(r.get("type")),
                            "name": _s(r.get("name")),
                            "model": _s(r.get("model")),
                            "specification": _s(r.get("specification")),
                            "color": _s(r.get("color")),
                            "unit": _s(r.get("unit")),
                            "orderRid": _s(r.get("orderRid")),
                        })

            if bad_rows:
                bad_rows.sort(key=lambda x: (x["rowNo"] is None, x["rowNo"] or 0))
                logger.logger.error(f"【盘库确认】存在 {len(bad_rows)} 行所属类型未配置 warehouse_id，拒绝入库")
                return jsonify({
                    "code": 400,
                    "message": "有材料类型未配置仓库（warehouse_id），请到【材料类型】维护仓库后重试",
                    "missingRows": bad_rows[:500],   # 限流避免响应过大
                    "missingCount": len(bad_rows)
                }), 400
            for idx, ((sup_id, wh_id), items) in enumerate(in_groups.items(), start=1):
                if wh_id is None:
                    # 理论上不会触发（前面已拦截），但双重保险
                    return jsonify({"code": 400, "message": "检测到未配置仓库的类型，已终止"}), 400

                irid = f"IR{ts_tag}INV-{idx:02d}"
                inbound = InboundRecord(
                    inbound_datetime=now,
                    inbound_type=4,
                    inbound_rid=irid,
                    inbound_batch_id=0,
                    supplier_id=sup_id,
                    warehouse_id=wh_id,
                    remark=f"盘库[{rid}]回填入库",
                    pay_method=None,
                    total_price=Decimal("0"),
                    approval_status=1,
                    is_sized_material=0,
                    # 如模型需要：approval_status=0, staff_id=None 等按你的表结构补
                )
                db.session.add(inbound); db.session.flush()

                total_in_price = Decimal("0")
                details_payload, inc_updates = [], []
                for it in items:
                    st = it["storage"]
                    qty = it["qty"]
                    unit_price = it["unit_price"]
                    item_total = it["total_price"]
                    total_in_price += item_total
                    details_payload.append({
                        "inbound_record_id": inbound.inbound_record_id,
                        "inbound_amount": qty,
                        "remark": "盘库入库",
                        "order_id": st.order_id,
                        "order_shoe_id": st.order_shoe_id,
                        "spu_material_id": st.spu_material_id,
                        "material_storage_id": st.material_storage_id,
                        "unit_price": unit_price,
                        "item_total_price": item_total,
                    })
                    inc_updates.append((st.material_storage_id, qty, unit_price))

                if details_payload:
                    db.session.bulk_insert_mappings(InboundRecordDetail, details_payload)

                # 批量加量/更新价格
                for batch in _chunk(inc_updates, 500):
                    ids = [x[0] for x in batch]
                    case_qty   = " ".join([f"WHEN {sid} THEN {str(qty)}" for sid, qty, _ in batch])
                    case_price = " ".join([f"WHEN {sid} THEN {str(up)}"  for sid, _, up in batch])
                    db.session.execute(
                        text(
                            f"""
                            UPDATE material_storage
                            SET inbound_amount = inbound_amount + CASE material_storage_id {case_qty} END,
                                current_amount = current_amount + CASE material_storage_id {case_qty} END,
                                unit_price     = CASE material_storage_id {case_price} END,
                                average_price  = CASE material_storage_id {case_price} END
                            WHERE material_storage_id IN :ids
                            """
                        ),
                        {"ids": tuple(ids)}
                    )

                inbound.total_price = total_in_price
                db.session.flush()
                created_in_records.append(inbound)

        # ==========================
        # C) 落盘 & 提交
        # ==========================
        rec.outbound_record_id = created_out_records[0].outbound_record_id if created_out_records else None
        rec.inbound_record_id  = created_in_records[0].inbound_record_id  if created_in_records else None
        rec.make_inventory_status = 2  # 等待确认
        db.session.commit()

        logger.logger.info(f"【盘库确认】完成：出库单={len(created_out_records)} 入库单={len(created_in_records)} 总耗时={time.time() - t0:.2f}s")

        return jsonify({
            "code": 200,
            "message": "ok",
            "outboundRecords": [
                {
                    "id": r.outbound_record_id,
                    "rid": r.outbound_rid,
                    "supplierId": r.supplier_id,
                    "warehouseId": getattr(r, "warehouse_id", None),
                    "totalPrice": str(r.total_price or "0"),
                } for r in created_out_records
            ],
            "inboundRecords": [
                {
                    "id": r.inbound_record_id,
                    "rid": r.inbound_rid,
                    "supplierId": r.supplier_id,
                    "warehouseId": r.warehouse_id,
                    "totalPrice": str(r.total_price or "0"),
                } for r in created_in_records
            ],
            "savedOnRecord": {
                "outboundRecordId": rec.outbound_record_id,
                "inboundRecordId": rec.inbound_record_id
            }
        })

    except Exception as e:
        logger.logger.exception(f"【盘库确认】发生异常，将回滚。错误：{e}")
        db.session.rollback()
        return jsonify({"code": 500, "message": f"确认盘库失败: {e}"}), 500




@make_inventory_bp.route("/warehouse/makeinventory/create", methods=["POST"])
def create_make_inventory():
    data = request.get_json() or {}
    record_date_s = (data.get("recordDate") or "").strip()
    reason = data.get("reason") or ""

    rid = _gen_rid()

    record_date = None
    if record_date_s:
        try:
            record_date = datetime.strptime(record_date_s, "%Y-%m-%d").date()
        except Exception:
            return jsonify({"code": 400, "message": "recordDate 格式应为 YYYY-MM-DD"}), 400

    record = MakeInventoryRecord(
        record_date=record_date,
        make_inventory_rid=rid,
        make_inventory_reason=reason,
        make_inventory_status=0,     # 未回传
        excel_reupload_status=0
    )
    db.session.add(record)
    db.session.commit()

    return jsonify({
        "code": 200,
        "message": "ok",
        "makeInventoryRecordId": record.make_inventory_record_id,
        "makeInventoryRid": record.make_inventory_rid,
    })

# ====== 导出（落盘后下载）======
@make_inventory_bp.route("/warehouse/inventorysummary/export", methods=["GET"])
def export_inventory_summary():
    keyword          = (request.args.get("keyword") or "").strip()
    sort_by          = (request.args.get("sortBy") or "total_current_amount").strip()
    sort_order       = (request.args.get("sortOrder") or "desc").strip().lower()
    inventory_date   = (request.args.get("inventoryDate") or "").strip()
    inventory_reason = (request.args.get("inventoryReason") or "").strip()
    record_id        = request.args.get("recordId", type=int)

    record = MakeInventoryRecord.query.get(record_id) if record_id else None

    _ensure_dir(EXPORT_DIR)
    filename = _inventory_export_filename(
        record=record, inventory_date=inventory_date, inventory_reason=inventory_reason
    )
    file_path = os.path.join(EXPORT_DIR, filename)

    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True, download_name=filename)

    rows, size_range = query_inventory_summary(
        keyword=keyword, sort_by=sort_by, sort_order=sort_order
    )

    bio = build_inventory_excel(
        rows=rows,
        size_range=size_range,
        inventory_date=inventory_date,
        inventory_reason=inventory_reason
    )
    bio.seek(0)
    with open(file_path, "wb") as f:
        f.write(bio.read())

    return send_file(
        file_path,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ====== 汇总查询（用于导出 & 对话框列表）======
def query_inventory_summary(*, keyword:str="", sort_by:str="total_current_amount", sort_order:str="desc"):
    size_sum_exprs = _sum_current_amount_cols()
    qty_sum_expr   = func.coalesce(func.sum(MaterialStorage.current_amount), 0)
    value_sum_expr = func.coalesce(func.sum(MaterialStorage.average_price * MaterialStorage.current_amount), 0)
    avg_price_expr = (value_sum_expr / func.nullif(qty_sum_expr, 0))

    base_q = (
        db.session.query(
            SPUMaterial.spu_material_id.label("spu_material_id"),
            SPUMaterial.spu_rid.label("spu_rid"),
            MaterialType.material_type_name.label("material_type_name"),
            Material.material_name.label("material_name"),
            Supplier.supplier_name.label("material_supplier"),
            SPUMaterial.material_model.label("material_model"),
            SPUMaterial.material_specification.label("material_specification"),
            SPUMaterial.color.label("color"),
            MaterialStorage.actual_inbound_unit.label("unit"),
            Order.order_rid.label("order_rid"),
            qty_sum_expr.label("total_current_amount"),
            value_sum_expr.label("total_value_amount"),
            avg_price_expr.label("avg_unit_price"),
            *size_sum_exprs
        )
        .select_from(MaterialStorage)
        .join(SPUMaterial, SPUMaterial.spu_material_id == MaterialStorage.spu_material_id)
        .join(Material, Material.material_id == SPUMaterial.material_id)
        .join(MaterialType, MaterialType.material_type_id == Material.material_type_id)
        .join(Supplier, Supplier.supplier_id == Material.material_supplier)
        .join(Order, Order.order_id == MaterialStorage.order_id, isouter=True)
    )

    if keyword:
        like = f"%{keyword}%"
        base_q = base_q.filter(
            or_(
                SPUMaterial.material_model.ilike(like),
                SPUMaterial.material_specification.ilike(like),
                SPUMaterial.color.ilike(like),
                SPUMaterial.spu_rid.ilike(like),
                Material.material_name.ilike(like),
                MaterialType.material_type_name.ilike(like),
                Supplier.supplier_name.ilike(like),
            )
        )

    base_q = base_q.group_by(
        SPUMaterial.spu_material_id,
        SPUMaterial.spu_rid,
        MaterialType.material_type_name,
        Material.material_name,
        Supplier.supplier_name,
        MaterialStorage.actual_inbound_unit,
        SPUMaterial.material_model,
        SPUMaterial.material_specification,
        SPUMaterial.color,
        Order.order_rid
    ).having(qty_sum_expr > 0)

    sub = base_q.subquery()
    SORTABLE = {
        "total_current_amount": sub.c.total_current_amount,
        "total_value_amount":   sub.c.total_value_amount,
        "avg_unit_price":       sub.c.avg_unit_price,
        "spu_material_id":      sub.c.spu_material_id,
        "spu_rid":              sub.c.spu_rid,
        "material_type_name":   sub.c.material_type_name,
        "material_name":        sub.c.material_name,
        "material_model":       sub.c.material_model,
        "material_specification": sub.c.material_specification,
        "color":                sub.c.color,
        "order_rid":            sub.c.order_rid,
    }
    for sz in SHOESIZERANGE:
        SORTABLE[f"size_{sz}"] = getattr(sub.c, f"size_{sz}")

    sort_col = SORTABLE.get(sort_by, sub.c.total_current_amount)
    order_by_expr = sort_col.desc() if (sort_order or "desc").lower() == "desc" else sort_col.asc()

    rows = db.session.query(sub).order_by(order_by_expr).all()
    return rows, SHOESIZERANGE

# ====== 导入（与服务器导出对比；缺失/新增行视为减少/新增材料）======
@make_inventory_bp.route("/warehouse/inventorysummary/import", methods=["POST"])
def import_inventory_summary():
    """
    接收前端上传的盘库 Excel：
      - 需要 recordId 以找到服务器上最近一次导出的 Excel 作为基线
      - 保存上传文件到 IMPORT_DIR
      - 逐行对比（主键：SPU_MATERIAL_ID），新增/缺失行分别标记为 added / removed
      - 生成 diff.json 存到 IMPORT_DIR/diff/<RID>_diff.json，并把 diff 返回给前端
      - 状态：0 -> 1（已回传），excel_reupload_status = 1
    返回：{ code, message, diff, diffSavedAs, updated, failed, errors }
    """
    file = request.files.get("file")
    if not file:
        return jsonify({"code": 400, "message": "缺少文件"}), 400

    record_id = request.form.get("recordId", type=int)
    if not record_id:
        return jsonify({"code": 400, "message": "recordId 必填（用于找到基线导出文件）"}), 400

    record = MakeInventoryRecord.query.get(record_id)
    if not record:
        return jsonify({"code": 404, "message": "盘库记录不存在"}), 404

    # 1) 基线文件：若找不到，视为“全新增”的对比（baseline 为空）
    baseline_path = _find_baseline_export_file(record)  # 可能为 None
    # 2) 保存本次上传
    original = secure_filename(file.filename or "inventory.xlsx")
    upload_name = f"回传_{record.make_inventory_rid}_{original}"
    upload_path = os.path.join(IMPORT_DIR, upload_name)
    file.save(upload_path)

    # 3) 解析两份文件
    try:
        before_map = {}
        if baseline_path and os.path.exists(baseline_path):
            before_map = _read_visible_sheet_to_map(baseline_path)
    except Exception as e:
        current_app.logger.exception("解析基线文件失败: %s", baseline_path)
        return jsonify({"code": 400, "message": f"解析基线Excel失败: {e}"}), 400

    try:
        after_map = _read_visible_sheet_to_map(upload_path)
    except Exception as e:
        current_app.logger.exception("解析上传文件失败: %s", upload_path)
        return jsonify({"code": 400, "message": f"解析上传Excel失败: {e}"}), 400

    # ==== 兼容新格式（含“订单号”）并按【SPU + 订单号】进行对比 ====
    # 说明：
    # - 旧版基线可能没有“订单号”，则按空字符串对待（即视为“无订单”维度）
    # - 新版上传包含“orderRid”（由导出函数写入第二列表头“订单号”）
    # - 我们在这里不强依赖 _read_visible_sheet_to_map 的键，只要 value 里带有 spuRid/orderRid/…即可
    def _to_float(v):
        if v is None:
            return 0.0
        if isinstance(v, (int, float, Decimal)):
            return float(v)
        try:
            s = str(v).strip().replace(",", "").replace("，", "").replace("\u3000", "")
            return float(s) if s else 0.0
        except Exception:
            return 0.0

    def _norm_order_rid(x):
        # 导入里“订单号”可能叫 orderRid / order_rid；都做兼容
        return (x.get("orderRid") or x.get("order_rid") or "").strip()

    def _norm_spu_id_key(k, v):
        # _read_visible_sheet_to_map 的 key 旧实现是 spuMaterialId；
        # 如果不是，从值里兜底取一次
        sid = k or v.get("spuMaterialId") or v.get("spu_material_id")
        try:
            return int(sid)
        except Exception:
            return str(sid or "").strip()

    def _remap_by_composite(src: dict):
        # 把 {sid: row} 变成 { (sid, orderRid): row }；orderRid 为空用 ""
        out = {}
        for k, v in (src or {}).items():
            sid = _norm_spu_id_key(k, v)
            ord_rid = _norm_order_rid(v)  # 旧基线没有订单号时为 ""
            out[(sid, ord_rid)] = v
        return out

    before_map2 = _remap_by_composite(before_map)
    after_map2  = _remap_by_composite(after_map)

    # 4) 生成 diff（按 (SPU, 订单号) 维度）
    items = []
    changed_count = added_count = removed_count = 0
    sum_delta_qty = 0.0
    sum_delta_value = 0.0

    keys = set(before_map2.keys()) | set(after_map2.keys())

    def sizes_delta(bsizes: dict, asizes: dict):
        all_keys = set((bsizes or {}).keys()) | set((asizes or {}).keys())
        return {k: _to_float((asizes or {}).get(k, 0)) - _to_float((bsizes or {}).get(k, 0)) for k in all_keys}

    for (sid, ord_rid) in sorted(keys, key=lambda x: (str(x[0]), str(x[1]))):
        b = before_map2.get((sid, ord_rid))
        a = after_map2.get((sid, ord_rid))

        # 统一把 orderRid 填回 before/after，便于前端展示
        if b is not None:
            b = {**b, "orderRid": ord_rid}
        if a is not None:
            a = {**a, "orderRid": ord_rid}

        if b and not a:
            # 移除（上传缺失该 (SPU, 订单号) 组合）
            removed_count += 1
            delta_qty   = -_to_float(b.get("totalCurrentAmount"))
            delta_value = -_to_float(b.get("totalValueAmount"))
            sum_delta_qty   += delta_qty
            sum_delta_value += delta_value

            items.append({
                "type": "removed",
                "spuMaterialId": sid,
                "orderRid": ord_rid,          # 新增：订单号
                "spuRid": b.get("spuRid"),
                "unit":  b.get("unit"),
                "before": b,
                "after":  None,
                "delta": {
                    "totalCurrentAmount": delta_qty,
                    "totalValueAmount":   delta_value,
                    "sizes": sizes_delta(b.get("sizes", {}), {})
                }
            })
            continue

        if a and not b:
            # 新增（上传新增该 (SPU, 订单号) 组合）
            added_count += 1
            delta_qty   = _to_float(a.get("totalCurrentAmount"))
            delta_value = _to_float(a.get("totalValueAmount"))
            sum_delta_qty   += delta_qty
            sum_delta_value += delta_value

            items.append({
                "type": "added",
                "spuMaterialId": sid,
                "orderRid": ord_rid,          # 新增：订单号
                "spuRid": a.get("spuRid"),
                "unit":  a.get("unit"),
                "before": None,
                "after":  a,
                "delta": {
                    "totalCurrentAmount": delta_qty,
                    "totalValueAmount":   delta_value
                },
                "isNewMaterial": not a.get("_resolvedToSpu", False)
            })
            continue

        # 常规变更
        delta_total = _to_float(a.get("totalCurrentAmount")) - _to_float(b.get("totalCurrentAmount"))
        delta_value = _to_float(a.get("totalValueAmount"))   - _to_float(b.get("totalValueAmount"))
        delta_sizes = sizes_delta(b.get("sizes", {}), a.get("sizes", {}))

        changed = (
            abs(delta_total) > 1e-9
            or abs(delta_value) > 1e-9
            or any(abs(v) > 1e-9 for v in delta_sizes.values())
        )
        if changed:
            changed_count += 1
            sum_delta_qty   += delta_total
            sum_delta_value += delta_value

        items.append({
            "type": "changed" if changed else "unchanged",
            "spuMaterialId": sid,
            "orderRid": ord_rid,              # 新增：订单号
            "spuRid": a.get("spuRid") or b.get("spuRid"),
            "unit":  a.get("unit")  or b.get("unit"),
            "before": b,
            "after":  a,
            "delta": {
                "totalCurrentAmount": delta_total,
                "totalValueAmount":   delta_value,
                "sizes":              delta_sizes
            }
        })

    diff = {
        "recordId": record_id,
        "rid": record.make_inventory_rid,
        "baselineFile": os.path.basename(baseline_path) if baseline_path else "",
        "uploadedFile": upload_name,
        "summary": {
            "added":   added_count,
            "removed": removed_count,
            "changed": changed_count,
            "totalDeltaQty":   sum_delta_qty,
            "totalDeltaValue": sum_delta_value
        },
        "items": items
    }

    # 5) 保存 diff.json
    diff_name = f"{record.make_inventory_rid}_diff.json"
    diff_path = os.path.join(DIFF_DIR, diff_name)
    with open(diff_path, "w", encoding="utf-8") as f:
        json.dump(diff, f, ensure_ascii=False, indent=2)

    # 6) 状态更新：0->1；再次回传 excel_reupload_status=1
    if record.make_inventory_status == 0:
        record.make_inventory_status = 1
    record.excel_reupload_status = 1
    db.session.commit()

    preview_size = 50
    items_preview = items[:preview_size]
    is_truncated = len(items) > preview_size

    # 6) 状态更新略...

    return jsonify({
        "code": 200,
        "message": "ok",
        "updated": 0, "failed": 0, "errors": [],
        # 只回摘要，不回整份 items
        "diffSummary": diff.get("summary", {}),
        "diffSavedAs": diff_name,
        "preview": items_preview,
        "previewTruncated": is_truncated
    })
    
# ===== Diff 预览：分页 / 分组 / 搜索 / 排序 =====
@make_inventory_bp.route("/warehouse/inventorysummary/diff/<rid>", methods=["GET"])
def get_inventory_diff(rid):
    """
    分页读取导入时保存的 diff.json
    支持：
      - 分页：page, pageSize
      - 分组(type)：added / removed / changed / unchanged
      - 关键字：q（在 spuRid / orderRid / 名称/型号/规格/颜色/单位 中模糊匹配）
      - 排序：sortBy, sortOrder（asc/desc）
    返回：
      { code, message, rid, baselineFile, uploadedFile, summary, counts, page, pageSize, total, items }
    """
    # ---- 读取参数 ----
    page = max(int(request.args.get("page", 1)), 1)
    page_size = max(int(request.args.get("pageSize", 100)), 1)
    group = (request.args.get("type") or "").strip().lower()  # added/removed/changed/unchanged
    q = (request.args.get("q") or request.args.get("keyword") or "").strip()
    sort_by = (request.args.get("sortBy") or "spuRid").strip()
    sort_order = (request.args.get("sortOrder") or "asc").strip().lower()
    sort_order = "desc" if sort_order not in ("asc", "desc") else sort_order

    diff_path = os.path.join(DIFF_DIR, f"{rid}_diff.json")
    if not os.path.exists(diff_path):
        return jsonify({"code": 404, "message": "diff 文件不存在", "rid": rid}), 404

    with open(diff_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    # ---- 头部信息（兼容旧字段名）----
    header = {
        "rid": rid,
        "baselineFile": data.get("baselineFile", ""),
        "uploadedFile": data.get("uploadedFile", data.get("uploaded_file", "")),
        "summary": data.get("summary", {}),
        "recordId": data.get("recordId"),
    }

    items = data.get("items", []) or []

    # ---- 页签计数（未过滤前计算）----
    def _count_by_type(arr):
        c = {"added": 0, "removed": 0, "changed": 0, "unchanged": 0}
        for it in arr:
            tp = (it.get("type") or "").lower()
            if tp in c:
                c[tp] += 1
        return c
    counts = _count_by_type(items)

    # ---- 过滤：分组(type) ----
    if group in ("added", "removed", "changed", "unchanged"):
        items = [it for it in items if (it.get("type") or "").lower() == group]

    # ---- 关键字检索（宽松字段）----
    def _s(v):
        # 安全转字符串
        if v is None:
            return ""
        try:
            return str(v)
        except Exception:
            return ""
    def _field(it, name):
        # 兼容放在根 / before / after 上的字段
        return it.get(name) or (it.get("after") or {}).get(name) or (it.get("before") or {}).get(name) or ""

    if q:
        ql = q.lower()
        filtered = []
        for it in items:
            hay = " ".join([
                _s(it.get("spuRid") or _field(it, "spuRid")),
                _s(it.get("orderRid") or _field(it, "orderRid")),
                _s(_field(it, "name")),
                _s(_field(it, "model")),
                _s(_field(it, "specification")),
                _s(_field(it, "color")),
                _s(_field(it, "unit")),
            ]).lower()
            if ql in hay:
                filtered.append(it)
        items = filtered

    # ---- 排序 ----
    def _num(x):
        try:
            return float(x)
        except Exception:
            return 0.0

    def _key(it):
        # 可扩展的排序键
        if sort_by == "orderRid":
            return _s(it.get("orderRid") or _field(it, "orderRid"))
        if sort_by == "type":
            return _s(it.get("type"))
        if sort_by in ("deltaQty", "deltaTotal", "totalCurrentAmountDelta"):
            return _num((it.get("delta") or {}).get("totalCurrentAmount"))
        if sort_by in ("deltaValue", "totalValueAmountDelta"):
            return _num((it.get("delta") or {}).get("totalValueAmount"))
        if sort_by == "name":
            return _s(_field(it, "name"))
        if sort_by == "model":
            return _s(_field(it, "model"))
        if sort_by == "specification":
            return _s(_field(it, "specification"))
        if sort_by == "color":
            return _s(_field(it, "color"))
        # 默认按 spuRid
        return _s(it.get("spuRid") or _field(it, "spuRid"))

    items.sort(key=_key, reverse=(sort_order == "desc"))

    # ---- 分页 ----
    total = len(items)
    start = (page - 1) * page_size
    end = min(start + page_size, total)
    page_items = items[start:end]

    # ---- 返回 ----
    return jsonify({
        "code": 200,
        "message": "ok",
        "rid": header["rid"],
        "baselineFile": header["baselineFile"],
        "uploadedFile": header["uploadedFile"],
        "summary": header["summary"],
        "counts": counts,           # 页签计数（全量，方便前端 tab 上显示）
        "page": page,
        "pageSize": page_size,
        "total": total,
        "items": page_items,
    })

