from __future__ import annotations

from datetime import datetime, timedelta, date, time as dt_time
from decimal import Decimal
import json
import os
import time
from typing import Optional, Dict, List

from flask import Blueprint, current_app, jsonify, request, send_file, abort, Response

from app_config import db
from constants import *
from models import *
from sqlalchemy import func, select, or_, and_, desc, literal, case
from sqlalchemy.sql import union_all
from file_locations import FILE_STORAGE_PATH
from general_document.accounting_warehouse_history_excel import (
    generate_accounting_warehouse_excel,
    generate_accounting_warehouse_grouped_excel,
)

material_storage_snapshot_bp = Blueprint("material_storage_snapshot_bp", __name__)


# -----------------------------
# Helpers
# -----------------------------

def _parse_date(s: str) -> date:
    return datetime.strptime(s, "%Y-%m-%d").date()


def _get_base_snapshot_date(target_date: date) -> Optional[date]:
    stmt = select(func.max(MaterialStorageSnapshot.snapshot_date)).where(
        MaterialStorageSnapshot.snapshot_date <= target_date
    )
    return db.session.execute(stmt).scalar_one_or_none()


def add_filter_to_query(query, filters: dict, _: str = "base"):
    """Apply optional filters to a SQLAlchemy selectable (no GROUP BY assumptions)."""
    warehouse_id = filters.get("warehouseIdFilter")
    supplier_name = filters.get("supplierNameFilter")
    material_name = filters.get("materialNameFilter")
    material_model = filters.get("materialModelFilter")
    material_specification = filters.get("materialSpecificationFilter")
    material_color = filters.get("materialColorFilter")
    customer_product_name = filters.get("customerProductNameFilter")
    order_rid = filters.get("orderRidFilter")
    shoe_rid = filters.get("shoeRidFilter")

    if warehouse_id:
        query = query.where(MaterialType.warehouse_id == warehouse_id)
    if supplier_name:
        query = query.where(Supplier.supplier_name.ilike(f"%{supplier_name}%"))
    if material_name:
        query = query.where(Material.material_name.ilike(f"%{material_name}%"))
    if material_model:
        query = query.where(SPUMaterial.material_model.ilike(f"%{material_model}%"))
    if material_specification:
        query = query.where(
            SPUMaterial.material_specification.ilike(f"%{material_specification}%")
        )
    if material_color:
        query = query.where(SPUMaterial.color.ilike(f"%{material_color}%"))
    if customer_product_name:
        query = query.where(
            OrderShoe.customer_product_name.ilike(f"%{customer_product_name}%")
        )
    if order_rid:
        query = query.where(Order.order_rid.ilike(f"%{order_rid}%"))
    if shoe_rid:
        query = query.where(Shoe.shoe_rid.ilike(f"%{shoe_rid}%"))

    return query


def _union_msids_subq(base_snapshot_date: date, target_date: date, filters: dict):
    """Return DISTINCT msid subquery entirely in DB (no roundtrip)."""
    # base-side msids (keep at least meaningful ones)
    base_q = (
        select(MaterialStorageSnapshot.material_storage_id.label("msid"))
        .join(SPUMaterial, SPUMaterial.spu_material_id == MaterialStorageSnapshot.spu_material_id)
        .join(Material, SPUMaterial.material_id == Material.material_id)
        .join(MaterialType, MaterialType.material_type_id == Material.material_type_id)
        .join(Supplier, Supplier.supplier_id == Material.material_supplier)
        .outerjoin(Order, MaterialStorageSnapshot.order_id == Order.order_id)
        .outerjoin(OrderShoe, MaterialStorageSnapshot.order_shoe_id == OrderShoe.order_shoe_id)
        .outerjoin(Shoe, Shoe.shoe_id == OrderShoe.shoe_id)
        .where(
            MaterialStorageSnapshot.snapshot_date == base_snapshot_date,
            or_(
                MaterialStorageSnapshot.pending_inbound != 0, # 过滤早期没有入库记录的库存
                MaterialStorageSnapshot.pending_outbound != 0,
                MaterialStorageSnapshot.inbound_amount != 0,
                MaterialStorageSnapshot.outbound_amount != 0,
                MaterialStorageSnapshot.current_amount != 0,
            )
        )
    )
    base_q = add_filter_to_query(base_q, filters, "base")

    # change-window msids (real-time from inbound/outbound records)
    start_change_date = base_snapshot_date + timedelta(days=1)
    start_dt = datetime.combine(start_change_date, dt_time.min)
    end_dt = datetime.combine(target_date + timedelta(days=1), dt_time.min)

    # NOTE:
    #   审核可能发生在“单据发生时间(inbound_datetime/outbound_datetime)”之后很多天。
    #   如果这里只按发生时间过滤，会漏掉“旧单据在窗口内被审核”的情况（典型：月末快照 pending，
    #   之后某天财务审核 -> pending 归零、inbound/outbound 增加）。
    #   因此：msid 候选集需要把 approval_datetime 落在窗口内的记录也纳入。
    inbound_msids_q = (
        select(InboundRecordDetail.material_storage_id.label("msid"))
        .join(InboundRecord, InboundRecord.inbound_record_id == InboundRecordDetail.inbound_record_id)
        .join(MaterialStorage, MaterialStorage.material_storage_id == InboundRecordDetail.material_storage_id)
        .join(SPUMaterial, SPUMaterial.spu_material_id == MaterialStorage.spu_material_id)
        .join(Material, SPUMaterial.material_id == Material.material_id)
        .join(MaterialType, MaterialType.material_type_id == Material.material_type_id)
        .join(Supplier, Supplier.supplier_id == Material.material_supplier)
        .outerjoin(Order, MaterialStorage.order_id == Order.order_id)
        .outerjoin(OrderShoe, MaterialStorage.order_shoe_id == OrderShoe.order_shoe_id)
        .outerjoin(Shoe, Shoe.shoe_id == OrderShoe.shoe_id)
        .where(
            InboundRecord.display == 1,
            or_(
                and_(InboundRecord.inbound_datetime >= start_dt, InboundRecord.inbound_datetime < end_dt),
                and_(InboundRecord.approval_datetime >= start_dt, InboundRecord.approval_datetime < end_dt),
            ),
        )
    )
    inbound_msids_q = add_filter_to_query(inbound_msids_q, filters, "change")

    outbound_msids_q = (
        select(OutboundRecordDetail.material_storage_id.label("msid"))
        .join(OutboundRecord, OutboundRecord.outbound_record_id == OutboundRecordDetail.outbound_record_id)
        .join(MaterialStorage, MaterialStorage.material_storage_id == OutboundRecordDetail.material_storage_id)
        .join(SPUMaterial, SPUMaterial.spu_material_id == MaterialStorage.spu_material_id)
        .join(Material, SPUMaterial.material_id == Material.material_id)
        .join(MaterialType, MaterialType.material_type_id == Material.material_type_id)
        .join(Supplier, Supplier.supplier_id == Material.material_supplier)
        .outerjoin(Order, MaterialStorage.order_id == Order.order_id)
        .outerjoin(OrderShoe, MaterialStorage.order_shoe_id == OrderShoe.order_shoe_id)
        .outerjoin(Shoe, Shoe.shoe_id == OrderShoe.shoe_id)
        .where(
            OutboundRecord.display == 1,
            or_(
                and_(OutboundRecord.outbound_datetime >= start_dt, OutboundRecord.outbound_datetime < end_dt),
                and_(OutboundRecord.approval_datetime >= start_dt, OutboundRecord.approval_datetime < end_dt),
            ),
        )
    )
    outbound_msids_q = add_filter_to_query(outbound_msids_q, filters, "change")

    change_u = union_all(inbound_msids_q, outbound_msids_q).subquery("u_change")
    change_q = select(change_u.c.msid)

    u = union_all(base_q, change_q).subquery("u_all")
    return select(u.c.msid).distinct().subquery("msids_all")


def _final_agg_for_msids_subq(msids_all, base_snapshot_date: date, start_change_date: date, target_date: date):
    """LEFT JOIN base + delta on full msid set, return final_agg subquery."""
    # base snapshot
    base_snap = (
        select(
            MaterialStorageSnapshot.material_storage_id.label("msid"),
            func.coalesce(MaterialStorageSnapshot.pending_inbound, 0).label("base_pending_in"),
            func.coalesce(MaterialStorageSnapshot.pending_outbound, 0).label("base_pending_out"),
            func.coalesce(MaterialStorageSnapshot.inbound_amount, 0).label("base_inbound"),
            func.coalesce(MaterialStorageSnapshot.outbound_amount, 0).label("base_outbound"),
            func.coalesce(MaterialStorageSnapshot.current_amount, 0).label("base_current"),
            func.coalesce(MaterialStorageSnapshot.make_inventory_inbound, 0).label("base_make_inbound"),
            func.coalesce(MaterialStorageSnapshot.make_inventory_outbound, 0).label("base_make_outbound"),
        )
        .where(MaterialStorageSnapshot.snapshot_date == base_snapshot_date)
    ).subquery()

    # interval deltas (approved amounts) + as-of pending amounts (absolute)
    #
    # 关键点：
    #   - pending_* 是“截至 target_date 当天结束仍未审核”的数量，应按 *发生时间* 截止+当前状态(approval_status==0)
    #     直接算成【绝对值】；不能用 base + window_delta 的方式，否则会漏掉“旧单在窗口内被审核”的 pending 下降。
    #   - inbound/outbound/make_inventory/return_out 等“已审核口径”在业务上以 *approval_datetime* 生效；
    #     因此窗口增量必须用 approval_datetime 过滤，才能捕捉“月末 pending、之后某天审核入账”的情况。
    start_dt = datetime.combine(start_change_date, dt_time.min)
    end_dt = datetime.combine(target_date + timedelta(days=1), dt_time.min)

    inbound_agg = (
        select(
            InboundRecordDetail.material_storage_id.label("msid"),
            # 截至 target_date 仍未审核的“待审核入库”绝对值（按发生时间统计）
            func.coalesce(
                func.sum(
                    case(
                        (
                            and_(
                                InboundRecord.approval_status == 0,
                                InboundRecord.inbound_datetime < end_dt,
                            ),
                            InboundRecordDetail.inbound_amount,
                        ),
                        else_=0,
                    )
                ),
                0,
            ).label("pending_in_asof"),
            func.coalesce(
                func.sum(
                    case(
                        (
                            and_(
                                InboundRecord.approval_status == 1,
                                InboundRecord.inbound_type.in_([0, 3]),
                                InboundRecord.approval_datetime >= start_dt,
                                InboundRecord.approval_datetime < end_dt,
                            ),
                            InboundRecordDetail.inbound_amount,
                        ),
                        else_=0,
                    )
                ),
                0,
            ).label("purchase_in"),
            func.coalesce(
                func.sum(
                    case(
                        (
                            and_(
                                InboundRecord.approval_status == 1,
                                InboundRecord.inbound_type == 4,
                                InboundRecord.approval_datetime >= start_dt,
                                InboundRecord.approval_datetime < end_dt,
                            ),
                            InboundRecordDetail.inbound_amount,
                        ),
                        else_=0,
                    )
                ),
                0,
            ).label("make_in"),
        )
        .join(InboundRecord, InboundRecord.inbound_record_id == InboundRecordDetail.inbound_record_id)
        .where(
            InboundRecord.display == 1,
            # pending 需要覆盖所有历史发生记录；approved delta 通过 approval_datetime 限制窗口
            InboundRecord.inbound_datetime < end_dt,
        )
        .group_by(InboundRecordDetail.material_storage_id)
    ).subquery("inbound_agg")

    outbound_agg = (
        select(
            OutboundRecordDetail.material_storage_id.label("msid"),
            # 截至 target_date 仍未审核的“待审核出库”绝对值（按发生时间统计）
            func.coalesce(
                func.sum(
                    case(
                        (
                            and_(
                                OutboundRecord.approval_status == 0,
                                OutboundRecord.outbound_datetime < end_dt,
                            ),
                            OutboundRecordDetail.outbound_amount,
                        ),
                        else_=0,
                    )
                ),
                0,
            ).label("pending_out_asof"),
            func.coalesce(
                func.sum(
                    case(
                        (
                            and_(
                                OutboundRecord.approval_status == 1,
                                OutboundRecord.outbound_type == 0,
                                OutboundRecord.approval_datetime >= start_dt,
                                OutboundRecord.approval_datetime < end_dt,
                            ),
                            OutboundRecordDetail.outbound_amount,
                        ),
                        else_=0,
                    )
                ),
                0,
            ).label("prod_out"),
            func.coalesce(
                func.sum(
                    case(
                        (
                            and_(
                                OutboundRecord.approval_status == 1,
                                OutboundRecord.outbound_type == 5,
                                OutboundRecord.approval_datetime >= start_dt,
                                OutboundRecord.approval_datetime < end_dt,
                            ),
                            OutboundRecordDetail.outbound_amount,
                        ),
                        else_=0,
                    )
                ),
                0,
            ).label("make_out"),
            func.coalesce(
                func.sum(
                    case(
                        (
                            and_(
                                OutboundRecord.approval_status == 1,
                                OutboundRecord.outbound_type == 4,
                                OutboundRecord.approval_datetime >= start_dt,
                                OutboundRecord.approval_datetime < end_dt,
                            ),
                            OutboundRecordDetail.outbound_amount,
                        ),
                        else_=0,
                    )
                ),
                0,
            ).label("return_out"),
        )
        .join(OutboundRecord, OutboundRecord.outbound_record_id == OutboundRecordDetail.outbound_record_id)
        .where(
            OutboundRecord.display == 1,
            # pending 需要覆盖所有历史发生记录；approved delta 通过 approval_datetime 限制窗口
            OutboundRecord.outbound_datetime < end_dt,
        )
        .group_by(OutboundRecordDetail.material_storage_id)
    ).subquery("outbound_agg")

    delta_sum = (
        select(
            msids_all.c.msid.label("msid"),
            # 采购入库 - (材料退回出库)
            (func.coalesce(inbound_agg.c.purchase_in, 0) - func.coalesce(outbound_agg.c.return_out, 0)).label("delta_in"),
            func.coalesce(outbound_agg.c.prod_out, 0).label("delta_out"),
            # pending_* 直接用 as-of 绝对值，不参与 base + delta
            func.coalesce(inbound_agg.c.pending_in_asof, 0).label("pending_in_asof"),
            func.coalesce(outbound_agg.c.pending_out_asof, 0).label("pending_out_asof"),
            func.coalesce(inbound_agg.c.make_in, 0).label("delta_make_inbound"),
            func.coalesce(outbound_agg.c.make_out, 0).label("delta_make_outbound"),
            (
                (func.coalesce(inbound_agg.c.purchase_in, 0) - func.coalesce(outbound_agg.c.return_out, 0))
                - func.coalesce(outbound_agg.c.prod_out, 0)
                + func.coalesce(inbound_agg.c.make_in, 0)
                - func.coalesce(outbound_agg.c.make_out, 0)
            ).label("delta_net"),
        )
        .join(inbound_agg, inbound_agg.c.msid == msids_all.c.msid, isouter=True)
        .join(outbound_agg, outbound_agg.c.msid == msids_all.c.msid, isouter=True)
    ).subquery("delta_sum")

    final_agg = (
        select(
            msids_all.c.msid,
            func.coalesce(delta_sum.c.pending_in_asof, 0).label("final_pending_in"),
            func.coalesce(delta_sum.c.pending_out_asof, 0).label("final_pending_out"),
            (func.coalesce(base_snap.c.base_inbound, 0) + func.coalesce(delta_sum.c.delta_in, 0)).label("final_inbound"),
            (func.coalesce(base_snap.c.base_outbound, 0) + func.coalesce(delta_sum.c.delta_out, 0)).label("final_outbound"),
            (func.coalesce(base_snap.c.base_current, 0) + func.coalesce(delta_sum.c.delta_net, 0)).label("final_current"),
            (func.coalesce(base_snap.c.base_make_inbound, 0) + func.coalesce(delta_sum.c.delta_make_inbound, 0)).label("final_make_inbound"),
            (func.coalesce(base_snap.c.base_make_outbound, 0) + func.coalesce(delta_sum.c.delta_make_outbound, 0)).label("final_make_outbound"),
        )
        .join(base_snap, base_snap.c.msid == msids_all.c.msid, isouter=True)
        .join(delta_sum, delta_sum.c.msid == msids_all.c.msid, isouter=True)
    ).subquery("final_agg")

    return final_agg


# -----------------------------
# Lightweight inventory (for grouped / period)
# -----------------------------

def _compute_inventory_lightweight(target_date: date, filters: dict) -> List[Dict]:
    """
    Compute inventory quantities + materialName + averagePrice for ALL matching items.
    Skips full decorations (warehouse, supplier, order, shoe, etc.) and unit_price
    for significantly better performance. Used by grouped and period computations.
    """
    base_snapshot_date = _get_base_snapshot_date(target_date)
    if not base_snapshot_date:
        return []

    start_change_date = base_snapshot_date + timedelta(days=1)
    msids_all = _union_msids_subq(base_snapshot_date, target_date, filters)
    final_agg = _final_agg_for_msids_subq(msids_all, base_snapshot_date, start_change_date, target_date)

    all_rows = db.session.execute(
        select(final_agg).order_by(final_agg.c.msid.asc())
    ).all()
    if not all_rows:
        return []

    msid_list = [r.msid for r in all_rows]

    # Lightweight materialName: only 3 joins (vs 7+ in full decoration)
    name_stmt = (
        select(
            MaterialStorageSnapshot.material_storage_id.label("msid"),
            Material.material_name,
        )
        .join(SPUMaterial, SPUMaterial.spu_material_id == MaterialStorageSnapshot.spu_material_id)
        .join(Material, Material.material_id == SPUMaterial.material_id)
        .where(
            MaterialStorageSnapshot.snapshot_date == base_snapshot_date,
            MaterialStorageSnapshot.material_storage_id.in_(msid_list),
        )
    )
    name_map = {r.msid: r.material_name for r in db.session.execute(name_stmt).all()}

    missing = [m for m in msid_list if m not in name_map]
    if missing:
        fallback_stmt = (
            select(
                MaterialStorage.material_storage_id.label("msid"),
                Material.material_name,
            )
            .join(SPUMaterial, SPUMaterial.spu_material_id == MaterialStorage.spu_material_id)
            .join(Material, Material.material_id == SPUMaterial.material_id)
            .where(MaterialStorage.material_storage_id.in_(missing))
        )
        for r in db.session.execute(fallback_stmt).all():
            name_map[r.msid] = r.material_name

    # Average price
    price_end_dt = datetime.combine(target_date + timedelta(days=1), dt_time.min)
    avg_price_stmt = (
        select(
            InboundRecordDetail.material_storage_id.label("msid"),
            func.coalesce(
                func.sum(InboundRecordDetail.inbound_amount * InboundRecordDetail.unit_price)
                / func.nullif(func.sum(InboundRecordDetail.inbound_amount), 0),
                0,
            ).label("average_price"),
        )
        .join(InboundRecord, InboundRecord.inbound_record_id == InboundRecordDetail.inbound_record_id)
        .where(
            InboundRecord.display == 1,
            InboundRecord.approval_status == 1,
            InboundRecord.inbound_type.in_([0, 3]),
            InboundRecord.inbound_datetime < price_end_dt,
            InboundRecordDetail.material_storage_id.in_(msid_list),
        )
        .group_by(InboundRecordDetail.material_storage_id)
    )
    price_map = {r.msid: r.average_price for r in db.session.execute(avg_price_stmt).all()}

    _zero = Decimal("0")
    items = []
    for r in all_rows:
        avg_price = price_map.get(r.msid, _zero) or _zero
        items.append({
            "materialStorageId": r.msid,
            "materialName": name_map.get(r.msid, "未知材料"),
            "pendingInbound": r.final_pending_in or _zero,
            "pendingOutbound": r.final_pending_out or _zero,
            "inboundAmount": r.final_inbound or _zero,
            "outboundAmount": r.final_outbound or _zero,
            "currentAmount": r.final_current or _zero,
            "makeInventoryInbound": r.final_make_inbound or _zero,
            "makeInventoryOutbound": r.final_make_outbound or _zero,
            "averagePrice": round(avg_price, 3),
        })

    return items


def _fetch_decoration_for_msids(msids: List[int], base_snapshot_date: date) -> Dict[int, Dict]:
    """Fetch decoration info (warehouse, supplier, type, model, spec, color) for a small set of msids."""
    if not msids or not base_snapshot_date:
        return {}

    base_stmt = (
        select(
            MaterialStorageSnapshot.material_storage_id.label("msid"),
            MaterialWarehouse.material_warehouse_name.label("warehouseName"),
            Supplier.supplier_name.label("supplierName"),
            MaterialType.material_type_name.label("materialType"),
            Material.material_name.label("materialName"),
            SPUMaterial.material_model.label("materialModel"),
            SPUMaterial.material_specification.label("materialSpecification"),
            SPUMaterial.color.label("materialColor"),
        )
        .join(SPUMaterial, SPUMaterial.spu_material_id == MaterialStorageSnapshot.spu_material_id)
        .join(Material, Material.material_id == SPUMaterial.material_id)
        .join(Supplier, Supplier.supplier_id == Material.material_supplier)
        .join(MaterialType, MaterialType.material_type_id == Material.material_type_id)
        .join(MaterialWarehouse, MaterialWarehouse.material_warehouse_id == MaterialType.warehouse_id)
        .where(
            MaterialStorageSnapshot.snapshot_date == base_snapshot_date,
            MaterialStorageSnapshot.material_storage_id.in_(msids),
        )
    )
    result = {}
    for r in db.session.execute(base_stmt).all():
        result[r.msid] = {
            "warehouseName": r.warehouseName,
            "supplierName": r.supplierName,
            "materialType": r.materialType,
            "materialName": r.materialName,
            "materialModel": r.materialModel,
            "materialSpecification": r.materialSpecification,
            "materialColor": r.materialColor,
        }

    missing = [m for m in msids if m not in result]
    if missing:
        storage_stmt = (
            select(
                MaterialStorage.material_storage_id.label("msid"),
                MaterialWarehouse.material_warehouse_name.label("warehouseName"),
                Supplier.supplier_name.label("supplierName"),
                MaterialType.material_type_name.label("materialType"),
                Material.material_name.label("materialName"),
                SPUMaterial.material_model.label("materialModel"),
                SPUMaterial.material_specification.label("materialSpecification"),
                SPUMaterial.color.label("materialColor"),
            )
            .join(SPUMaterial, SPUMaterial.spu_material_id == MaterialStorage.spu_material_id)
            .join(Material, Material.material_id == SPUMaterial.material_id)
            .join(Supplier, Supplier.supplier_id == Material.material_supplier)
            .join(MaterialType, MaterialType.material_type_id == Material.material_type_id)
            .join(MaterialWarehouse, MaterialWarehouse.material_warehouse_id == MaterialType.warehouse_id)
            .where(MaterialStorage.material_storage_id.in_(missing))
        )
        for r in db.session.execute(storage_stmt).all():
            result[r.msid] = {
                "warehouseName": r.warehouseName,
                "supplierName": r.supplierName,
                "materialType": r.materialType,
                "materialName": r.materialName,
                "materialModel": r.materialModel,
                "materialSpecification": r.materialSpecification,
                "materialColor": r.materialColor,
            }

    return result


# -----------------------------
# Core
# -----------------------------

def compute_inventory_as_of(target_date: date, filters: dict, paginate: bool) -> dict:
    """
    DB-side DISTINCT + aggregate + filter + ORDER BY + LIMIT/OFFSET pagination.
    """
    page = filters.get("page", 1, type=int)
    page_size = filters.get("pageSize", 20, type=int)
    quantity_filters_raw = filters.get("quantityFilters", "[]")
    quantity_filters = json.loads(quantity_filters_raw)

    base_snapshot_date = _get_base_snapshot_date(target_date)
    if not base_snapshot_date:
        abort(Response(json.dumps({"message": "没有该日期历史库存记录"}), 400))

    start_change_date = base_snapshot_date + timedelta(days=1)

    # 1) Candidate msids in DB
    msids_all = _union_msids_subq(base_snapshot_date, target_date, filters)

    # 2) Final aggregation for all msids
    final_agg = _final_agg_for_msids_subq(msids_all, base_snapshot_date, start_change_date, target_date)

    # 3) quantity filters
    cond = literal(True)
    if len(quantity_filters) > 0:
        quantity_filter_conditions = []
        for q_filter in quantity_filters:
            field = q_filter.get('field')
            op = q_filter.get('op')
            column = None
            if field == 'pending_inbound':
                column = final_agg.c.final_pending_in
            elif field == 'pending_outbound':
                column = final_agg.c.final_pending_out
            elif field == 'inbound_amount':
                column = final_agg.c.final_inbound
            elif field == 'outbound_amount':
                column = final_agg.c.final_outbound
            elif field == 'make_inventory_inbound':
                column = final_agg.c.final_make_inbound
            elif field == 'make_inventory_outbound':
                column = final_agg.c.final_make_outbound
            elif field == 'current_amount':
                column = final_agg.c.final_current
            else:
                continue  # 跳过未知字段
            if op == 'eq_zero':
                condition = column == 0
            elif op == 'neq_zero':
                condition = column != 0
            else:
                continue  # 跳过未知操作符
            quantity_filter_conditions.append(condition)
        cond = and_(*quantity_filter_conditions)

    # 4) Total count in DB
    total_stmt = select(func.count()).select_from(
        select(final_agg.c.msid).where(cond).subquery()
    )
    total = db.session.execute(total_stmt).scalar_one()

    # 5) ORDER & paginate in DB
    offset = (page - 1) * page_size if paginate else 0
    limit = page_size if paginate else None

    page_base = select(final_agg).where(cond).order_by(final_agg.c.msid.asc())
    if limit is not None:
        page_base = page_base.limit(limit).offset(offset)

    page_rows = db.session.execute(page_base).all()
    msids_page = [r.msid for r in page_rows]
    # Map of final metrics
    final_map: Dict[int, Dict[str, Decimal]] = {
        r.msid: {
            "final_pending_in": r.final_pending_in,
            "final_pending_out": r.final_pending_out,
            "final_inbound": r.final_inbound,
            "final_outbound": r.final_outbound,
            "final_current": r.final_current,
            "final_make_inbound": r.final_make_inbound,
            "final_make_outbound": r.final_make_outbound,
        }
        for r in page_rows
    }

    if total == 0 or not msids_page:
        return {"items": [], "total": total, "page": page, "pageSize": page_size}

    # 6) Base snapshot info for decorations
    base_stmt = (
        select(
            MaterialStorageSnapshot,
            MaterialWarehouse.material_warehouse_name,
            Supplier.supplier_name,
            MaterialType.material_type_name,
            Material.material_name,
            SPUMaterial.material_model,
            SPUMaterial.material_specification,
            SPUMaterial.color,
            Order.order_rid,
            OrderShoe.customer_product_name,
            Shoe.shoe_rid,
        )
        .join(SPUMaterial, SPUMaterial.spu_material_id == MaterialStorageSnapshot.spu_material_id)
        .join(Material, Material.material_id == SPUMaterial.material_id)
        .join(Supplier, Supplier.supplier_id == Material.material_supplier)
        .join(MaterialType, MaterialType.material_type_id == Material.material_type_id)
        .join(MaterialWarehouse, MaterialWarehouse.material_warehouse_id == MaterialType.warehouse_id)
        .outerjoin(Order, MaterialStorageSnapshot.order_id == Order.order_id)
        .outerjoin(OrderShoe, MaterialStorageSnapshot.order_shoe_id == OrderShoe.order_shoe_id)
        .outerjoin(Shoe, Shoe.shoe_id == OrderShoe.shoe_id)
        .where(
            MaterialStorageSnapshot.snapshot_date == base_snapshot_date,
            MaterialStorageSnapshot.material_storage_id.in_(msids_page),
        )
    )
    base_rows = db.session.execute(base_stmt).all()
    base_map: Dict[int, Dict[str, object]] = {}
    COLS = [c.name for c in MaterialStorageSnapshot.__table__.columns]
    for row in base_rows:
        snapshot = row.MaterialStorageSnapshot
        base_map[snapshot.material_storage_id] = {
            **{c: getattr(snapshot, c) for c in COLS},
            "material_warehouse_name": row.material_warehouse_name,
            "supplier_name": row.supplier_name,
            "material_type_name": row.material_type_name,
            "material_name": row.material_name,
            "material_model": row.material_model,
            "material_specification": row.material_specification,
            "material_color": row.color,
            "order_rid": row.order_rid,
            "customer_product_name": row.customer_product_name,
            "shoe_rid": row.shoe_rid,
        }

    # 7) Fallback: for msids without base snapshot row, fetch from MaterialStorage
    missing_msids = [m for m in msids_page if m not in base_map]
    if missing_msids:
        storage_stmt = (
            select(
                MaterialStorage,
                MaterialWarehouse.material_warehouse_name,
                Supplier.supplier_name,
                MaterialType.material_type_name,
                Material.material_name,
                SPUMaterial.material_model,
                SPUMaterial.material_specification,
                SPUMaterial.color,
                Order.order_rid,
                OrderShoe.customer_product_name,
                Shoe.shoe_rid,
            )
            .join(SPUMaterial, SPUMaterial.spu_material_id == MaterialStorage.spu_material_id)
            .join(Material, Material.material_id == SPUMaterial.material_id)
            .join(Supplier, Supplier.supplier_id == Material.material_supplier)
            .join(MaterialType, MaterialType.material_type_id == Material.material_type_id)
            .join(MaterialWarehouse, MaterialWarehouse.material_warehouse_id == MaterialType.warehouse_id)
            .outerjoin(Order, MaterialStorage.order_id == Order.order_id)
            .outerjoin(OrderShoe, MaterialStorage.order_shoe_id == OrderShoe.order_shoe_id)
            .outerjoin(Shoe, Shoe.shoe_id == OrderShoe.shoe_id)
            .where(MaterialStorage.material_storage_id.in_(missing_msids))
        )
        storage_rows = db.session.execute(storage_stmt).all()
        for row in storage_rows:
            storage = row.MaterialStorage
            base_map[storage.material_storage_id] = {
                "material_storage_id": storage.material_storage_id,
                "material_warehouse_name": row.material_warehouse_name,
                "supplier_name": row.supplier_name,
                "material_type_name": row.material_type_name,
                "material_name": row.material_name,
                "material_model": row.material_model,
                "material_specification": row.material_specification,
                "material_color": row.color,
                "order_rid": row.order_rid,
                "customer_product_name": row.customer_product_name,
                "shoe_rid": row.shoe_rid,
                "actual_inbound_unit": getattr(storage, "actual_inbound_unit", None),
            }

    # 8) Prices (real-time from approved purchase inbound records, latest <= target_date)
    price_end_dt = datetime.combine(target_date + timedelta(days=1), dt_time.min)

    latest_inbound_dt_subq = (
        select(
            InboundRecordDetail.material_storage_id.label("msid"),
            func.max(InboundRecord.inbound_datetime).label("latest_dt"),
        )
        .join(InboundRecord, InboundRecord.inbound_record_id == InboundRecordDetail.inbound_record_id)
        .where(
            InboundRecord.display == 1,
            InboundRecord.approval_status == 1,
            InboundRecord.inbound_type.in_([0, 3]),
            InboundRecord.inbound_datetime < price_end_dt,
            InboundRecordDetail.material_storage_id.in_(msids_page),
        )
        .group_by(InboundRecordDetail.material_storage_id)
        .subquery("latest_inbound_dt")
    )

    latest_price_stmt = (
        select(
            InboundRecordDetail.material_storage_id.label("msid"),
            func.coalesce(InboundRecordDetail.unit_price, 0).label("unit_price"),
        )
        .join(InboundRecord, InboundRecord.inbound_record_id == InboundRecordDetail.inbound_record_id)
        .join(
            latest_inbound_dt_subq,
            and_(
                InboundRecordDetail.material_storage_id == latest_inbound_dt_subq.c.msid,
                InboundRecord.inbound_datetime == latest_inbound_dt_subq.c.latest_dt,
            ),
        )
    )

    avg_price_stmt = (
        select(
            InboundRecordDetail.material_storage_id.label("msid"),
            func.coalesce(
                (
                    func.sum(InboundRecordDetail.inbound_amount * InboundRecordDetail.unit_price)
                    / func.nullif(func.sum(InboundRecordDetail.inbound_amount), 0)
                ),
                0,
            ).label("average_price"),
        )
        .join(InboundRecord, InboundRecord.inbound_record_id == InboundRecordDetail.inbound_record_id)
        .where(
            InboundRecord.display == 1,
            InboundRecord.approval_status == 1,
            InboundRecord.inbound_type.in_([0, 3]),
            InboundRecord.inbound_datetime < price_end_dt,
            InboundRecordDetail.material_storage_id.in_(msids_page),
        )
        .group_by(InboundRecordDetail.material_storage_id)
    )

    latest_price_rows = db.session.execute(latest_price_stmt).all()
    avg_price_rows = db.session.execute(avg_price_stmt).all()

    price_map: Dict[int, Dict[str, Decimal]] = {msid: {"unit_price": Decimal("0"), "average_price": Decimal("0")} for msid in msids_page}
    for r in latest_price_rows:
        price_map[r.msid]["unit_price"] = r.unit_price
    for r in avg_price_rows:
        price_map[r.msid]["average_price"] = r.average_price
    # price_map already prefilled with zeros for all msids_page

    # 9) Assemble items
    items: List[Dict[str, object]] = []
    # print([id for id in msids_page])
    for msid in msids_page:
        base = base_map.get(msid, None)
        fin = final_map.get(msid, {})
        price_info = price_map.get(msid, {"unit_price": Decimal("0"), "average_price": Decimal("0")})

        if not base:
            # Again, extreme fallback to avoid KeyError
            base = {
                "material_storage_id": msid,
                "material_warehouse_name": None,
                "supplier_name": None,
                "material_type_name": None,
                "material_name": None,
                "material_model": None,
                "material_specification": None,
                "material_color": None,
                "order_rid": None,
                "customer_product_name": None,
                "shoe_rid": None,
                "actual_inbound_unit": None,
            }

        final_pending_in = fin.get("final_pending_in", Decimal("0"))
        final_pending_out = fin.get("final_pending_out", Decimal("0"))
        final_inbound = fin.get("final_inbound", Decimal("0"))
        final_outbound = fin.get("final_outbound", Decimal("0"))
        final_current = fin.get("final_current", Decimal("0"))
        final_make_inbound = fin.get("final_make_inbound", Decimal("0"))
        final_make_outbound = fin.get("final_make_outbound", Decimal("0"))

        items.append(
            {
                "materialStorageId": msid,
                "warehouseName": base.get("material_warehouse_name"),
                "supplierName": base.get("supplier_name"),
                "materialType": base.get("material_type_name"),
                "materialName": base.get("material_name"),
                "materialModel": base.get("material_model"),
                "materialSpecification": base.get("material_specification"),
                "materialColor": base.get("material_color"),
                "orderRId": base.get("order_rid"),
                "customerProductName": base.get("customer_product_name"),
                "shoeRId": base.get("shoe_rid"),
                "actualInboundUnit": base.get("actual_inbound_unit"),
                "unitPrice": round(price_info.get("unit_price", Decimal("0")), 3),
                "averagePrice": round(price_info.get("average_price", Decimal("0")), 3),
                "pendingInbound": final_pending_in,
                "pendingOutbound": final_pending_out,
                "inboundAmount": final_inbound,
                "outboundAmount": final_outbound,
                "currentAmount": final_current,
                "makeInventoryInbound": final_make_inbound,
                "makeInventoryOutbound": final_make_outbound,
                "inboundedItemTotalPrice": round(final_inbound * price_info.get("average_price", Decimal("0")), 3),
                "currentItemTotalPrice": round(final_current * price_info.get("average_price", Decimal("0")), 3),
            }
        )

    return {"items": items, "total": total, "page": page, "pageSize": page_size}


def compute_inventory_grouped_by_name(target_date: date, filters: dict, paginate: bool) -> dict:
    """
    Lightweight grouped inventory: uses _compute_inventory_lightweight (skips full
    decorations) then groups by materialName with weighted-average pricing.
    """
    page = int(filters.get("page", 1))
    page_size = int(filters.get("pageSize", 20))

    all_items = _compute_inventory_lightweight(target_date, filters)

    _zero = Decimal("0")
    groups: Dict[str, Dict] = {}
    for item in all_items:
        name = item.get("materialName") or "未知材料"
        if name not in groups:
            groups[name] = {
                "materialName": name,
                "pendingInbound": _zero,
                "pendingOutbound": _zero,
                "inboundAmount": _zero,
                "outboundAmount": _zero,
                "currentAmount": _zero,
                "makeInventoryInbound": _zero,
                "makeInventoryOutbound": _zero,
                "currentItemTotalPrice": _zero,
                "_weighted_price_sum": _zero,
                "_total_inbound_for_avg": _zero,
            }
        g = groups[name]
        g["pendingInbound"] += item.get("pendingInbound") or _zero
        g["pendingOutbound"] += item.get("pendingOutbound") or _zero
        g["inboundAmount"] += item.get("inboundAmount") or _zero
        g["outboundAmount"] += item.get("outboundAmount") or _zero
        current = item.get("currentAmount") or _zero
        avg_price = item.get("averagePrice") or _zero
        g["currentAmount"] += current
        g["makeInventoryInbound"] += item.get("makeInventoryInbound") or _zero
        g["makeInventoryOutbound"] += item.get("makeInventoryOutbound") or _zero
        g["currentItemTotalPrice"] += current * avg_price

        inbound_amt = item.get("inboundAmount") or _zero
        g["_weighted_price_sum"] += inbound_amt * avg_price
        g["_total_inbound_for_avg"] += inbound_amt

    result_items: List[Dict] = []
    for g in groups.values():
        total_inbound = g.pop("_total_inbound_for_avg")
        weighted_sum = g.pop("_weighted_price_sum")
        g["averagePrice"] = round(
            (weighted_sum / total_inbound) if total_inbound else _zero, 3
        )
        g["currentItemTotalPrice"] = round(g["currentItemTotalPrice"], 3)
        result_items.append(g)

    result_items.sort(key=lambda x: x["materialName"])

    total = len(result_items)
    if paginate:
        offset = (page - 1) * page_size
        result_items = result_items[offset : offset + page_size]

    return {"items": result_items, "total": total, "page": page, "pageSize": page_size}


# -----------------------------
# Routes
# -----------------------------

@material_storage_snapshot_bp.route("/warehouse/getmaterialstroagebydate", methods=["GET"])
def get_material_storage_by_date():
    """
    GET /warehouse/getmaterialstroagebydate?snapshotDate=2025-10-04&...
    Return: inventory state list (one row per material_storage_id) for given date.
    """
    date_str = request.args.get("snapshotDate")
    if not date_str:
        return jsonify({"code": 400, "msg": "缺少参数 date(YYYY-MM-DD)"}), 400

    try:
        target_date = _parse_date(date_str)
    except Exception:
        return jsonify({"code": 400, "msg": "date 格式错误，需 YYYY-MM-DD"}), 400

    data = compute_inventory_as_of(target_date, filters=request.args, paginate=True)
    return jsonify({"code": 200, "msg": "success", "data": data})


@material_storage_snapshot_bp.route("/warehouse/exportinventoryhistory", methods=["GET"])
def export_inventory_history():
    date_str = request.args.get("snapshotDate")
    warehouse_id = request.args.get("warehouseIdFilter")

    warehouse_name = None
    if warehouse_id:
        warehouse_name = (
            db.session.query(MaterialWarehouse.material_warehouse_name)
            .filter(MaterialWarehouse.material_warehouse_id == warehouse_id)
            .scalar()
        )

    supplier_name_filter = request.args.get("supplierNameFilter")

    if not date_str:
        return jsonify({"code": 400, "msg": "缺少参数 date(YYYY-MM-DD)"}), 400

    try:
        target_date = _parse_date(date_str)
    except Exception:
        return jsonify({"code": 400, "msg": "date 格式错误，需 YYYY-MM-DD"}), 400

    data = compute_inventory_as_of(target_date, filters=request.args, paginate=False)

    template_path = os.path.join(FILE_STORAGE_PATH, "财务历史库存模板.xlsx")
    timestamp = str(time.time())
    new_file_name = f"财务历史库存输出_{timestamp}.xlsx"
    save_path = os.path.join(FILE_STORAGE_PATH, "财务部文件", "库存总单", new_file_name)
    time_range_string = date_str

    generate_accounting_warehouse_excel(
        template_path,
        save_path,
        warehouse_name,
        supplier_name_filter,
        time_range_string,
        data["items"],
    )
    return send_file(save_path, as_attachment=True, download_name=new_file_name)


@material_storage_snapshot_bp.route("/warehouse/getmaterialstoragebydate/grouped", methods=["GET"])
def get_material_storage_by_date_grouped():
    """
    GET /warehouse/getmaterialstoragebydate/grouped?snapshotDate=2025-10-04&...
    Return: inventory grouped by materialName with weighted average price.
    """
    date_str = request.args.get("snapshotDate")
    if not date_str:
        return jsonify({"code": 400, "msg": "缺少参数 date(YYYY-MM-DD)"}), 400

    try:
        target_date = _parse_date(date_str)
    except Exception:
        return jsonify({"code": 400, "msg": "date 格式错误，需 YYYY-MM-DD"}), 400

    data = compute_inventory_grouped_by_name(target_date, filters=request.args, paginate=True)
    return jsonify({"code": 200, "msg": "success", "data": data})


@material_storage_snapshot_bp.route("/warehouse/exportinventoryhistory/grouped", methods=["GET"])
def export_inventory_history_grouped():
    date_str = request.args.get("snapshotDate")
    warehouse_id = request.args.get("warehouseIdFilter")

    warehouse_name = None
    if warehouse_id:
        warehouse_name = (
            db.session.query(MaterialWarehouse.material_warehouse_name)
            .filter(MaterialWarehouse.material_warehouse_id == warehouse_id)
            .scalar()
        )

    supplier_name_filter = request.args.get("supplierNameFilter")

    if not date_str:
        return jsonify({"code": 400, "msg": "缺少参数 date(YYYY-MM-DD)"}), 400

    try:
        target_date = _parse_date(date_str)
    except Exception:
        return jsonify({"code": 400, "msg": "date 格式错误，需 YYYY-MM-DD"}), 400

    data = compute_inventory_grouped_by_name(target_date, filters=request.args, paginate=False)

    timestamp = str(time.time())
    new_file_name = f"财务历史库存名称汇总_{timestamp}.xlsx"
    save_path = os.path.join(FILE_STORAGE_PATH, "财务部文件", "库存总单", new_file_name)
    time_range_string = date_str

    generate_accounting_warehouse_grouped_excel(
        save_path,
        warehouse_name,
        supplier_name_filter,
        time_range_string,
        data["items"],
    )
    return send_file(save_path, as_attachment=True, download_name=new_file_name)


# ——————————————————————————————————————————————
# 区间变化（期初 - 变化 - 期末）
# ——————————————————————————————————————————————

def _compute_period_diffs(start_date: date, end_date: date, filters: dict) -> List[Dict]:
    """
    Compute period diffs for ALL items using lightweight queries (no full decorations).
    Returns list of {msid, materialName, openingAmount, period changes, closingAmount, averagePrice, closingTotalPrice}.
    """
    opening_date = start_date - timedelta(days=1)
    opening_items = _compute_inventory_lightweight(opening_date, filters)
    closing_items = _compute_inventory_lightweight(end_date, filters)

    opening_map: Dict[int, Dict] = {
        item["materialStorageId"]: item for item in opening_items
    }
    closing_map: Dict[int, Dict] = {
        item["materialStorageId"]: item for item in closing_items
    }
    all_msids = sorted(set(opening_map.keys()) | set(closing_map.keys()))

    _zero = Decimal("0")
    items: List[Dict] = []
    for msid in all_msids:
        opening = opening_map.get(msid, {})
        closing = closing_map.get(msid, {})

        opening_current = opening.get("currentAmount", _zero) or _zero
        closing_current = closing.get("currentAmount", _zero) or _zero

        if opening_current == 0 and closing_current == 0:
            continue

        period_inbound = (closing.get("inboundAmount", _zero) or _zero) - (opening.get("inboundAmount", _zero) or _zero)
        period_outbound = (closing.get("outboundAmount", _zero) or _zero) - (opening.get("outboundAmount", _zero) or _zero)
        period_make_inbound = (closing.get("makeInventoryInbound", _zero) or _zero) - (opening.get("makeInventoryInbound", _zero) or _zero)
        period_make_outbound = (closing.get("makeInventoryOutbound", _zero) or _zero) - (opening.get("makeInventoryOutbound", _zero) or _zero)
        net_change = closing_current - opening_current

        ref = closing if closing else opening
        avg_price = ref.get("averagePrice", _zero) or _zero

        items.append({
            "materialStorageId": msid,
            "materialName": ref.get("materialName"),
            "openingAmount": opening_current,
            "periodInbound": period_inbound,
            "periodOutbound": period_outbound,
            "periodMakeInbound": period_make_inbound,
            "periodMakeOutbound": period_make_outbound,
            "netChange": net_change,
            "closingAmount": closing_current,
            "averagePrice": round(avg_price, 3),
            "closingTotalPrice": round(closing_current * avg_price, 3),
        })

    return items


def _compute_period_detail(start_date: date, end_date: date, filters: dict, paginate: bool) -> dict:
    """
    Period detail view. Uses lightweight queries, then fetches decorations only
    for the paginated page (or all items for export).
    """
    page = int(filters.get("page", 1))
    page_size = int(filters.get("pageSize", 20))

    all_diffs = _compute_period_diffs(start_date, end_date, filters)
    total = len(all_diffs)

    if paginate:
        offset = (page - 1) * page_size
        page_items = all_diffs[offset : offset + page_size]
    else:
        page_items = all_diffs

    # Fetch decorations only for the items we're returning
    msids = [item["materialStorageId"] for item in page_items]
    closing_base = _get_base_snapshot_date(end_date)
    decorations = _fetch_decoration_for_msids(msids, closing_base) if closing_base and msids else {}

    for item in page_items:
        dec = decorations.get(item["materialStorageId"], {})
        item["warehouseName"] = dec.get("warehouseName")
        item["supplierName"] = dec.get("supplierName")
        item["materialType"] = dec.get("materialType")
        item["materialModel"] = dec.get("materialModel")
        item["materialSpecification"] = dec.get("materialSpecification")
        item["materialColor"] = dec.get("materialColor")

    return {"items": page_items, "total": total, "page": page, "pageSize": page_size}


def _compute_period_grouped(start_date: date, end_date: date, filters: dict, paginate: bool) -> dict:
    """区间变化按材料名称汇总，价格使用加权平均。Uses lightweight period diffs."""
    page = int(filters.get("page", 1))
    page_size = int(filters.get("pageSize", 20))

    all_diffs = _compute_period_diffs(start_date, end_date, filters)

    groups: Dict[str, Dict] = {}
    _zero = Decimal("0")
    for item in all_diffs:
        name = item.get("materialName") or "未知材料"
        if name not in groups:
            groups[name] = {
                "materialName": name,
                "openingAmount": _zero,
                "periodInbound": _zero,
                "periodOutbound": _zero,
                "periodMakeInbound": _zero,
                "periodMakeOutbound": _zero,
                "netChange": _zero,
                "closingAmount": _zero,
                "closingTotalPrice": _zero,
                "_weighted_price_sum": _zero,
                "_total_closing_for_avg": _zero,
            }
        g = groups[name]
        g["openingAmount"] += item.get("openingAmount") or _zero
        g["periodInbound"] += item.get("periodInbound") or _zero
        g["periodOutbound"] += item.get("periodOutbound") or _zero
        g["periodMakeInbound"] += item.get("periodMakeInbound") or _zero
        g["periodMakeOutbound"] += item.get("periodMakeOutbound") or _zero
        g["netChange"] += item.get("netChange") or _zero
        g["closingAmount"] += item.get("closingAmount") or _zero
        g["closingTotalPrice"] += item.get("closingTotalPrice") or _zero

        closing_amt = item.get("closingAmount") or _zero
        avg_price = item.get("averagePrice") or _zero
        if closing_amt > 0:
            g["_weighted_price_sum"] += closing_amt * avg_price
            g["_total_closing_for_avg"] += closing_amt

    result_items: List[Dict] = []
    for g in groups.values():
        total_closing = g.pop("_total_closing_for_avg")
        weighted_sum = g.pop("_weighted_price_sum")
        g["averagePrice"] = round((weighted_sum / total_closing) if total_closing else _zero, 3)
        g["closingTotalPrice"] = round(g["closingTotalPrice"], 3)
        result_items.append(g)

    result_items.sort(key=lambda x: x["materialName"])

    total = len(result_items)
    if paginate:
        offset = (page - 1) * page_size
        result_items = result_items[offset : offset + page_size]
    return {"items": result_items, "total": total, "page": page, "pageSize": page_size}


def _parse_period_dates():
    start_str = request.args.get("startDate")
    end_str = request.args.get("endDate")
    if not start_str or not end_str:
        abort(Response(json.dumps({"message": "缺少 startDate / endDate 参数"}), 400))
    try:
        return _parse_date(start_str), _parse_date(end_str)
    except Exception:
        abort(Response(json.dumps({"message": "日期格式错误，需 YYYY-MM-DD"}), 400))


@material_storage_snapshot_bp.route("/warehouse/getmaterialstorageperiod", methods=["GET"])
def get_material_storage_period():
    start_date, end_date = _parse_period_dates()
    data = _compute_period_detail(start_date, end_date, filters=request.args, paginate=True)
    return jsonify({"code": 200, "msg": "success", "data": data})


@material_storage_snapshot_bp.route("/warehouse/getmaterialstorageperiod/grouped", methods=["GET"])
def get_material_storage_period_grouped():
    start_date, end_date = _parse_period_dates()
    data = _compute_period_grouped(start_date, end_date, filters=request.args, paginate=True)
    return jsonify({"code": 200, "msg": "success", "data": data})


@material_storage_snapshot_bp.route("/warehouse/exportmaterialstorageperiod", methods=["GET"])
def export_material_storage_period():
    start_date, end_date = _parse_period_dates()
    data = _compute_period_detail(start_date, end_date, filters=request.args, paginate=False)

    timestamp = str(time.time())
    new_file_name = f"财务历史库存区间明细_{timestamp}.xlsx"
    save_path = os.path.join(FILE_STORAGE_PATH, "财务部文件", "库存总单", new_file_name)

    _generate_period_excel(save_path, f"{start_date} ~ {end_date}", data["items"], grouped=False)
    return send_file(save_path, as_attachment=True, download_name=new_file_name)


@material_storage_snapshot_bp.route("/warehouse/exportmaterialstorageperiod/grouped", methods=["GET"])
def export_material_storage_period_grouped():
    start_date, end_date = _parse_period_dates()
    data = _compute_period_grouped(start_date, end_date, filters=request.args, paginate=False)

    timestamp = str(time.time())
    new_file_name = f"财务历史库存区间名称汇总_{timestamp}.xlsx"
    save_path = os.path.join(FILE_STORAGE_PATH, "财务部文件", "库存总单", new_file_name)

    _generate_period_excel(save_path, f"{start_date} ~ {end_date}", data["items"], grouped=True)
    return send_file(save_path, as_attachment=True, download_name=new_file_name)


def _generate_period_excel(save_path: str, time_range: str, items: list, grouped: bool):
    """Generate Excel for period (opening-change-closing) data."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    os.makedirs(os.path.dirname(save_path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "区间变化"

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11)
    ws["A1"] = "财务部历史库存 - 区间变化"
    ws["A1"].font = title_font
    ws["A2"] = "时间范围"
    ws["B2"] = time_range

    if grouped:
        headers = ["材料名称", "期初库存", "采购入库变化", "生产出库变化",
                    "盘库入库变化", "盘库出库变化", "净变化", "期末库存", "加权均价", "期末总金额"]
        keys = ["materialName", "openingAmount", "periodInbound", "periodOutbound",
                "periodMakeInbound", "periodMakeOutbound", "netChange", "closingAmount",
                "averagePrice", "closingTotalPrice"]
    else:
        headers = ["仓库", "供应商", "材料类型", "材料名称", "材料型号", "材料规格", "材料颜色",
                    "期初库存", "采购入库变化", "生产出库变化", "盘库入库变化", "盘库出库变化",
                    "净变化", "期末库存", "加权均价", "期末总金额"]
        keys = ["warehouseName", "supplierName", "materialType", "materialName",
                "materialModel", "materialSpecification", "materialColor",
                "openingAmount", "periodInbound", "periodOutbound",
                "periodMakeInbound", "periodMakeOutbound", "netChange",
                "closingAmount", "averagePrice", "closingTotalPrice"]

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font

    for row_idx, item in enumerate(items, start=5):
        for col_idx, key in enumerate(keys, start=1):
            val = item.get(key, "")
            if isinstance(val, Decimal):
                val = float(val)
            ws.cell(row=row_idx, column=col_idx, value=val)

    wb.save(save_path)
